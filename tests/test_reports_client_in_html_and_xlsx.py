from io import BytesIO
import openpyxl

def test_reports_shows_client_icon_in_html(client, db, create_user, create_project, create_task, login):
    admin = create_user(email='radmin@example.com', is_internal=True)
    client_user = create_user(email='rclient@example.com', is_internal=False, first_name='ReporteC')
    login(admin)
    p = create_project(name='P-report-client')

    from app.models import Project, db as _db
    proj = Project.query.get(p['id'])
    proj.clients.append(client_user)
    _db.session.commit()

    # Create task assigned to client
    rv = client.post('/task', data={'project_id': p['id'], 'title': 'RTask', 'assigned_client_id': str(client_user.id)}, follow_redirects=True)
    assert rv.status_code == 200

    rv2 = client.get(f"/reports?project_id={p['id']}")
    assert rv2.status_code == 200
    html = rv2.get_data(as_text=True)

    assert 'class="client-icon"' in html
    assert 'ReporteC (Cliente Externo)' in html or 'rclient@example.com (Cliente Externo)' in html


def test_reports_xlsx_includes_client_name_with_label(client, db, create_user, create_project, create_task, login):
    admin = create_user(email='rxadmin@example.com', is_internal=True)
    client_user = create_user(email='rxclient@example.com', is_internal=False, first_name='ReporteX')
    login(admin)
    p = create_project(name='P-report-client-x')

    from app.models import Project, db as _db
    proj = Project.query.get(p['id'])
    proj.clients.append(client_user)
    _db.session.commit()

    rv = client.post('/task', data={'project_id': p['id'], 'title': 'RExport', 'assigned_client_id': str(client_user.id)}, follow_redirects=True)
    assert rv.status_code == 200

    rv2 = client.get(f"/reports?project_id={p['id']}&export=xlsx")
    assert rv2.status_code == 200
    wb = openpyxl.load_workbook(filename=BytesIO(rv2.data))
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert 'Assigned Client' in headers or 'Assigned Client' in headers

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    found = False
    for r in rows:
        if r[1] == 'RExport':
            idx = headers.index('Assigned Client')
            assert r[idx] == 'ReporteX (Cliente Externo)' or r[idx] == 'rxclient@example.com (Cliente Externo)'
            found = True
            break
    assert found
