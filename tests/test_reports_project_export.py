from datetime import date, timedelta
from io import BytesIO
import openpyxl


def test_project_export_xlsx(client, db, create_user, create_project, create_task, login):
    u = create_user(email='rep@example.com', is_internal=True)
    login(u)
    p = create_project(name='ExportP')
    due = date.today() - timedelta(days=3)
    t = create_task(project_id=p['id'], title='ExportTask', due_date=due, assigned_client_id=None)

    rv = client.get(f"/reports?project_id={p['id']}&export=xlsx")
    assert rv.status_code == 200
    content_type = rv.headers.get('Content-Type', '')
    assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type
    cd = rv.headers.get('Content-Disposition', '')
    assert f'project_{p["id"]}_summary.xlsx' in cd or f'project_{p["id"]}_summary.xlsx' in cd

    # Verify XLSX content contains our new header and the correct overdue value
    wb = openpyxl.load_workbook(filename=BytesIO(rv.data))
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert 'Atraso (días)' in headers

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    found = False
    for r in rows:
        # Title is at index 1
        if r[1] == 'ExportTask':
            idx = headers.index('Atraso (días)')
            assert int(r[idx]) == 3
            found = True
            break
    assert found
