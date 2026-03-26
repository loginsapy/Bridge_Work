from datetime import datetime, timedelta
import os
import uuid
import mimetypes
import shutil
from flask import render_template, jsonify, request, redirect, url_for, flash, abort, current_app, send_from_directory, send_file
from io import BytesIO
import requests
from flask_login import current_user, login_required
from sqlalchemy import func
from werkzeug.utils import secure_filename

from . import main_bp
from .. import db
from ..models import Project, Task, TimeEntry, User, Role, TaskAttachment, AuditLog, SystemNotification, SystemSettings, HourlyRate, project_clients, ProjectTemplate, ProjectTemplateTask, ProjectRisk
from ..metrics import calculate_project_metrics
from ..services import NotificationService
from ..auth.decorators import internal_required


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config.get('ALLOWED_EXTENSIONS', set())


def get_unique_filename(task_id, filename):
    """Generate a unique filename using UUID to avoid race conditions."""
    safe_name = secure_filename(filename)
    if not safe_name:
        safe_name = 'file'

    _, ext = os.path.splitext(safe_name)
    # Use UUID to guarantee uniqueness without filesystem race conditions
    final_name = f"{uuid.uuid4().hex}{ext}"

    task_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], f'task_{task_id}')
    os.makedirs(task_folder, exist_ok=True)

    return final_name, task_folder


def notify_clients_task_completed(task, completed_by_user=None):
    """Notifica a los clientes asociados al proyecto cuando una tarea se completa"""
    project = task.project
    
    # Marcar la tarea como pendiente de aprobación
    if task.requires_approval and task.is_external_visible:
        task.approval_status = 'PENDING'
        
        # Usar el servicio de notificaciones para notificar a clientes
        NotificationService.notify_task_completed(
            task=task,
            completed_by_user=completed_by_user,
            notify_client=True,
            send_email=SystemSettings.get('notify_task_approved', True)
        )
    else:
        # Notificar al creador del proyecto aunque no requiera aprobación
        NotificationService.notify_task_completed(
            task=task,
            completed_by_user=completed_by_user,
            notify_client=False,
            send_email=SystemSettings.get('notify_task_completed', True)
        )


@main_bp.route('/')
@login_required
def dashboard():
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None
    
    # Usuario sin rol: mostrar dashboard vacío con mensaje
    if not user_role:
        return render_template(
            'dashboard.html',
            no_role=True,
            active_projects_count=0,
            hours_this_week=0,
            tasks_completed_count=0,
            avg_budget_usage=0,
            recent_projects=[],
            project_status_data={},
            budget_chart_labels=[],
            budget_chart_data_used=[],
            budget_chart_data_planned=[],
            recent_activity=[],
            projects_by_status={'PLANNING': [], 'ACTIVE': [], 'COMPLETED': [], 'ARCHIVED': []},
            active_team_members=[],
            now=datetime.now()
        )
    
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday())

    # Filtrar datos según rol
    if user_role in ['PMP', 'Admin']:
        # PMP/Admin ve todo
        active_projects_count = db.session.query(func.count(Project.id)).filter(Project.status == 'ACTIVE').scalar()
        tasks_completed_count = db.session.query(func.count(Task.id)).filter(
            Task.status == 'COMPLETED',
            Task.assigned_to_id == current_user.id
        ).scalar()
        recent_projects = Project.query.order_by(Project.start_date.desc()).limit(5).all()
        recent_activity = Task.query.filter_by(status='COMPLETED').order_by(Task.due_date.desc()).limit(5).all()
        projects_by_status = {
            'PLANNING': Project.query.filter_by(status='PLANNING').order_by(Project.start_date.desc()).all(),
            'ACTIVE': Project.query.filter_by(status='ACTIVE').order_by(Project.start_date.desc()).all(),
            'COMPLETED': Project.query.filter_by(status='COMPLETED').order_by(Project.end_date.desc()).all(),
            'ARCHIVED': Project.query.filter_by(status='ARCHIVED').order_by(Project.end_date.desc()).all()
        }
    elif user_role == 'Supervisor':
        # Supervisor ve proyectos donde es miembro
        supervisor_project_ids = db.session.query(Project.id).filter(
            Project.members.contains(current_user)
        ).scalar_subquery()

        active_projects_count = db.session.query(func.count(Project.id)).filter(
            Project.status == 'ACTIVE',
            Project.id.in_(supervisor_project_ids)
        ).scalar()
        tasks_completed_count = db.session.query(func.count(Task.id)).filter(
            Task.status == 'COMPLETED',
            Task.project_id.in_(supervisor_project_ids)
        ).scalar()
        recent_projects = Project.query.filter(Project.id.in_(supervisor_project_ids)).order_by(Project.start_date.desc()).limit(5).all()
        recent_activity = Task.query.filter(
            Task.status == 'COMPLETED',
            Task.project_id.in_(supervisor_project_ids)
        ).order_by(Task.due_date.desc()).limit(5).all()
        projects_by_status = {
            'PLANNING': Project.query.filter(Project.status == 'PLANNING', Project.id.in_(supervisor_project_ids)).order_by(Project.start_date.desc()).all(),
            'ACTIVE': Project.query.filter(Project.status == 'ACTIVE', Project.id.in_(supervisor_project_ids)).order_by(Project.start_date.desc()).all(),
            'COMPLETED': Project.query.filter(Project.status == 'COMPLETED', Project.id.in_(supervisor_project_ids)).order_by(Project.end_date.desc()).all(),
            'ARCHIVED': Project.query.filter(Project.status == 'ARCHIVED', Project.id.in_(supervisor_project_ids)).order_by(Project.end_date.desc()).all()
        }
    elif user_role == 'Participante':
        # Participante ve solo proyectos donde tiene tareas asignadas (incluyendo multi-asignados)
        user_project_ids = db.session.query(Task.project_id).filter(
            (Task.assigned_to_id == current_user.id) | (Task.assignees.any(User.id == current_user.id))
        ).distinct().scalar_subquery()
        
        active_projects_count = db.session.query(func.count(Project.id)).filter(
            Project.status == 'ACTIVE',
            Project.id.in_(user_project_ids)
        ).scalar()
        tasks_completed_count = db.session.query(func.count(Task.id)).filter(
            Task.status == 'COMPLETED',
            (Task.assigned_to_id == current_user.id) | (Task.assignees.any(User.id == current_user.id))
        ).scalar()
        recent_projects = Project.query.filter(Project.id.in_(user_project_ids)).order_by(Project.start_date.desc()).limit(5).all()
        recent_activity = Task.query.filter(
            Task.status == 'COMPLETED',
            (Task.assigned_to_id == current_user.id) | (Task.assignees.any(User.id == current_user.id))
        ).order_by(Task.due_date.desc()).limit(5).all()
        projects_by_status = {
            'PLANNING': Project.query.filter(Project.status == 'PLANNING', Project.id.in_(user_project_ids)).order_by(Project.start_date.desc()).all(),
            'ACTIVE': Project.query.filter(Project.status == 'ACTIVE', Project.id.in_(user_project_ids)).order_by(Project.start_date.desc()).all(),
            'COMPLETED': Project.query.filter(Project.status == 'COMPLETED', Project.id.in_(user_project_ids)).order_by(Project.end_date.desc()).all(),
            'ARCHIVED': Project.query.filter(Project.status == 'ARCHIVED', Project.id.in_(user_project_ids)).order_by(Project.end_date.desc()).all()
        }
    else:
        # Cliente u otros roles: proyectos donde es cliente
        client_project_ids = db.session.query(Project.id).filter(
            Project.clients.contains(current_user)
        ).scalar_subquery()
        
        active_projects_count = db.session.query(func.count(Project.id)).filter(
            Project.status == 'ACTIVE',
            Project.id.in_(client_project_ids)
        ).scalar()
        tasks_completed_count = 0
        recent_projects = Project.query.filter(Project.id.in_(client_project_ids)).order_by(Project.start_date.desc()).limit(5).all()
        recent_activity = []
        projects_by_status = {
            'PLANNING': Project.query.filter(Project.status == 'PLANNING', Project.id.in_(client_project_ids)).order_by(Project.start_date.desc()).all(),
            'ACTIVE': Project.query.filter(Project.status == 'ACTIVE', Project.id.in_(client_project_ids)).order_by(Project.start_date.desc()).all(),
            'COMPLETED': Project.query.filter(Project.status == 'COMPLETED', Project.id.in_(client_project_ids)).order_by(Project.end_date.desc()).all(),
            'ARCHIVED': Project.query.filter(Project.status == 'ARCHIVED', Project.id.in_(client_project_ids)).order_by(Project.end_date.desc()).all()
        }

    # KPI data común
    hours_this_week = db.session.query(func.sum(TimeEntry.hours)).filter(
        TimeEntry.user_id == current_user.id,
        TimeEntry.date >= start_of_week
    ).scalar() or 0

    # Budget usage calculation
    projects_with_budget = Project.query.filter(Project.budget_hours.isnot(None)).all()
    total_projects_with_budget = len(projects_with_budget)
    total_budget_usage_percent = 0

    for p in projects_with_budget:
        if p.budget_hours > 0:
            total_hours_spent = db.session.query(func.sum(TimeEntry.hours)).join(Task).filter(Task.project_id == p.id).scalar() or 0
            usage_percent = (total_hours_spent / p.budget_hours) * 100
            total_budget_usage_percent += usage_percent

    avg_budget_usage = (total_budget_usage_percent / total_projects_with_budget) if total_projects_with_budget > 0 else 0

    # Calcular progreso de proyectos recientes (tareas completadas / total tareas)
    for p in recent_projects:
        total_tasks = db.session.query(func.count(Task.id)).filter(Task.project_id == p.id).scalar() or 0
        completed_tasks = db.session.query(func.count(Task.id)).filter(
            Task.project_id == p.id,
            Task.status.in_(['COMPLETED', 'DONE', 'ACCEPTED'])
        ).scalar() or 0
        p.progress = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0

    # Project status distribution
    project_status_counts = db.session.query(Project.status, func.count(Project.id)).group_by(Project.status).all()
    project_status_data = {status: count for status, count in project_status_counts}

    # Budget chart data (last 7 days)
    budget_chart_labels = []
    budget_chart_data_used = []
    budget_chart_data_planned = [] # Mocked planned data
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        budget_chart_labels.append(day.strftime('%a'))
        daily_hours = db.session.query(func.sum(TimeEntry.hours)).filter(TimeEntry.date == day).scalar() or 0
        budget_chart_data_used.append(float(daily_hours))
        # Simple linear progression for planned budget
        budget_chart_data_planned.append(float(sum(budget_chart_data_used) * 0.95 / (7-i)))

    # Usuarios internos con actividad reciente (time entries en las últimas 2 semanas)
    two_weeks_ago = today - timedelta(days=14)
    active_team_members = db.session.query(User).join(TimeEntry, TimeEntry.user_id == User.id).filter(
        User.is_internal == True,
        TimeEntry.date >= two_weeks_ago
    ).distinct().limit(6).all()

    # Agregar estadísticas a cada miembro en una sola query agregada (evitar N+1)
    if active_team_members:
        member_ids = [m.id for m in active_team_members]
        member_stats = db.session.query(
            TimeEntry.user_id,
            func.count(func.distinct(TimeEntry.task_id)).label('tasks_count'),
            func.sum(TimeEntry.hours).label('hours_total')
        ).filter(
            TimeEntry.user_id.in_(member_ids),
            TimeEntry.date >= two_weeks_ago
        ).group_by(TimeEntry.user_id).all()

        stats_map = {row.user_id: row for row in member_stats}
        for member in active_team_members:
            row = stats_map.get(member.id)
            member.recent_tasks_count = int(row.tasks_count) if row else 0
            member.hours_logged = float(row.hours_total) if row and row.hours_total else 0


    return render_template(
        'dashboard.html',
        no_role=False,
        active_projects_count=active_projects_count,
        hours_this_week=hours_this_week,
        tasks_completed_count=tasks_completed_count,
        avg_budget_usage=avg_budget_usage,
        recent_projects=recent_projects,
        project_status_data=project_status_data,
        budget_chart_labels=budget_chart_labels,
        budget_chart_data_used=budget_chart_data_used,
        budget_chart_data_planned=budget_chart_data_planned,
        recent_activity=recent_activity,
        projects_by_status=projects_by_status,
        active_team_members=active_team_members,
        now=datetime.now()
    )


@main_bp.route('/profile')
@login_required
def profile():
    """User profile page"""
    # Get user's tasks statistics
    user_tasks = Task.query.filter(
        (Task.assigned_to_id == current_user.id) | (Task.assignees.any(User.id == current_user.id))
    )
    total_tasks = user_tasks.count()
    completed_tasks = user_tasks.filter(Task.status == 'COMPLETED').count()
    
    # Get user's time entries statistics
    total_hours = db.session.query(func.sum(TimeEntry.hours)).filter(
        TimeEntry.user_id == current_user.id
    ).scalar() or 0
    
    # Get user's projects count
    user_project_ids = db.session.query(Task.project_id).filter(
        (Task.assigned_to_id == current_user.id) | (Task.assignees.any(User.id == current_user.id))
    ).distinct().scalar_subquery()
    projects_count = Project.query.filter(Project.id.in_(user_project_ids)).count()
    
    # Build stats object for template
    stats = {
        'tasks_assigned': total_tasks,
        'tasks_completed': completed_tasks,
        'projects_managed': projects_count,
        'total_hours': total_hours
    }
    
    return render_template('profile.html', stats=stats)


# Mockup preview routes for design review
@main_bp.route('/mock/dashboard')
def mock_dashboard():
    return render_template('mockups/dashboard_mockup.html')


@main_bp.route('/mock/project')
def mock_project():
    return render_template('mockups/project_mockup.html')


@main_bp.route('/project/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_project(project_id):
    project = Project.query.get_or_404(project_id)
    
    # Solo PMP o Admin pueden editar proyectos
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None
    if user_role not in ['PMP', 'Admin']:
        flash('Solo usuarios PMP o Admin pueden editar proyectos.', 'danger')
        return redirect(url_for('main.project_detail', project_id=project.id))
    
    # Obtener usuarios con rol Cliente para el selector
    client_role = Role.query.filter_by(name='Cliente').first()
    available_clients = User.query.filter_by(role_id=client_role.id).order_by(User.first_name).all() if client_role else []

    # Obtener usuarios internos activos para selector de miembros (mantener miembros actuales aunque cambiara su rol)
    internal_members = User.query.filter_by(is_internal=True, is_active=True).order_by(User.first_name).all()
    member_map = {u.id: u for u in internal_members}
    for m in project.members:
        member_map.setdefault(m.id, m)
    available_members = sorted(member_map.values(), key=lambda u: (u.first_name or '', u.last_name or ''))
    
    if request.method == 'POST':
        try:
            # Capturar valores anteriores para auditoría
            old_client_ids = [c.id for c in project.clients]
            old_member_ids = [m.id for m in project.members]
            old_values = {
                'name': project.name,
                'description': project.description,
                'budget_hours': float(project.budget_hours) if project.budget_hours else None,
                'status': project.status,
                'manager_id': project.manager_id,
                'client_ids': old_client_ids,
                'member_ids': old_member_ids
            }
            
            # Aplicar cambios
            project.name = request.form.get('name') or project.name
            project.description = request.form.get('description')
            project.budget_hours = float(request.form.get('budget_hours')) if request.form.get('budget_hours') and request.form.get('budget_hours').strip() else project.budget_hours
            project.status = request.form.get('status', project.status)
            
            # Update dates
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            
            if start_date_str:
                project.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            if end_date_str:
                project.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            # Actualizar responsable del proyecto
            manager_id = request.form.get('manager_id')
            
            # Fix 8: Validación fechas en edición de proyecto
            if project.start_date and project.end_date and project.start_date > project.end_date:
                 flash('La fecha de fin no puede ser anterior a la de inicio.', 'warning')
                 return redirect(url_for('main.edit_project', project_id=project.id))

            project.manager_id = int(manager_id) if manager_id and manager_id.strip() else None
            
            # Actualizar clientes asociados
            client_ids = request.form.getlist('client_ids')
            new_clients = User.query.filter(User.id.in_(client_ids)).all() if client_ids else []
            project.clients = new_clients

            # Actualizar miembros (PMPs adicionales + Supervisores + Participantes)
            pmp_ids = [int(x) for x in request.form.getlist('pmp_ids') if x.strip()]
            supervisor_ids = [int(x) for x in request.form.getlist('supervisor_ids') if x.strip()]
            member_ids = [int(x) for x in request.form.getlist('member_ids') if x.strip()]
            all_member_ids = list(set(pmp_ids + supervisor_ids + member_ids))
            new_members = User.query.filter(User.id.in_(all_member_ids)).all() if all_member_ids else []
            project.members = new_members

            # Departamento (opcional)
            dept_id_raw = request.form.get('department_id', '').strip()
            project.department_id = int(dept_id_raw) if dept_id_raw else None
            
            # Registrar cambios en auditoría
            new_client_ids = [c.id for c in project.clients]
            new_member_ids = [m.id for m in project.members]
            new_values = {
                'name': project.name,
                'description': project.description,
                'budget_hours': float(project.budget_hours) if project.budget_hours else None,
                'status': project.status,
                'manager_id': project.manager_id,
                'client_ids': new_client_ids,
                'member_ids': new_member_ids
            }
            
            changes = {}
            for field, old_val in old_values.items():
                new_val = new_values[field]
                if old_val != new_val:
                    changes[field] = {'old': old_val, 'new': new_val}
            
            if changes:
                audit = AuditLog(
                    entity_type='Project',
                    entity_id=project.id,
                    action='UPDATE',
                    user_id=current_user.id,
                    changes=changes
                )
                db.session.add(audit)
            
            db.session.commit()
            flash(f"Proyecto '{project.name}' actualizado.", 'success')

            # Webhook: project.status_changed / project.completed
            try:
                from app.services import webhook_service
                if 'status' in changes:
                    event = 'project.completed' if project.status == 'COMPLETED' else 'project.status_changed'
                    webhook_service.dispatch(event, {
                        'project_id': project.id,
                        'project_name': project.name,
                        'user_name': current_user.name or current_user.email,
                        'old_status': changes['status']['old'],
                        'new_status': project.status,
                    })
            except Exception:
                current_app.logger.exception('Failed to dispatch project status webhook')

            return redirect(url_for('main.project_detail', project_id=project.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar: {str(e)}', 'danger')
    
    # Show all internal active users as available members, but include existing project members
    # so we don't lose members whose role changed or who became inactive.
    internal_members = User.query.filter_by(is_internal=True, is_active=True).order_by(User.first_name).all()
    member_map = {u.id: u for u in internal_members}
    for m in project.members:
        member_map.setdefault(m.id, m)
    all_members = sorted(member_map.values(), key=lambda u: (u.first_name or '', u.last_name or ''))

    # Group members by role for clear display in the template
    MEMBER_ROLE_GROUPS = ['PMP', 'Supervisor', 'Participante']
    members_by_role = {r: [] for r in MEMBER_ROLE_GROUPS}
    members_other = []
    for u in all_members:
        role_name = u.role.name if u.role else None
        if role_name in members_by_role:
            members_by_role[role_name].append(u)
        else:
            members_other.append(u)

    # IDs of currently assigned members (for checkbox checked state)
    project_member_ids = [m.id for m in project.members]

    # Responsable del proyecto: solo usuarios con rol PMP
    _pmp_role = Role.query.filter_by(name='PMP').first()
    available_managers = User.query.filter_by(is_internal=True, is_active=True, role_id=_pmp_role.id).order_by(User.first_name).all() if _pmp_role else []

    # Áreas
    from ..models import Department
    available_departments = Department.query.order_by(Department.name).all()

    return render_template('project_edit.html', project=project, available_clients=available_clients,
                           members_by_role=members_by_role,
                           members_other=members_other,
                           project_member_ids=project_member_ids,
                           available_managers=available_managers,
                           available_departments=available_departments)


@main_bp.route('/project/<int:project_id>/delete', methods=['POST'])
@login_required
def delete_project(project_id):
    project = Project.query.get_or_404(project_id)
    
    # Validar permisos - solo Admin y PMP pueden eliminar
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None
    if user_role not in ['PMP', 'Admin']:
        flash('Solo usuarios PMP o Admin pueden eliminar proyectos.', 'danger')
        return redirect(url_for('main.projects'))
    
    try:
        project_name = project.name
        project_id_backup = project.id
        
        # Registrar auditoría antes de eliminar
        audit = AuditLog(
            entity_type='Project',
            entity_id=project_id_backup,
            action='DELETE',
            user_id=current_user.id,
            changes={'name': project_name, 'status': project.status}
        )
        db.session.add(audit)
        
        # Cascade delete: remove tasks, their time entries and attachments, then delete the project
        tasks = Task.query.filter(Task.project_id == project.id).all()
        for t in tasks:
            try:
                # Delete time entries
                TimeEntry.query.filter(TimeEntry.task_id == t.id).delete(synchronize_session=False)

                # Delete attachments records and files
                atts = TaskAttachment.query.filter(TaskAttachment.task_id == t.id).all()
                for a in atts:
                    try:
                        task_folder = os.path.join(current_app.config.get('UPLOAD_FOLDER', ''), f'task_{t.id}')
                        if task_folder and os.path.exists(os.path.join(task_folder, a.stored_filename)):
                            os.remove(os.path.join(task_folder, a.stored_filename))
                    except Exception:
                        current_app.logger.exception('Failed to remove attachment file for %s', a.id)
                TaskAttachment.query.filter(TaskAttachment.task_id == t.id).delete(synchronize_session=False)

                # Audit task deletion
                audit_t = AuditLog(
                    entity_type='task',
                    entity_id=t.id,
                    action='DELETE',
                    user_id=current_user.id,
                    changes={'task_title': t.title}
                )
                db.session.add(audit_t)

                # Finally delete the task
                db.session.delete(t)
            except Exception:
                current_app.logger.exception('Error while cascading delete for task %s', t.id)

        db.session.delete(project)
        db.session.commit()
        flash(f"Proyecto '{project_name}' y sus tareas asociadas fueron eliminados.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}', 'danger')
    
    return redirect(url_for('main.projects'))


# ---------------------------------------------------------------------------
# PROJECT TEMPLATES
# ---------------------------------------------------------------------------

@main_bp.route('/project-templates')
@login_required
def project_templates():
    """List all project templates."""
    templates = ProjectTemplate.query.order_by(ProjectTemplate.created_at.desc()).all()
    return render_template('project_templates.html', templates=templates)


@main_bp.route('/project/<int:project_id>/save-as-template', methods=['POST'])
@login_required
def save_project_as_template(project_id):
    """Save an existing project (and its tasks) as a reusable template."""
    project = Project.query.get_or_404(project_id)
    user_role = current_user.role.name if current_user.role else None
    if user_role not in ['PMP', 'Admin']:
        flash('Solo usuarios PMP o Admin pueden crear plantillas.', 'danger')
        return redirect(url_for('main.project_detail', project_id=project_id))

    name = request.form.get('template_name', '').strip() or f'Plantilla: {project.name}'
    description = request.form.get('template_description', '').strip()

    try:
        template = ProjectTemplate(
            name=name,
            description=description or project.description,
            project_type=project.project_type,
            created_by_id=current_user.id,
        )
        db.session.add(template)
        db.session.flush()  # get template.id before bulk insert

        # Only copy top-level and child tasks (skip tasks with missing titles)
        tasks = Task.query.filter_by(project_id=project.id).order_by(Task.position).all()
        proj_start = project.start_date or datetime.now().date()

        for pos, task in enumerate(tasks):
            def _days(dt):
                if dt is None:
                    return None
                d = dt.date() if hasattr(dt, 'date') else dt
                return (d - proj_start).days

            tt = ProjectTemplateTask(
                template_id=template.id,
                source_task_id=task.id,
                parent_source_id=task.parent_task_id,
                title=task.title,
                description=task.description,
                priority=task.priority,
                estimated_hours=task.estimated_hours,
                relative_start_days=_days(task.start_date),
                relative_due_days=_days(task.due_date),
                position=pos,
                is_external_visible=task.is_external_visible,
                requires_approval=task.requires_approval,
            )
            db.session.add(tt)

        db.session.commit()
        flash(f"Plantilla '{template.name}' creada con {len(tasks)} tarea(s).", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al guardar la plantilla: {str(e)}', 'danger')

    return redirect(url_for('main.project_templates'))


@main_bp.route('/project-templates/<int:template_id>/create-project', methods=['POST'])
@login_required
def create_project_from_template(template_id):
    """Create a new project (with tasks) from a template."""
    template = ProjectTemplate.query.get_or_404(template_id)
    user_role = current_user.role.name if current_user.role else None
    if user_role not in ['PMP', 'Admin']:
        flash('Solo usuarios PMP o Admin pueden crear proyectos.', 'danger')
        return redirect(url_for('main.project_templates'))

    name = request.form.get('name', '').strip()
    if not name:
        flash('El nombre del proyecto es obligatorio.', 'danger')
        return redirect(url_for('main.project_templates'))

    description = request.form.get('description', '').strip()
    start_date_str = request.form.get('start_date', '').strip()
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else datetime.now().date()
    except ValueError:
        start_date = datetime.now().date()

    try:
        project = Project(
            name=name,
            description=description or template.description,
            project_type=template.project_type,
            manager_id=current_user.id,
            status='PLANNING',
            start_date=start_date,
        )
        db.session.add(project)
        db.session.flush()  # get project.id

        # Map source_task_id → new Task so we can re-link parents
        source_to_new = {}

        for tt in sorted(template.tasks, key=lambda t: t.position):
            def _abs_date(offset):
                if offset is None:
                    return None
                return datetime.combine(start_date + timedelta(days=offset), datetime.min.time())

            task = Task(
                project_id=project.id,
                title=tt.title,
                description=tt.description,
                priority=tt.priority,
                estimated_hours=tt.estimated_hours,
                start_date=_abs_date(tt.relative_start_days),
                due_date=_abs_date(tt.relative_due_days),
                status='BACKLOG',
                position=tt.position,
                is_external_visible=tt.is_external_visible,
                requires_approval=tt.requires_approval,
            )
            db.session.add(task)
            db.session.flush()
            if tt.source_task_id:
                source_to_new[tt.source_task_id] = task

        # Re-link parent relationships using the source→new map
        for tt in template.tasks:
            if tt.parent_source_id and tt.source_task_id:
                child = source_to_new.get(tt.source_task_id)
                parent = source_to_new.get(tt.parent_source_id)
                if child and parent:
                    child.parent_task_id = parent.id

        db.session.add(AuditLog(
            entity_type='Project',
            entity_id=project.id,
            action='CREATE',
            user_id=current_user.id,
            changes={'name': project.name, 'from_template': template.name},
        ))
        db.session.commit()
        flash(f"Proyecto '{project.name}' creado desde la plantilla '{template.name}'.", 'success')
        return redirect(url_for('main.project_detail', project_id=project.id))
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear el proyecto: {str(e)}', 'danger')
        return redirect(url_for('main.project_templates'))


@main_bp.route('/project-templates/<int:template_id>/delete', methods=['POST'])
@login_required
def delete_project_template(template_id):
    """Delete a project template."""
    template = ProjectTemplate.query.get_or_404(template_id)
    user_role = current_user.role.name if current_user.role else None
    if user_role not in ['PMP', 'Admin']:
        flash('Solo usuarios PMP o Admin pueden eliminar plantillas.', 'danger')
        return redirect(url_for('main.project_templates'))

    try:
        name = template.name
        db.session.delete(template)
        db.session.commit()
        flash(f"Plantilla '{name}' eliminada.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}', 'danger')

    return redirect(url_for('main.project_templates'))


@main_bp.route('/api/kpi/velocity')
def kpi_velocity():
    """Provides data for the velocity chart on the dashboard."""
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    
    # Días en español
    dias_semana = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    
    labels = []
    data = []
    
    for i in range(7):
        day = start_of_week + timedelta(days=i)
        labels.append(dias_semana[i])
        
        daily_hours = db.session.query(func.sum(TimeEntry.hours)).filter(
            TimeEntry.date == day
        ).scalar() or 0
        
        data.append(float(daily_hours))
        
    return jsonify({'labels': labels, 'data': data})


@main_bp.route('/client/dashboard')
@login_required
def client_dashboard():
    """Portal simplificado para usuarios externos (clientes)."""
    # Redirect internal users to main projects page
    if current_user.is_internal:
        return redirect(url_for('main.projects'))

    # Projects this client is associated with
    user_projects = current_user.associated_projects or []

    # Tasks where this client is the assigned client — sorted None-safe by due_date
    assigned_tasks = sorted(
        Task.query.filter_by(assigned_client_id=current_user.id).all(),
        key=lambda t: (t.due_date is None, t.due_date or datetime.min)
    )

    # Tasks requiring approval from this client
    pending_approval = [t for t in assigned_tasks
                        if t.requires_approval and t.approval_status == 'PENDING']

    # Overdue tasks (not completed)
    now_dt = datetime.now()
    overdue_tasks = [t for t in assigned_tasks
                     if t.due_date and (t.due_date.replace(tzinfo=None) if t.due_date.tzinfo else t.due_date) < now_dt and t.status != 'COMPLETED']

    # Project metrics (only non-internal tasks visible to clients)
    project_metrics = {}
    for proj in user_projects:
        tasks = Task.query.filter(
            Task.project_id == proj.id,
            Task.is_internal_only != True
        ).all()
        total = len(tasks)
        completed = sum(1 for t in tasks if t.status == 'COMPLETED')
        project_metrics[proj.id] = {
            'total': total,
            'completed': completed,
            'pct': int(completed / total * 100) if total > 0 else 0,
        }

    overdue_ids = {t.id for t in overdue_tasks}

    return render_template(
        'client_dashboard.html',
        projects=user_projects,
        assigned_tasks=assigned_tasks,
        pending_approval=pending_approval,
        overdue_tasks=overdue_tasks,
        overdue_ids=overdue_ids,
        project_metrics=project_metrics,
        now=now_dt,
    )


@main_bp.route('/projects')
@login_required
def projects():
    """
    Visibilidad de proyectos según rol:
    - Admin: puede ver todos los proyectos
    - PMP: solo proyectos que gestiona (manager_id) o donde es miembro
    - Supervisor: solo proyectos donde es miembro
    - Participante: solo proyectos donde tiene tareas asignadas o es miembro
    - Cliente: solo proyectos donde es cliente
    - Sin Rol: no puede ver nada
    """
    from ..auth.decorators import _get_user_from_session
    from ..models import Department
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None

    # Usuario sin rol: mostrar mensaje y lista vacía
    if not user_role:
        flash('No tienes un rol asignado. Contacta al administrador para obtener acceso.', 'warning')
        return render_template('projects.html', projects=[], no_role=True)

    # Paginación: página actual desde querystring
    page = request.args.get('page', 1, type=int)
    per_page = 10

    # Admin: puede ver todo
    if user_role == 'Admin':
        base_query = Project.query.order_by(Project.start_date.desc())
    # PMP: solo proyectos que él gestiona o donde fue invitado como miembro
    elif user_role == 'PMP':
        base_query = Project.query.filter(
            (Project.manager_id == current_user.id) | (Project.members.contains(current_user))
        ).order_by(Project.start_date.desc())
    # Supervisor: solo proyectos donde es miembro
    elif user_role == 'Supervisor':
        base_query = Project.query.filter(
            Project.members.contains(current_user)
        ).order_by(Project.start_date.desc())
    # Participante: solo proyectos donde tiene tareas asignadas
    elif user_role == 'Participante':
        project_ids = db.session.query(Task.project_id).filter(
            Task.assigned_to_id == current_user.id
        ).distinct().scalar_subquery()
        base_query = Project.query.filter(
            (Project.id.in_(project_ids)) | (Project.members.contains(current_user))
        ).order_by(Project.start_date.desc())
    # Cliente: solo proyectos donde es cliente
    elif user_role == 'Cliente' or not current_user.is_internal:
        base_query = Project.query.filter(Project.clients.contains(current_user)).order_by(Project.start_date.desc())
    else:
        base_query = Project.query.filter(False)

    # Filtro por área
    area_id = request.args.get('area_id', '', type=str).strip()
    if area_id == 'none':
        base_query = base_query.filter(Project.department_id == None)
    elif area_id:
        try:
            base_query = base_query.filter(Project.department_id == int(area_id))
        except ValueError:
            area_id = ''

    # Evitar duplicados y problemas con JSON en Postgres.
    # Estrategia: contar project ids distintos, obtener los ids paginados mediante GROUP BY,
    # y luego cargar los objetos Project por esos ids.
    from ..models import Project as _Project
    from sqlalchemy import desc
    import math
    from types import SimpleNamespace

    # Total de proyectos distintos para la paginación
    ids_subq = base_query.with_entities(_Project.id).group_by(_Project.id).subquery()
    total = db.session.query(func.count()).select_from(ids_subq).scalar() or 0

    # Obtener ids paginados (ordenados por fecha de inicio más reciente)
    offset = (page - 1) * per_page
    id_rows = base_query.with_entities(_Project.id, func.max(_Project.start_date).label('sdate'))
    id_rows = id_rows.group_by(_Project.id).order_by(desc('sdate')).limit(per_page).offset(offset).all()
    ids = [r[0] for r in id_rows]

    # Cargar los proyectos correspondientes (mantener orden por start_date desc)
    if ids:
        projects = Project.query.filter(Project.id.in_(ids)).order_by(Project.start_date.desc()).all()
    else:
        projects = []

    # Construir objeto de paginación mínimo para la plantilla
    pages = int(math.ceil(total / float(per_page))) if per_page else 0
    pagination = SimpleNamespace(
        page=page,
        pages=pages,
        has_prev=(page > 1),
        has_next=(page < pages),
        prev_num=(page - 1 if page > 1 else None),
        next_num=(page + 1 if page < pages else None),
        total=total
    )

    # Calcular progreso y horas para cada proyecto en tiempo real
    # En la grilla de proyectos, "avance" debe reflejar ejecución de tareas
    # (tareas completadas / tareas totales), no consumo de presupuesto.
    for p in projects:
        total_hours = db.session.query(func.sum(TimeEntry.hours)).join(Task).filter(Task.project_id == p.id).scalar() or 0
        p.hours_spent = total_hours
        # Conteo de tareas para mostrar en listado y calcular avance
        try:
            total_tasks = int(db.session.query(func.count(Task.id)).filter(Task.project_id == p.id).scalar() or 0)
            completed_tasks = int(
                db.session.query(func.count(Task.id)).filter(
                    Task.project_id == p.id,
                    Task.status == 'COMPLETED'
                ).scalar() or 0
            )
            p.task_count = total_tasks
            p.completed_task_count = completed_tasks
        except Exception:
            total_tasks = 0
            p.task_count = 0
            p.completed_task_count = 0

        p.progress = ((p.completed_task_count / total_tasks) * 100) if total_tasks > 0 else 0
            
    # Obtener usuarios cliente para el modal de creación
    client_role = Role.query.filter_by(name='Cliente').first()
    available_clients = User.query.filter_by(role_id=client_role.id).order_by(User.first_name).all() if client_role else []
    
    # Obtener usuarios internos para el modal de creación, agrupados por rol
    _all_members = User.query.filter_by(is_internal=True, is_active=True).order_by(User.first_name).all()
    MEMBER_ROLE_GROUPS = ['PMP', 'Supervisor', 'Participante']
    members_by_role = {r: [] for r in MEMBER_ROLE_GROUPS}
    members_other = []
    for _u in _all_members:
        _rn = _u.role.name if _u.role else None
        if _rn in members_by_role:
            members_by_role[_rn].append(_u)
        else:
            members_other.append(_u)

    # Áreas — respetar el toggle de admin settings para la vista lista
    # Mantener el mismo default que Admin Settings (habilitado por defecto)
    _dept_raw = SystemSettings.get('departments_enabled', True)
    departments_enabled = (_dept_raw.lower() not in ('false', '0', 'no')) if isinstance(_dept_raw, str) else bool(_dept_raw)

    from ..models import Department
    available_departments = Department.query.order_by(Department.name).all() if departments_enabled else []
    area_filter = next((d for d in available_departments if str(d.id) == area_id), None)

    # Agrupar por área solo si el feature está activo y no hay filtro puntual aplicado
    if departments_enabled and available_departments and not area_id:
        groups = []
        for dept in available_departments:
            dept_projects = [p for p in projects if p.department_id == dept.id]
            if dept_projects:
                groups.append({'dept': dept, 'projects': dept_projects})
        no_dept = [p for p in projects if not p.department_id]
        if no_dept:
            groups.append({'dept': None, 'projects': no_dept})
        project_groups = groups
    else:
        project_groups = None

    # Provide current time context for templates that compare dates
    return render_template('projects.html',
                           projects=projects,
                           no_role=False,
                           now=datetime.now(),
                           current_user_role_name=user_role,
                           available_clients=available_clients,
                           members_by_role=members_by_role,
                           members_other=members_other,
                           pagination=pagination,
                           per_page=per_page,
                           available_departments=available_departments,
                           area_id=area_id,
                           area_filter=area_filter,
                           project_groups=project_groups)


@main_bp.route('/project/<int:project_id>')
@login_required
def project_detail(project_id):
    project = Project.query.get(project_id)
    if not project:
        from app.models import AuditLog
        any_log = AuditLog.query.filter_by(entity_type='Project', entity_id=project_id).order_by(AuditLog.created_at.desc()).first()
        if not any_log:
            return render_template('item_status.html', entity_type='proyecto', entity_id=project_id, status='never', canonical_entity='Project'), 404
        deletion = AuditLog.query.filter_by(entity_type='Project', entity_id=project_id, action='DELETE').order_by(AuditLog.created_at.desc()).first()
        if deletion:
            deleted_by = deletion.user.name if deletion.user else (f'Usuario {deletion.user_id}' if deletion.user_id else None)
            deleted_at = deletion.created_at
            return render_template('item_status.html', entity_type='proyecto', entity_id=project_id, status='deleted', deleted_by=deleted_by, deleted_at=deleted_at, canonical_entity='Project'), 404
        return render_template('item_status.html', entity_type='proyecto', entity_id=project_id, status='unavailable', canonical_entity='Project'), 404
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None

    # Usuario sin rol: no puede ver nada
    if not user_role:
        flash('No tienes un rol asignado. Contacta al administrador para obtener acceso.', 'warning')
        return redirect(url_for('main.projects'))

    # Control de acceso según rol
    if user_role == 'Admin':
        # Admin puede ver cualquier proyecto
        pass
    elif user_role == 'PMP':
        # PMP solo puede ver proyectos que gestiona o donde es miembro
        if project.manager_id != current_user.id and current_user not in project.members:
            flash('No tienes permiso para ver este proyecto.', 'danger')
            return redirect(url_for('main.projects'))
    elif user_role == 'Supervisor':
        # Supervisor solo puede ver proyectos donde es miembro
        if current_user not in project.members:
            flash('No tienes permiso para ver este proyecto.', 'danger')
            return redirect(url_for('main.projects'))
    elif user_role == 'Participante':
        # Participante solo puede ver proyectos donde tiene tareas asignadas o es miembro
        has_tasks = Task.query.filter(Task.project_id == project_id).filter(
            (Task.assigned_to_id == current_user.id) | (Task.assignees.any(User.id == current_user.id))
        ).first()
        is_member = current_user in project.members
        if not has_tasks and not is_member:
            flash('No tienes permiso para ver este proyecto.', 'danger')
            return redirect(url_for('main.projects'))
    elif user_role == 'Cliente' or not (user.is_internal if user else True):
        # Cliente solo puede ver sus proyectos
        client_ids = [u.id for u in project.clients]
        if (user.id if user else None) not in client_ids:
            flash('No tienes permiso para ver este proyecto.', 'danger')
            return redirect(url_for('main.projects'))
    else:
        flash('No tienes permiso para ver este proyecto.', 'danger')
        return redirect(url_for('main.projects'))

    # Filtrar tareas según rol
    if user_role in ['PMP', 'Admin', 'Supervisor']:
        # PMP/Admin/Supervisor ve todas las tareas del proyecto
        tasks = Task.query.filter_by(project_id=project_id).order_by(Task.status, Task.priority.desc()).all()
    elif user_role == 'Cliente' or not (user.is_internal if user else True):
        # Cliente ve solo tareas explícitamente visibles: marcadas como externas o asignadas al cliente
        # Se excluyen siempre las marcadas como solo internas
        client_user_id = user.id if user else None
        from sqlalchemy import or_
        tasks = Task.query.filter(
            Task.project_id == project_id,
            Task.is_internal_only != True,
            or_(
                Task.is_external_visible == True,
                Task.assigned_client_id == client_user_id
            )
        ).order_by(Task.status, Task.priority.desc()).all()
    else:
        # Participantes que acceden al proyecto pueden ver todas las tareas del proyecto
        tasks = Task.query.filter_by(project_id=project_id).order_by(Task.status, Task.priority.desc()).all()

    # Sugeridos para asignación: usuarios internos
    if project.members:
        assignees = list(set(project.members + ([project.manager] if project.manager else [])))
        assignees.sort(key=lambda u: u.first_name or '')
    else:
        assignees = User.query.filter_by(is_internal=True, is_active=True).order_by(User.first_name).all()
    # Candidate predecessors: all tasks in this project (for parent selection and dependencies)
    candidate_predecessors = Task.query.filter(Task.project_id == project_id).order_by(Task.title).all()
    
    # Build nested task tree using parent_task_id (hierarchy) or predecessor relationships
    def build_task_tree(task_list):
        """Build a tree structure using parent_task_id (WBS hierarchy) or predecessor relationships.

        This creates a proper parent-child tree where:
        - Tasks with parent_task_id are children of that parent
        - Tasks without parent_task_id but with a single predecessor are displayed as
          children of that predecessor (predecessor-as-parent display model)
        Also assigns WBS numbers (1, 1.1, 1.2, 2, 2.1.1)
        """
        tasks_by_id = {t.id: t for t in task_list}
        children_map = {t.id: [] for t in task_list}
        root_ids = []

        for t in task_list:
            if t.parent_task_id and t.parent_task_id in tasks_by_id:
                # Has a valid explicit parent in the visible task list
                children_map[t.parent_task_id].append(t)
            else:
                # Check if this task has predecessors within this task list that can serve as display parent
                pred_ids = [p.id for p in (getattr(t, 'predecessors', None) or []) if p.id in tasks_by_id]
                if len(pred_ids) == 1:
                    # Single predecessor: display this task as a child of that predecessor
                    children_map[pred_ids[0]].append(t)
                else:
                    # No parent or predecessor - treat as root
                    root_ids.append(t.id)
        
        # Sort function: prefer position, then status, priority, title
        status_order = {'BACKLOG': 0, 'IN_PROGRESS': 1, 'IN_REVIEW': 2, 'COMPLETED': 3}
        priority_order = {'CRITICAL': 3, 'HIGH': 2, 'MEDIUM': 1, 'LOW': 0}
        def sort_key(x):
            # if a manual position is set, respect it
            if getattr(x, 'position', None) is not None:
                return (0, x.position)
            return (1, status_order.get(x.status, 99), -priority_order.get(x.priority, 0), x.title or '')
        
        # Build recursive nodes with WBS numbering
        def build_node(tid, depth=0, wbs_prefix='', index=1):
            t = tasks_by_id[tid]
            t.tree_depth = depth  # Attach depth for indentation
            
            # Assign WBS number (e.g., "1", "1.1", "1.2", "2", "2.1.1")
            if wbs_prefix:
                t.wbs_number = f"{wbs_prefix}.{index}"
            else:
                t.wbs_number = str(index)
            
            children = sorted(children_map[tid], key=sort_key)
            child_nodes = []
            for i, c in enumerate(children, 1):
                child_nodes.append(build_node(c.id, depth + 1, t.wbs_number, i))
            
            return {'task': t, 'children': child_nodes}
        
        roots = [tasks_by_id[rid] for rid in root_ids]
        roots_sorted = sorted(roots, key=sort_key)
        forest = []
        for i, r in enumerate(roots_sorted, 1):
            forest.append(build_node(r.id, 0, '', i))
        
        return forest

    tasks_tree = build_task_tree(tasks)
    
    # Calculate predecessor order info for each task
    for task in tasks:
        if task.predecessors:
            # Sort predecessors by their WBS number if available, otherwise by ID
            sorted_preds = sorted(task.predecessors, key=lambda p: (getattr(p, 'wbs_number', '999'), p.id))
            task.predecessor_order = [(i+1, p) for i, p in enumerate(sorted_preds)]
        else:
            task.predecessor_order = []

    # Calcular métricas del proyecto
    metrics = calculate_project_metrics(project_id)

    # Participantes disponibles para que el Supervisor invite (todos los internos con rol Participante no miembros aún)
    from ..models import Role as _Role
    _part_role = _Role.query.filter_by(name='Participante').first()
    available_participantes = []
    if user_role == 'Supervisor' and _part_role:
        current_member_ids = {m.id for m in project.members}
        available_participantes = User.query.filter_by(
            is_internal=True, is_active=True, role_id=_part_role.id
        ).order_by(User.first_name).all()
        available_participantes = [u for u in available_participantes if u.id not in current_member_ids]

    # Render Monday-style board view
    return render_template('board.html', project=project, tasks=tasks, tasks_tree=tasks_tree, metrics=metrics,
                          users=assignees, candidate_predecessors=candidate_predecessors, now=datetime.now(),
                          project_id=project.id, user_role=user_role,
                          available_participantes=available_participantes)


@main_bp.route('/project/<int:project_id>/tasks/reorder', methods=['POST'])
@login_required
def reorder_project_tasks(project_id):
    """Persist order of tasks in a project based on the provided array of task IDs.
    Body: { "ordered_task_ids": [id1, id2, ...] }
    This sets the `position` field for the tasks according to the list order.
    """
    # Only internal users with role 'PMP' or 'Admin' can reorder tasks
    if not current_user.is_internal:
        return jsonify({'error': 'Solo usuarios internos pueden reordenar tareas', 'permission_denied': True}), 403
    if not getattr(current_user, 'role', None) or current_user.role.name not in ['PMP', 'Admin']:
        return jsonify({'error': 'Solo usuarios con rol PMP o Admin pueden reordenar tareas', 'permission_denied': True}), 403

    data = request.get_json() or {}
    ids = data.get('ordered_task_ids')
    if not isinstance(ids, list):
        return jsonify({'error': 'ordered_task_ids must be a list'}), 400

    # Validate tasks belong to this project
    tasks = Task.query.filter(Task.id.in_(ids), Task.project_id == project_id).all()
    ids_set = set(ids)
    if len(tasks) != len(ids):
        return jsonify({'error': 'Invalid task ids or mismatch with project'}), 400

    try:
        for idx, tid in enumerate(ids):
            t = next((x for x in tasks if x.id == tid), None)
            if t:
                t.position = idx

        # Audit task order changes from Board drag/drop.
        audit = AuditLog(
            entity_type='Project',
            entity_id=project_id,
            action='UPDATE',
            user_id=current_user.id if getattr(current_user, 'is_authenticated', False) else None,
            changes={
                'message': 'Orden de tareas actualizado en vista Board',
                'ordered_task_ids': ids,
            }
        )
        db.session.add(audit)
        db.session.commit()

        # Recompute WBS numbers for current project view and return mapping so client can update UI
        tasks_all = Task.query.filter_by(project_id=project_id).all()
        tasks_by_id = {t.id: t for t in tasks_all}
        children_map = {t.id: [] for t in tasks_all}
        root_ids = []
        for t in tasks_all:
            if t.parent_task_id and t.parent_task_id in tasks_by_id:
                children_map[t.parent_task_id].append(t)
            else:
                root_ids.append(t.id)

        status_order = {'BACKLOG': 0, 'IN_PROGRESS': 1, 'IN_REVIEW': 2, 'COMPLETED': 3}
        priority_order = {'CRITICAL': 3, 'HIGH': 2, 'MEDIUM': 1, 'LOW': 0}
        def sort_key(x):
            if getattr(x, 'position', None) is not None:
                return (0, x.position)
            return (1, status_order.get(x.status, 99), -priority_order.get(x.priority, 0), x.title or '')

        wbs_map = {}
        def build_and_assign(tid, depth=0, prefix='', index=1):
            t = tasks_by_id[tid]
            wbs = f"{prefix}.{index}" if prefix else str(index)
            wbs_map[tid] = wbs
            children = sorted(children_map[tid], key=sort_key)
            for i, c in enumerate(children, 1):
                build_and_assign(c.id, depth+1, wbs, i)

        roots = [tasks_by_id[rid] for rid in root_ids]
        roots_sorted = sorted(roots, key=sort_key)
        for i, r in enumerate(roots_sorted, 1):
            build_and_assign(r.id, 0, '', i)

        return jsonify({'status': 'ok', 'wbs': wbs_map})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@main_bp.route('/project/<int:project_id>/kanban')
@login_required
def project_kanban(project_id):
    project = Project.query.get_or_404(project_id)
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None

    # Usuario sin rol: no puede ver nada
    if not user_role:
        flash('No tienes un rol asignado. Contacta al administrador para obtener acceso.', 'warning')
        return redirect(url_for('main.projects'))

    # Control de acceso según rol
    if user_role in ['PMP', 'Admin']:
        pass
    elif user_role == 'Supervisor':
        if current_user not in project.members:
            flash('No tienes permiso para ver este proyecto.', 'danger')
            return redirect(url_for('main.projects'))
    elif user_role == 'Participante':
        has_tasks = Task.query.filter(Task.project_id == project_id).filter(
            (Task.assigned_to_id == current_user.id) | (Task.assignees.any(User.id == current_user.id))
        ).first()
        is_member = current_user in project.members
        if not has_tasks and not is_member:
            flash('No tienes permiso para ver este proyecto.', 'danger')
            return redirect(url_for('main.projects'))
    elif user_role == 'Cliente' or not current_user.is_internal:
        if current_user not in project.clients:
            flash('No tienes permiso para ver este proyecto.', 'danger')
            return redirect(url_for('main.projects'))
    else:
        flash('No tienes permiso para ver este proyecto.', 'danger')
        return redirect(url_for('main.projects'))

    # Filtrar tareas según rol
    if user_role in ['PMP', 'Admin', 'Supervisor']:
        tasks = Task.query.filter_by(project_id=project_id).all()
    elif user_role == 'Participante':
        tasks = Task.query.filter_by(project_id=project_id).all()
    elif user_role == 'Cliente' or not current_user.is_internal:
        tasks = Task.query.filter(Task.project_id == project_id).filter(
            (Task.is_external_visible == True) | (Task.assigned_client_id == current_user.id)
        ).all()
    else:
        tasks = Task.query.filter_by(project_id=project_id, assigned_to_id=current_user.id).all()
    
    assignees = User.query.filter_by(is_internal=True, is_active=True).order_by(User.first_name).all()
    metrics = calculate_project_metrics(project_id)
    
    # Agrupar tareas por estado
    tasks_by_status = {
        'BACKLOG': [t for t in tasks if t.status == 'BACKLOG'],
        'IN_PROGRESS': [t for t in tasks if t.status == 'IN_PROGRESS'],
        'IN_REVIEW': [t for t in tasks if t.status == 'IN_REVIEW'],
        'COMPLETED': [t for t in tasks if t.status == 'COMPLETED']
    }
    
    meta = project.metadata_json or {}
    wip_limits = {
        'BACKLOG':     meta.get('wip_limits', {}).get('BACKLOG'),
        'IN_PROGRESS': meta.get('wip_limits', {}).get('IN_PROGRESS'),
        'IN_REVIEW':   meta.get('wip_limits', {}).get('IN_REVIEW'),
        'COMPLETED':   meta.get('wip_limits', {}).get('COMPLETED'),
    }
    return render_template('kanban.html', project=project, tasks_by_status=tasks_by_status,
                          metrics=metrics, users=assignees, now=datetime.now(),
                          wip_limits=wip_limits)


@main_bp.route('/project/<int:project_id>/gantt')
@login_required
def project_gantt(project_id):
    """Vista Gantt del proyecto con timeline de tareas."""
    project = Project.query.get_or_404(project_id)
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None

    # Usuario sin rol: no puede ver nada
    if not user_role:
        flash('No tienes un rol asignado. Contacta al administrador para obtener acceso.', 'warning')
        return redirect(url_for('main.projects'))

    # Control de acceso según rol
    if user_role in ['PMP', 'Admin']:
        pass
    elif user_role == 'Supervisor':
        if current_user not in project.members:
            flash('No tienes permiso para ver este proyecto.', 'danger')
            return redirect(url_for('main.projects'))
    elif user_role == 'Participante':
        has_tasks = Task.query.filter(Task.project_id == project_id).filter(
            (Task.assigned_to_id == current_user.id) | (Task.assignees.any(User.id == current_user.id))
        ).first()
        is_member = current_user in project.members
        if not has_tasks and not is_member:
            flash('No tienes permiso para ver este proyecto.', 'danger')
            return redirect(url_for('main.projects'))
    elif user_role == 'Cliente' or not current_user.is_internal:
        if current_user not in project.clients:
            flash('No tienes permiso para ver este proyecto.', 'danger')
            return redirect(url_for('main.projects'))
    else:
        flash('No tienes permiso para ver este proyecto.', 'danger')
        return redirect(url_for('main.projects'))

    # Filtrar tareas según rol
    if user_role in ['PMP', 'Admin', 'Supervisor']:
        tasks = Task.query.filter_by(project_id=project_id).order_by(Task.position, Task.id).all()
    elif user_role == 'Participante':
        tasks = Task.query.filter_by(project_id=project_id).order_by(Task.position, Task.id).all()
    elif user_role == 'Cliente' or not current_user.is_internal:
        from sqlalchemy import or_
        tasks = Task.query.filter(
            Task.project_id == project_id,
            or_(Task.is_internal_only == False, Task.is_internal_only == None)
        ).order_by(Task.position, Task.id).all()
    else:
        tasks = Task.query.filter_by(project_id=project_id, assigned_to_id=current_user.id).order_by(Task.position, Task.id).all()
    
    assignees = User.query.filter_by(is_internal=True, is_active=True).order_by(User.first_name).all()
    metrics = calculate_project_metrics(project_id)

    # Cargar horas reales por tarea en una sola query (evitar N+1)
    task_ids = [t.id for t in tasks]
    actual_hours_map = {}
    if task_ids:
        from sqlalchemy import func as sqlfunc
        rows = db.session.query(
            TimeEntry.task_id,
            sqlfunc.sum(TimeEntry.hours).label('total')
        ).filter(TimeEntry.task_id.in_(task_ids)).group_by(TimeEntry.task_id).all()
        actual_hours_map = {row.task_id: float(row.total) for row in rows}

    # Preparar datos para el Gantt
    gantt_tasks = []
    tasks_without_dates = []
    for task in tasks:
        start = task.start_date
        end = task.due_date

        if not start and not end:
            tasks_without_dates.append({
                'id': task.id,
                'title': task.title,
                'status': task.status,
                'priority': task.priority,
                'assignee': task.assigned_to.first_name if task.assigned_to else None,
            })
            continue

        if not start:
            start = end
        if not end:
            days = int(task.estimated_hours / 8) if task.estimated_hours else 1
            end = start + timedelta(days=max(1, days))

        # Progreso basado en horas reales vs estimadas (si disponible), sino por status
        actual_h = actual_hours_map.get(task.id, 0)
        estimated_h = float(task.estimated_hours) if task.estimated_hours else 0
        if task.status == 'COMPLETED':
            progress = 100
        elif estimated_h > 0:
            progress = min(95, int((actual_h / estimated_h) * 100))
        elif task.status == 'IN_REVIEW':
            progress = 75
        elif task.status == 'IN_PROGRESS':
            progress = 40
        else:
            progress = 0

        dependencies = ','.join([str(p.id) for p in task.predecessors]) if task.predecessors else ''

        gantt_tasks.append({
            'id': str(task.id),
            'name': task.title,
            'start': start.strftime('%Y-%m-%d') if hasattr(start, 'strftime') else str(start)[:10],
            'end': end.strftime('%Y-%m-%d') if hasattr(end, 'strftime') else str(end)[:10],
            'progress': progress,
            'dependencies': dependencies,
            'status': task.status,
            'priority': task.priority,
            'assignee': task.assigned_to.first_name if task.assigned_to else None,
            'estimated_hours': estimated_h,
            'actual_hours': actual_h,
            'parent_id': task.parent_task_id,
            'is_blocked': task.is_blocked(),
        })

    return render_template('gantt.html', project=project, tasks=tasks, gantt_tasks=gantt_tasks,
                           tasks_without_dates=tasks_without_dates,
                           metrics=metrics, users=assignees, now=datetime.now())


@main_bp.route('/project/<int:project_id>/report')
@login_required
def project_report(project_id):
    """Vista de reporte imprimible del proyecto (genera PDF via window.print())."""
    project = Project.query.get_or_404(project_id)
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None

    # Control de acceso
    if user_role in ['PMP', 'Admin']:
        pass
    elif user_role == 'Participante':
        has_tasks = Task.query.filter(
            Task.project_id == project_id,
            (Task.assigned_to_id == current_user.id) | (Task.assignees.any(User.id == current_user.id))
        ).first()
        if not has_tasks and current_user not in project.members:
            abort(403)
    elif not current_user.is_internal:
        if current_user not in project.clients:
            abort(403)
    else:
        abort(403)

    # Tareas según rol
    from sqlalchemy import or_
    if user_role in ['PMP', 'Admin', 'Participante']:
        tasks = Task.query.filter_by(project_id=project_id).order_by(Task.position, Task.id).all()
    else:
        tasks = Task.query.filter(
            Task.project_id == project_id,
            or_(Task.is_internal_only == False, Task.is_internal_only == None)
        ).order_by(Task.position, Task.id).all()

    metrics = calculate_project_metrics(project_id)

    # Horas reales por tarea en una query
    task_ids = [t.id for t in tasks]
    actual_hours_map = {}
    if task_ids:
        rows = db.session.query(
            TimeEntry.task_id,
            func.sum(TimeEntry.hours).label('total')
        ).filter(TimeEntry.task_id.in_(task_ids)).group_by(TimeEntry.task_id).all()
        actual_hours_map = {row.task_id: float(row.total) for row in rows}

    # Enriquecer tasks con horas reales para el template
    for t in tasks:
        t.actual_hours = actual_hours_map.get(t.id, 0)

    # Agrupar por status
    tasks_by_status = {
        'BACKLOG': [t for t in tasks if t.status == 'BACKLOG'],
        'IN_PROGRESS': [t for t in tasks if t.status == 'IN_PROGRESS'],
        'IN_REVIEW': [t for t in tasks if t.status == 'IN_REVIEW'],
        'COMPLETED': [t for t in tasks if t.status == 'COMPLETED'],
    }

    return render_template('project_report.html',
                           project=project,
                           tasks=tasks,
                           tasks_by_status=tasks_by_status,
                           metrics=metrics,
                           actual_hours_map=actual_hours_map,
                           now=datetime.now())


# Crear tarea desde modal en project_detail
@main_bp.route('/task', methods=['POST'])
@login_required
def create_task():
    project_id = request.form.get('project_id')
    project = Project.query.get_or_404(project_id)

    # Only internal users with appropriate roles can create tasks
    if (not current_user.is_internal) or (current_user.role and current_user.role.name == 'Participante'):
        flash('No tienes permiso para crear tareas en este proyecto.', 'danger')
        return redirect(url_for('main.project_detail', project_id=project_id))

    try:
        title = request.form.get('title', '').strip()
        if not title:
            flash('El título de la tarea es obligatorio.', 'danger')
            return redirect(url_for('main.project_detail', project_id=project_id))

        # Allow setting status from form (for Kanban view)
        status = request.form.get('status', 'BACKLOG')
        if status not in ['BACKLOG', 'IN_PROGRESS', 'IN_REVIEW', 'COMPLETED']:
            status = 'BACKLOG'

        task = Task(
            project_id=project.id,
            title=title,
            description=request.form.get('description') or None,
            status=status,
            priority=request.form.get('priority') or 'MEDIUM'
        )
        
        if status == 'COMPLETED':
            task.completed_at = datetime.now()

        # Assign clients (multi-select); keep assigned_client_id as primary/legacy
        client_ids = [int(x) for x in request.form.getlist('assigned_client_id') if x and x.strip()]
        if client_ids:
            client_users = User.query.filter(User.id.in_(client_ids)).all()
            task.assigned_clients = client_users
            task.assigned_client_id = client_ids[0]

        start_date_str = request.form.get('start_date')
        due_date_str = request.form.get('due_date')

        # Parse dates if provided and validate
        start_dt = None
        due_dt = None
        if start_date_str:
            try:
                start_dt = datetime.fromisoformat(start_date_str)
                task.start_date = start_dt
            except Exception:
                flash('Formato de fecha de inicio inválido.', 'danger')
                return redirect(url_for('main.project_detail', project_id=project_id))

        if due_date_str:
            try:
                due_dt = datetime.fromisoformat(due_date_str)
                task.due_date = due_dt
            except Exception:
                flash('Formato de fecha de vencimiento inválido.', 'danger')
                return redirect(url_for('main.project_detail', project_id=project_id))

        # Fix 8: Validación de fechas: if both provided ensure start <= due
        if start_dt and due_dt and start_dt > due_dt:
            flash('La fecha de vencimiento no puede ser anterior a la de inicio.', 'danger')
            return redirect(url_for('main.project_detail', project_id=project_id))

        estimated_hours = request.form.get('estimated_hours')
        if estimated_hours and estimated_hours.strip():
            task.estimated_hours = float(estimated_hours)

        # Assign to a user(s) if provided (multi-select)
        assignee_ids = [int(x) for x in request.form.getlist('assignees') if x and x.strip()]
        if assignee_ids:
            users = User.query.filter(User.id.in_(assignee_ids)).all()
            task.assignees = users
            # Keep compatibility with single assigned_to_id: use first selected if any
            task.assigned_to_id = users[0].id if users else None
            # Auto-add assignees to project members if not already members
            existing_member_ids = {m.id for m in project.members}
            for u in users:
                if u.id not in existing_member_ids:
                    project.members.append(u)
        else:
            # Handle single assignee from assigned_to_id field (for backwards compatibility)
            assigned_to_id = request.form.get('assigned_to_id')
            if assigned_to_id and assigned_to_id.strip():
                task.assigned_to_id = int(assigned_to_id)
                u = User.query.get(int(assigned_to_id))
                if u and u not in project.members:
                    project.members.append(u)

        # Parent task (hierarchy)
        parent_task_id = request.form.get('parent_task_id')
        if parent_task_id and parent_task_id.strip():
            parent_id = int(parent_task_id)
            parent_task = Task.query.get(parent_id)
            if not parent_task or parent_task.project_id != project.id:
                flash('Tarea padre inválida.', 'danger')
                return redirect(url_for('main.project_detail', project_id=project_id))
            task.parent_task_id = parent_id

        # Internal only flag - only visible for PMP/Admin
        task.is_internal_only = request.form.get('is_internal_only') == 'on'

        db.session.add(task)
        db.session.flush()  # Get task ID

        # Assign position at the end of the list
        max_position = db.session.query(db.func.max(Task.position)).filter(Task.project_id == project.id).scalar()
        task.position = (max_position + 1) if max_position is not None else 0

        # Handle predecessors at creation time (if provided)
        predecessor_ids = [int(x) for x in request.form.getlist('predecessor_ids') if x and x.strip()]
        if predecessor_ids:
            task.validate_predecessor_ids(predecessor_ids)
            preds = Task.query.filter(Task.id.in_(predecessor_ids)).all()
            task.predecessors = preds

        # Registrar auditoría de creación
        audit = AuditLog(
            entity_type='Task',
            entity_id=task.id,
            action='CREATE',
            user_id=current_user.id,
            changes={'title': task.title, 'project_id': task.project_id, 'status': task.status}
        )
        db.session.add(audit)
        db.session.commit()

        # Dispatch webhook task.created (non-blocking)
        try:
            from app.services import webhook_service
            webhook_service.dispatch('task.created', {
                'task_id': task.id,
                'task_title': task.title,
                'project_id': task.project_id,
                'project_name': project.name if project else None,
                'user_name': current_user.name if current_user else None,
                'new_status': task.status,
            })
        except Exception:
            current_app.logger.exception('Error dispatching webhook for created task %s', task.id)

        # Handle file attachments
        files = request.files.getlist('attachments')
        attachment_count = 0
        for file in files:
            if file and file.filename and allowed_file(file.filename):
                try:
                    stored_filename, task_folder = get_unique_filename(task.id, file.filename)
                    filepath = os.path.join(task_folder, stored_filename)
                    file.save(filepath)
                    
                    file_size = os.path.getsize(filepath)
                    mime_type = mimetypes.guess_type(file.filename)[0] or 'application/octet-stream'
                    
                    attachment = TaskAttachment(
                        task_id=task.id,
                        filename=file.filename,
                        stored_filename=stored_filename,
                        file_size=file_size,
                        mime_type=mime_type,
                        uploaded_by_id=current_user.id
                    )
                    db.session.add(attachment)
                    attachment_count += 1
                except Exception as e:
                    current_app.logger.error(f"Error uploading attachment: {str(e)}")

        if attachment_count > 0:
            db.session.commit()
            flash(f"Tarea '{task.title}' creada con {attachment_count} archivo(s) adjunto(s).", 'success')
        else:
            flash(f"Tarea '{task.title}' creada.", 'success')

        send_email_setting = SystemSettings.get('notify_task_assigned', 'true')
        send_email = send_email_setting == 'true' or send_email_setting == True
        email_sent = False

        # If we created with multiple assignees via form, notify each (including self-assignment)
        if assignee_ids:
            project_obj = Project.query.get(task.project_id) if task.project_id else None
            for uid in set(assignee_ids):
                try:
                    NotificationService.notify(
                        user_id=uid,
                        title='Nueva tarea asignada',
                        message=f"Se te ha asignado la tarea '{task.title}'{(' en el proyecto '+project_obj.name) if project_obj else ''}",
                        notification_type=NotificationService.TASK_ASSIGNED,
                        related_entity_type='task',
                        related_entity_id=task.id,
                        send_email=send_email,
                        email_context={'task': task, 'project': project_obj, 'assigned_by': current_user, 'task_url': NotificationService._build_task_url(task)}
                    )
                    email_sent = True
                except Exception:
                    current_app.logger.exception('Failed to notify new assignee %s for task %s', uid, task.id)
        else:
            # Backwards compatibility: notify single assigned_to_id if provided
            if task.assigned_to_id:
                NotificationService.notify_task_assigned(
                    task=task,
                    assigned_by_user=current_user,
                    send_email=send_email,
                    notify_client=False
                )
                email_sent = True

        # Notificar al cliente asignado (si existe y es diferente al creador)
        if task.assigned_client_id and task.assigned_client_id != current_user.id:
            NotificationService.notify_task_assigned(
                task=task,
                assigned_by_user=current_user, # Fix 10: Asegurar notificación al cliente
                send_email=send_email,
                notify_client=True
            )
            email_sent = True

        if email_sent and send_email:
            flash('Se ha enviado una notificación por correo al usuario asignado.', 'info')

        # Asegurar que las notificaciones generadas se reflejen en la BD inmediatamente
        db.session.commit()

    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear tarea: {str(e)}', 'danger')

    return redirect(url_for('main.project_detail', project_id=project.id))


# --- Task Import/Export ---

@main_bp.route('/project/<int:project_id>/tasks/template-xlsx')
@login_required
def download_tasks_template(project_id):
    """Download a template Excel file for bulk task import"""
    project = Project.query.get_or_404(project_id)

    # Permission check: only internal users (not Participante)
    if not current_user.is_internal:
        abort(403)
    if current_user.role and current_user.role.name == 'Participante':
        abort(403)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        abort(500, 'openpyxl not installed')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tareas"

    # Define headers
    headers = ['Título', 'Estado', 'Prioridad', 'Inicio', 'Vencimiento']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add example row
    example = ['Tarea de ejemplo', 'BACKLOG', 'MEDIUM', '01/03/2026', '15/03/2026']
    ws.append(example)

    # Auto width columns
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(['Título', 'Estado', 'Prioridad', 'Inicio', 'Vencimiento'], 1):
        col_letter = get_column_letter(i)
        if i == 1:  # Título gets wider
            ws.column_dimensions[col_letter].width = 40
        else:
            ws.column_dimensions[col_letter].width = 15

    # Add reference sheet
    ref_sheet = wb.create_sheet("Referencia")
    ref_sheet.append(['Estado - Valores válidos:'])
    for status in ['BACKLOG', 'IN_PROGRESS', 'IN_REVIEW', 'COMPLETED']:
        ref_sheet.append([status])

    ref_sheet.append([])
    ref_sheet.append(['Prioridad - Valores válidos:'])
    for priority in ['LOW', 'MEDIUM', 'HIGH', 'CRITICAL']:
        ref_sheet.append([priority])

    ref_sheet.append([])
    ref_sheet.append(['Fechas en formato DD/MM/YYYY'])

    ref_sheet.column_dimensions['A'].width = 30

    # Send file
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"tareas_template_{project.id}.xlsx"
    return send_file(
        bio,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@main_bp.route('/project/<int:project_id>/tasks/import-xlsx', methods=['POST'])
@login_required
def import_tasks_xlsx(project_id):
    """Import tasks from an Excel file"""
    project = Project.query.get_or_404(project_id)

    # Permission check: only internal users (not Participante)
    if not current_user.is_internal:
        abort(403)
    if current_user.role and current_user.role.name == 'Participante':
        abort(403)

    # Check if file was provided
    if 'excel_file' not in request.files:
        flash('Por favor selecciona un archivo Excel', 'danger')
        return redirect(url_for('main.project_detail', project_id=project.id))

    file = request.files['excel_file']

    if file.filename == '':
        flash('Por favor selecciona un archivo Excel', 'danger')
        return redirect(url_for('main.project_detail', project_id=project.id))

    # Validate file extension
    if not file.filename.lower().endswith('.xlsx'):
        flash('El archivo debe ser un archivo Excel (.xlsx)', 'danger')
        return redirect(url_for('main.project_detail', project_id=project.id))

    try:
        import openpyxl
        from datetime import datetime

        # Read workbook
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        created_count = 0
        error_count = 0
        errors = []

        # Skip header row (row 1), process from row 2 onwards
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            try:
                # Extract values
                title = row[0] if row and len(row) > 0 else None
                status = row[1] if row and len(row) > 1 else None
                priority = row[2] if row and len(row) > 2 else None
                start_date_val = row[3] if row and len(row) > 3 else None
                due_date_val = row[4] if row and len(row) > 4 else None

                # Validate title is not empty
                if not title or not str(title).strip():
                    error_count += 1
                    errors.append(f"Fila {row_idx}: Falta el título")
                    continue

                title = str(title).strip()

                # Normalize status
                if status:
                    status = status.strip().upper()
                    status = Task.normalize_status(status) if hasattr(Task, 'normalize_status') else status
                else:
                    status = 'BACKLOG'

                # Validate status
                valid_statuses = ['BACKLOG', 'IN_PROGRESS', 'IN_REVIEW', 'COMPLETED']
                if status not in valid_statuses:
                    status = 'BACKLOG'

                # Handle priority
                if priority:
                    priority = str(priority).strip().upper()
                else:
                    priority = 'MEDIUM'

                valid_priorities = ['LOW', 'MEDIUM', 'HIGH', 'CRITICAL']
                if priority not in valid_priorities:
                    priority = 'MEDIUM'

                # Parse dates
                start_date = None
                due_date = None

                if start_date_val:
                    start_date = _parse_date(start_date_val)

                if due_date_val:
                    due_date = _parse_date(due_date_val)

                # Validate date order
                if start_date and due_date and start_date > due_date:
                    error_count += 1
                    errors.append(f"Fila {row_idx} ({title}): Inicio no puede ser después de vencimiento")
                    continue

                # Create task
                task = Task(
                    project_id=project.id,
                    title=title,
                    status=status,
                    priority=priority,
                    start_date=start_date,
                    due_date=due_date
                )
                db.session.add(task)
                db.session.flush()  # Get the task ID

                # Create audit log
                audit = AuditLog(
                    entity_type='Task',
                    entity_id=task.id,
                    action='CREATE',
                    user_id=current_user.id,
                    changes={
                        'title': task.title,
                        'project_id': task.project_id,
                        'status': task.status,
                        'source': 'bulk_import'
                    }
                )
                db.session.add(audit)
                created_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f"Fila {row_idx}: {str(e)}")
                current_app.logger.exception(f"Error importing task at row {row_idx}")

        # Commit all tasks
        db.session.commit()

        # Build flash message
        message = f"Se importaron {created_count} tarea(s)"
        if error_count > 0:
            message += f" con {error_count} error(es)"

        if error_count > 0 and errors:
            # Show first 3 errors
            error_details = "; ".join(errors[:3])
            if len(errors) > 3:
                error_details += f"; y {len(errors) - 3} error(es) más"
            flash(f"{message}. Errores: {error_details}", 'warning')
        else:
            flash(message, 'success')

    except Exception as e:
        current_app.logger.exception("Error importing tasks from Excel")
        flash(f'Error al importar archivo: {str(e)}', 'danger')

    return redirect(url_for('main.project_detail', project_id=project.id))


def _parse_date(date_val):
    """Parse date from various formats"""
    if isinstance(date_val, str):
        # Try DD/MM/YYYY
        try:
            return datetime.strptime(date_val.strip(), '%d/%m/%Y')
        except:
            pass

        # Try YYYY-MM-DD
        try:
            return datetime.strptime(date_val.strip(), '%Y-%m-%d')
        except:
            pass

        return None
    elif hasattr(date_val, 'isoformat'):
        # Already a date/datetime object from openpyxl
        return date_val

    return None


# --- Admin: User and Role Management (PMP or Admin) ---

def _ensure_pmp():
    """Ensure user has PMP or Admin role for management features"""
    from flask import session
    uid = session.get('_user_id') or current_user.get_id()
    if not uid:
        abort(403)
    user = User.query.get(int(uid))
    if not user or not user.is_internal:
        abort(403)
    if not user.role_id:
        abort(403)
    role = Role.query.get(user.role_id)
    if not role or role.name not in ('PMP', 'Admin'):
        abort(403)
    return True


# ========== ADMIN DASHBOARD ==========

@main_bp.route('/admin')
@login_required
def admin_dashboard():
    if not _ensure_pmp():
        return redirect(url_for('main.dashboard'))
    
    # Estadísticas generales
    stats = {
        'total_users': User.query.count(),
        'active_users': User.query.filter_by(is_active=True).count(),
        'total_projects': Project.query.count(),
        'total_hours': db.session.query(func.sum(TimeEntry.hours)).scalar() or 0
    }
    
    return render_template('admin/index.html', stats=stats)


@main_bp.route('/admin/users')
@login_required
def admin_users():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    users = User.query.order_by(User.email).all()
    roles = Role.query.order_by(Role.name).all()
    return render_template('admin/users.html', users=users, roles=roles)


@main_bp.route('/admin/user/new', methods=['GET', 'POST'])
@login_required
def admin_create_user():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    roles = Role.query.order_by(Role.name).all()

    if request.method == 'POST':
        try:
            email = request.form.get('email', '').strip()
            if not email:
                flash('El email es obligatorio.', 'danger')
                return render_template('admin/user_edit.html', user=None, roles=roles)

            if User.query.filter_by(email=email).first():
                flash('Ya existe un usuario con ese email.', 'danger')
                return render_template('admin/user_edit.html', user=None, roles=roles)

            from werkzeug.security import generate_password_hash
            password = request.form.get('password', '').strip()
            if not password:
                password = 'changeme123'  # Default password

            user = User(
                email=email,
                first_name=request.form.get('first_name', '').strip() or None,
                last_name=request.form.get('last_name', '').strip() or None,
                password_hash=generate_password_hash(password),
                is_internal=request.form.get('is_internal') == 'on',
                is_active=request.form.get('is_active') == 'on'
            )
            role_id = request.form.get('role_id')
            if role_id:
                user.role_id = int(role_id)

            db.session.add(user)
            db.session.flush()  # Get user ID
            
            # Registrar auditoría
            audit = AuditLog(
                entity_type='User',
                entity_id=user.id,
                action='CREATE',
                user_id=current_user.id,
                changes={'email': user.email, 'is_internal': user.is_internal, 'role_id': user.role_id}
            )
            db.session.add(audit)
            db.session.commit()
            flash(f'Usuario {email} creado exitosamente.', 'success')
            return redirect(url_for('main.admin_users'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear usuario: {str(e)}', 'danger')

    return render_template('admin/user_edit.html', user=None, roles=roles)


@main_bp.route('/admin/user/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def admin_edit_user(user_id):
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    user = User.query.get_or_404(user_id)
    roles = Role.query.order_by(Role.name).all()

    if request.method == 'POST':
        try:
            # Guardar valores anteriores para auditoría
            old_values = {
                'role_id': user.role_id,
                'is_internal': user.is_internal,
                'is_active': user.is_active
            }
            
            role_id = request.form.get('role_id')
            if role_id:
                user.role_id = int(role_id)
            user.is_internal = True if request.form.get('is_internal') == 'on' else False
            user.is_active = True if request.form.get('is_active') == 'on' else False
            
            # Detectar cambios para auditoría
            new_values = {
                'role_id': user.role_id,
                'is_internal': user.is_internal,
                'is_active': user.is_active
            }
            changes = {}
            for field, old_val in old_values.items():
                if old_val != new_values[field]:
                    changes[field] = {'old': old_val, 'new': new_values[field]}
            
            # Cambiar contraseña solo para usuarios locales (sin azure_oid)
            new_password = request.form.get('password', '').strip()
            if new_password and not user.azure_oid:
                user.set_password(new_password)
                changes['password'] = {'old': '***', 'new': '***'}
            
            if changes:
                audit = AuditLog(
                    entity_type='User',
                    entity_id=user.id,
                    action='UPDATE',
                    user_id=current_user.id,
                    changes=changes
                )
                db.session.add(audit)
            
            db.session.commit()
            flash('Usuario actualizado.', 'success')
            return redirect(url_for('main.admin_users'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar usuario: {str(e)}', 'danger')

    return render_template('admin/user_edit.html', user=user, roles=roles)


@main_bp.route('/admin/user/<int:user_id>/delete', methods=['POST'])
@login_required
def admin_delete_user(user_id):
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    user = User.query.get_or_404(user_id)
    
    # No permitir eliminar al usuario actual
    if user.id == current_user.id:
        flash('No puedes eliminar tu propio usuario.', 'danger')
        return redirect(url_for('main.admin_users'))
    
    try:
        email = user.email
        user_id_backup = user.id
        
        # Registrar auditoría antes de eliminar
        audit = AuditLog(
            entity_type='User',
            entity_id=user_id_backup,
            action='DELETE',
            user_id=current_user.id,
            changes={'email': email, 'is_internal': user.is_internal}
        )
        db.session.add(audit)
        
        db.session.delete(user)
        db.session.commit()
        flash(f'Usuario {email} eliminado.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar usuario: {str(e)}', 'danger')
    
    return redirect(url_for('main.admin_users'))


@main_bp.route('/admin/roles', methods=['GET', 'POST'])
@login_required
def admin_roles():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    if request.method == 'POST':
        name = request.form.get('name')
        if name:
            if Role.query.filter_by(name=name).first():
                flash('El rol ya existe.', 'warning')
            else:
                role = Role(name=name)
                db.session.add(role)
                db.session.flush()
                
                # Registrar auditoría
                audit = AuditLog(
                    entity_type='Role',
                    entity_id=role.id,
                    action='CREATE',
                    user_id=current_user.id,
                    changes={'name': name}
                )
                db.session.add(audit)
                db.session.commit()
                flash('Rol creado.', 'success')
        return redirect(url_for('main.admin_roles'))
    roles = Role.query.order_by(Role.name).all()
    return render_template('admin/roles.html', roles=roles)


@main_bp.route('/admin/roles/<int:role_id>/delete', methods=['POST'])
@login_required
def admin_delete_role(role_id):
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    role = Role.query.get_or_404(role_id)
    
    # No permitir eliminar roles en uso
    if User.query.filter_by(role_id=role_id).first():
        flash('No se puede eliminar un rol que está asignado a usuarios.', 'danger')
        return redirect(url_for('main.admin_roles'))
    
    try:
        name = role.name
        role_id_backup = role.id
        
        # Registrar auditoría antes de eliminar
        audit = AuditLog(
            entity_type='Role',
            entity_id=role_id_backup,
            action='DELETE',
            user_id=current_user.id,
            changes={'name': name}
        )
        db.session.add(audit)
        
        db.session.delete(role)
        db.session.commit()
        flash(f'Rol "{name}" eliminado.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}', 'danger')
    
    return redirect(url_for('main.admin_roles'))


# ========== ADMIN: TARIFAS ==========

@main_bp.route('/admin/rates', methods=['GET', 'POST'])
@login_required
def admin_rates():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            try:
                rate = HourlyRate(
                    name=request.form.get('name'),
                    rate=float(request.form.get('rate', 0)),
                    currency=request.form.get('currency', 'USD'),
                    description=request.form.get('description'),
                    is_default=request.form.get('is_default') == 'on'
                )
                
                # Si es default, quitar default de los demás
                if rate.is_default:
                    HourlyRate.query.update({HourlyRate.is_default: False})
                
                db.session.add(rate)
                db.session.flush()
                
                # Auditoría
                audit = AuditLog(
                    entity_type='HourlyRate',
                    entity_id=rate.id,
                    action='CREATE',
                    user_id=current_user.id,
                    changes={'name': rate.name, 'rate': rate.rate, 'currency': rate.currency}
                )
                db.session.add(audit)
                db.session.commit()
                flash('Tarifa creada exitosamente.', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error: {str(e)}', 'danger')
        
        elif action == 'update':
            rate_id = request.form.get('rate_id')
            rate = HourlyRate.query.get(rate_id)
            if rate:
                old_values = {'name': rate.name, 'rate': float(rate.rate), 'currency': rate.currency, 'is_active': rate.is_active}
                
                rate.name = request.form.get('name')
                rate.rate = float(request.form.get('rate', 0))
                rate.currency = request.form.get('currency', 'USD')
                rate.description = request.form.get('description')
                rate.is_active = request.form.get('is_active') == 'on'
                
                if request.form.get('is_default') == 'on':
                    HourlyRate.query.update({HourlyRate.is_default: False})
                    rate.is_default = True
                
                new_values = {'name': rate.name, 'rate': float(rate.rate), 'currency': rate.currency, 'is_active': rate.is_active}
                changes = {k: {'old': old_values[k], 'new': new_values[k]} for k in old_values if old_values[k] != new_values[k]}
                
                if changes:
                    audit = AuditLog(
                        entity_type='HourlyRate',
                        entity_id=rate.id,
                        action='UPDATE',
                        user_id=current_user.id,
                        changes=changes
                    )
                    db.session.add(audit)
                
                db.session.commit()
                flash('Tarifa actualizada.', 'success')
        
        elif action == 'delete':
            rate_id = request.form.get('rate_id')
            rate = HourlyRate.query.get(rate_id)
            if rate:
                audit = AuditLog(
                    entity_type='HourlyRate',
                    entity_id=rate.id,
                    action='DELETE',
                    user_id=current_user.id,
                    changes={'name': rate.name, 'rate': float(rate.rate)}
                )
                db.session.add(audit)
                db.session.delete(rate)
                db.session.commit()
                flash('Tarifa eliminada.', 'success')
        
        return redirect(url_for('main.admin_rates'))
    
    rates = HourlyRate.query.order_by(HourlyRate.name).all()
    currencies = ['USD', 'PYG', 'EUR']
    currency_names = {'USD': 'Dólares', 'PYG': 'Guaraníes (PYG)', 'EUR': 'Euro'}
    return render_template('admin/rates.html', rates=rates, currencies=currencies, currency_names=currency_names)


# route to serve branding assets (logos, favicons) stored in uploads/branding
@main_bp.route('/uploads/branding/<path:filename>')
# branding images are public; no authentication required
# we won't enforce login here to let public see the logo
def branding_asset(filename):
    # note: do not allow traversal outside branding folder
    branding_folder = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), 'branding')
    return send_from_directory(branding_folder, filename)

# ========== ADMIN: BRANDING ==========

@main_bp.route('/admin/branding', methods=['GET', 'POST'])
@login_required
def admin_branding():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    
    if request.method == 'POST':
        try:
            # Guardar configuraciones de branding
            SystemSettings.set('app_name', request.form.get('app_name', 'BridgeWork'), 'branding', 'Nombre de la aplicación', user_id=current_user.id)
            SystemSettings.set('app_subtitle', request.form.get('app_subtitle', 'Project Manager'), 'branding', 'Subtítulo', user_id=current_user.id)
            SystemSettings.set('primary_color', request.form.get('primary_color', '#0d6efd'), 'branding', 'Color primario', user_id=current_user.id)
            SystemSettings.set('secondary_color', request.form.get('secondary_color', '#6c757d'), 'branding', 'Color secundario', user_id=current_user.id)
            SystemSettings.set('sidebar_color', request.form.get('sidebar_color', '#1a1d29'), 'branding', 'Color del sidebar', user_id=current_user.id)
            
            # Logo upload
            if 'logo' in request.files:
                file = request.files['logo']
                if file and file.filename:
                    from werkzeug.utils import secure_filename
                    filename = secure_filename(file.filename)
                    logo_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), 'branding')
                    os.makedirs(logo_path, exist_ok=True)
                    filepath = os.path.join(logo_path, f'logo_{filename}')
                    file.save(filepath)
                    SystemSettings.set('logo_path', f'/uploads/branding/logo_{filename}', 'branding', 'Ruta del logo', user_id=current_user.id)
            
            # Favicon upload
            if 'favicon' in request.files:
                file = request.files['favicon']
                if file and file.filename:
                    from werkzeug.utils import secure_filename
                    filename = secure_filename(file.filename)
                    logo_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), 'branding')
                    os.makedirs(logo_path, exist_ok=True)
                    filepath = os.path.join(logo_path, f'favicon_{filename}')
                    file.save(filepath)
                    SystemSettings.set('favicon_path', f'/uploads/branding/favicon_{filename}', 'branding', 'Ruta del favicon', user_id=current_user.id)
            
            # Auditoría de cambios de branding
            audit = AuditLog(
                entity_type='SystemSettings',
                entity_id=0,
                action='UPDATE',
                user_id=current_user.id,
                changes={'category': 'branding', 'app_name': request.form.get('app_name')}
            )
            db.session.add(audit)
            
            db.session.commit()
            flash('Configuración de apariencia guardada.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
        
        return redirect(url_for('main.admin_branding'))
    
    # Obtener configuraciones actuales
    settings = {
        'app_name': SystemSettings.get('app_name', 'BridgeWork'),
        'app_subtitle': SystemSettings.get('app_subtitle', 'Project Manager'),
        'primary_color': SystemSettings.get('primary_color', '#0d6efd'),
        'secondary_color': SystemSettings.get('secondary_color', '#6c757d'),
        'sidebar_color': SystemSettings.get('sidebar_color', '#1a1d29'),
        'logo_path': SystemSettings.get('logo_path'),
        'favicon_path': SystemSettings.get('favicon_path'),
    }
    
    return render_template('admin/branding.html', settings=settings)


# ========== ADMIN: CONFIGURACIÓN GENERAL ==========

@main_bp.route('/admin/general', methods=['GET', 'POST'])
@login_required
def admin_general():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    
    if request.method == 'POST':
        try:
            SystemSettings.set('default_currency', request.form.get('default_currency', 'USD'), 'general', 'Moneda predeterminada', user_id=current_user.id)
            SystemSettings.set('date_format', request.form.get('date_format', 'DD/MM/YYYY'), 'general', 'Formato de fecha', user_id=current_user.id)
            SystemSettings.set('time_format', request.form.get('time_format', '24h'), 'general', 'Formato de hora', user_id=current_user.id)
            SystemSettings.set('language', request.form.get('language', 'es'), 'general', 'Idioma', user_id=current_user.id)
            SystemSettings.set('week_start', request.form.get('week_start', 'monday'), 'general', 'Inicio de semana', user_id=current_user.id)
            SystemSettings.set('default_task_status', request.form.get('default_task_status', 'BACKLOG'), 'general', 'Estado inicial de tareas', user_id=current_user.id)
            SystemSettings.set('require_task_approval', request.form.get('require_task_approval', 'true'), 'general', 'Requerir aprobación de tareas', 'boolean', user_id=current_user.id)
            
            # Auditoría
            audit = AuditLog(
                entity_type='SystemSettings',
                entity_id=0,
                action='UPDATE',
                user_id=current_user.id,
                changes={'category': 'general', 'language': request.form.get('language')}
            )
            db.session.add(audit)
            
            db.session.commit()
            flash('Configuración general guardada.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
        
        return redirect(url_for('main.admin_general'))
    
    settings = {
        'default_currency': SystemSettings.get('default_currency', 'USD'),
        'date_format': SystemSettings.get('date_format', 'DD/MM/YYYY'),
        'time_format': SystemSettings.get('time_format', '24h'),
        'language': SystemSettings.get('language', 'es'),
        'week_start': SystemSettings.get('week_start', 'monday'),
        'default_task_status': SystemSettings.get('default_task_status', 'BACKLOG'),
        'require_task_approval': SystemSettings.get('require_task_approval', True),
    }
    
    currencies = ['USD', 'PYG', 'EUR']
    currency_names = {'USD': 'Dólares', 'PYG': 'Guaraníes (PYG)', 'EUR': 'Euro'}
    return render_template('admin/general.html', settings=settings, currencies=currencies, currency_names=currency_names)


# ========== ADMIN: CONTENIDO Y TEXTOS ==========

@main_bp.route('/admin/content', methods=['GET', 'POST'])
@login_required
def admin_content():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    
    if request.method == 'POST':
        try:
            SystemSettings.set('footer_text', request.form.get('footer_text', ''), 'content', 'Texto del footer', user_id=current_user.id)
            SystemSettings.set('copyright_text', request.form.get('copyright_text', '© 2025 BridgeWork'), 'content', 'Texto de copyright', user_id=current_user.id)
            SystemSettings.set('support_email', request.form.get('support_email', ''), 'content', 'Email de soporte', user_id=current_user.id)
            SystemSettings.set('support_phone', request.form.get('support_phone', ''), 'content', 'Teléfono de soporte', user_id=current_user.id)
            SystemSettings.set('terms_url', request.form.get('terms_url', ''), 'content', 'URL de términos', user_id=current_user.id)
            SystemSettings.set('privacy_url', request.form.get('privacy_url', ''), 'content', 'URL de privacidad', user_id=current_user.id)
            SystemSettings.set('welcome_message', request.form.get('welcome_message', ''), 'content', 'Mensaje de bienvenida', user_id=current_user.id)
            
            # Auditoría
            audit = AuditLog(
                entity_type='SystemSettings',
                entity_id=0,
                action='UPDATE',
                user_id=current_user.id,
                changes={'category': 'content'}
            )
            db.session.add(audit)
            
            db.session.commit()
            flash('Contenido actualizado.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
        
        return redirect(url_for('main.admin_content'))
    
    settings = {
        'footer_text': SystemSettings.get('footer_text', ''),
        'copyright_text': SystemSettings.get('copyright_text', '© 2025 BridgeWork'),
        'support_email': SystemSettings.get('support_email', ''),
        'support_phone': SystemSettings.get('support_phone', ''),
        'terms_url': SystemSettings.get('terms_url', ''),
        'privacy_url': SystemSettings.get('privacy_url', ''),
        'welcome_message': SystemSettings.get('welcome_message', ''),
    }
    
    return render_template('admin/content.html', settings=settings)


# ========== ADMIN: NOTIFICACIONES ==========

@main_bp.route('/admin/notifications', methods=['GET', 'POST'])
@login_required
def admin_notifications():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    
    if request.method == 'POST':
        # Validate base_url first (if provided)
        base_url = (request.form.get('base_url') or '').strip()
        if base_url and not base_url.lower().startswith(('http://', 'https://')):
            flash('La Base URL debe comenzar con http:// o https://', 'danger')
            return redirect(url_for('main.admin_notifications'))

        try:
            SystemSettings.set('email_provider', request.form.get('email_provider', 'stub'), 'notifications', 'Proveedor de email', user_id=current_user.id)
            SystemSettings.set('sendgrid_api_key', request.form.get('sendgrid_api_key', ''), 'notifications', 'API Key de SendGrid', user_id=current_user.id)
            SystemSettings.set('email_from', request.form.get('email_from', ''), 'notifications', 'Email remitente', user_id=current_user.id)
            SystemSettings.set('email_from_name', request.form.get('email_from_name', 'BridgeWork'), 'notifications', 'Nombre remitente', user_id=current_user.id)
            # Normalize and save base_url (remove trailing slash)
            if base_url:
                base_url = base_url.rstrip('/')
            SystemSettings.set('base_url', base_url, 'notifications', 'Base URL para enlaces en emails', user_id=current_user.id)

            SystemSettings.set('notify_task_assigned', request.form.get('notify_task_assigned', 'true'), 'notifications', 'Notificar asignación', 'boolean', user_id=current_user.id)
            SystemSettings.set('notify_task_completed', request.form.get('notify_task_completed', 'true'), 'notifications', 'Notificar completado', 'boolean', user_id=current_user.id)
            SystemSettings.set('notify_task_status_changed', request.form.get('notify_task_status_changed', 'true'), 'notifications', 'Notificar cambio de estado', 'boolean', user_id=current_user.id)
            SystemSettings.set('notify_task_approved', request.form.get('notify_task_approved', 'true'), 'notifications', 'Notificar aprobación', 'boolean', user_id=current_user.id)
            SystemSettings.set('notify_task_rejected', request.form.get('notify_task_rejected', 'true'), 'notifications', 'Notificar rechazo', 'boolean', user_id=current_user.id)
            SystemSettings.set('notify_task_comment', request.form.get('notify_task_comment', 'true'), 'notifications', 'Notificar comentarios', 'boolean', user_id=current_user.id)
            SystemSettings.set('notify_task_mention', request.form.get('notify_task_mention', 'true'), 'notifications', 'Notificar menciones', 'boolean', user_id=current_user.id)
            SystemSettings.set('notify_due_date_reminder', request.form.get('notify_due_date_reminder', 'true'), 'notifications', 'Recordatorio de vencimiento', 'boolean', user_id=current_user.id)
            
            # Auditoría
            audit = AuditLog(
                entity_type='SystemSettings',
                entity_id=0,
                action='UPDATE',
                user_id=current_user.id,
                changes={'category': 'notifications', 'email_provider': request.form.get('email_provider')}
            )
            db.session.add(audit)
            
            db.session.commit()
            flash('Configuración de notificaciones guardada.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
        
        return redirect(url_for('main.admin_notifications'))
    
    settings = {
        'email_provider': SystemSettings.get('email_provider', 'stub'),
        'sendgrid_api_key': SystemSettings.get('sendgrid_api_key', ''),
        'email_from': SystemSettings.get('email_from', ''),
        'email_from_name': SystemSettings.get('email_from_name', 'BridgeWork'),
        'notify_task_assigned': SystemSettings.get('notify_task_assigned', True),
        'notify_task_completed': SystemSettings.get('notify_task_completed', True),
        'notify_task_status_changed': SystemSettings.get('notify_task_status_changed', True),
        'notify_task_approved': SystemSettings.get('notify_task_approved', True),
        'notify_task_rejected': SystemSettings.get('notify_task_rejected', True),
        'notify_task_comment': SystemSettings.get('notify_task_comment', True),
        'notify_task_mention': SystemSettings.get('notify_task_mention', True),
        'notify_due_date_reminder': SystemSettings.get('notify_due_date_reminder', True),
    }
    
    return render_template('admin/notifications_config.html', settings=settings)


# ========== ADMIN: MANTENIMIENTO ==========

@main_bp.route('/admin/maintenance', methods=['GET', 'POST'])
@login_required
def admin_maintenance():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'clear_old_notifications':
            # Eliminar notificaciones leídas de más de 30 días
            cutoff = datetime.now() - timedelta(days=30)
            deleted = SystemNotification.query.filter(
                SystemNotification.is_read == True,
                SystemNotification.created_at < cutoff
            ).delete()
            db.session.commit()
            flash(f'{deleted} notificaciones antiguas eliminadas.', 'success')
        
        elif action == 'clear_old_audit':
            # Eliminar logs de auditoría de más de 90 días
            cutoff = datetime.now() - timedelta(days=90)
            deleted = AuditLog.query.filter(AuditLog.created_at < cutoff).delete()
            db.session.commit()
            flash(f'{deleted} registros de auditoría eliminados.', 'success')
        
        return redirect(url_for('main.admin_maintenance'))
    
    # Estadísticas del sistema
    stats = {
        'db_notifications': SystemNotification.query.count(),
        'db_audit_logs': AuditLog.query.count(),
        'db_time_entries': TimeEntry.query.count(),
        'db_tasks': Task.query.count(),
        'db_projects': Project.query.count(),
        'db_users': User.query.count(),
    }
    
    return render_template('admin/maintenance.html', stats=stats)


@main_bp.route('/projects/new', methods=['POST'])
@login_required
def create_project():
    name = request.form.get('name')
    description = request.form.get('description')
    budget_hours = request.form.get('budget_hours')
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    project_type = request.form.get('project_type', 'APP_DEVELOPMENT')
    client_ids = request.form.getlist('client_ids')
    member_ids = request.form.getlist('member_ids')
    pmp_ids = request.form.getlist('pmp_ids')
    supervisor_ids = request.form.getlist('supervisor_ids')
    department_id_raw = request.form.get('department_id', '').strip()
    # Sanitize incoming id lists: remove empty strings and cast to int
    try:
        client_ids = [int(x) for x in client_ids if x and str(x).strip()]
    except ValueError:
        client_ids = []
    try:
        member_ids = [int(x) for x in member_ids if x and str(x).strip()]
    except ValueError:
        member_ids = []
    try:
        pmp_ids = [int(x) for x in pmp_ids if x and str(x).strip()]
    except ValueError:
        pmp_ids = []
    try:
        supervisor_ids = [int(x) for x in supervisor_ids if x and str(x).strip()]
    except ValueError:
        supervisor_ids = []
    department_id = int(department_id_raw) if department_id_raw else None
    
    if not name:
        flash('El nombre del proyecto es un campo obligatorio.', 'danger')
        return redirect(url_for('main.projects'))
    
    # Fix 8: Validación de coherencia de fechas
    if start_date_str and end_date_str and start_date_str > end_date_str:
        flash('La fecha de finalización no puede ser anterior a la fecha de inicio.', 'danger')
        return redirect(url_for('main.projects'))

    # Only internal users with appropriate roles can create projects
    if not current_user.is_internal or (current_user.role and current_user.role.name in ['Participante', 'Cliente']):
        flash('No tienes permiso para crear proyectos.', 'danger')
        return redirect(url_for('main.projects'))
    
    try:
        new_project = Project(
            name=name,
            description=description,
            budget_hours=float(budget_hours) if budget_hours and budget_hours.strip() else None,
            status='ACTIVE',
            project_type=project_type,
            manager_id=current_user.id if current_user.is_internal else None,
            start_date=datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else datetime.now().date(),
            end_date=datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None,
            department_id=department_id
        )
        db.session.add(new_project)
        db.session.flush()  # Get the project ID
        
        # Associate clients
        if client_ids:
            clients = User.query.filter(User.id.in_(client_ids)).all()
            new_project.clients = clients
            
        # Associate members (PMPs adicionales + Supervisores + Participantes)
        all_member_ids = list(set(pmp_ids + supervisor_ids + member_ids))
        if all_member_ids:
            members = User.query.filter(User.id.in_(all_member_ids)).all()
            new_project.members = members
        
        # Registrar auditoría de creación
        audit = AuditLog(
            entity_type='Project',
            entity_id=new_project.id,
            action='CREATE',
            user_id=current_user.id,
            changes={'name': name, 'status': 'ACTIVE', 'budget_hours': float(new_project.budget_hours) if new_project.budget_hours else None}
        )
        db.session.add(audit)
        
        db.session.commit()
        flash(f"Proyecto '{name}' creado con éxito.", 'success')

        # Fix 6: Notificación de creación de proyecto a clientes
        # Build external project URL for emails (fallback to base_url or config)
        try:
            project_url = url_for('main.project_detail', project_id=new_project.id, _external=True)
        except Exception:
            try:
                from ..models import SystemSettings
                host = SystemSettings.get('base_url') or current_app.config.get('SERVER_NAME') or current_app.config.get('BASE_URL')
            except Exception:
                host = current_app.config.get('SERVER_NAME') or current_app.config.get('BASE_URL')
            if host:
                host = host.rstrip('/')
                project_url = f"{current_app.config.get('PREFERRED_URL_SCHEME','https')}://{host}/project/{new_project.id}"
            else:
                project_url = f"/project/{new_project.id}"

        if new_project.clients:
            for client in new_project.clients:
                try:
                    NotificationService.notify(
                        user_id=client.id,
                        title='Nuevo Proyecto Asignado',
                        message=f"Has sido asignado al proyecto '{new_project.name}'.",
                        notification_type=NotificationService.PROJECT_CREATED,
                        related_entity_type='project',
                        related_entity_id=new_project.id,
                        send_email=True,
                        email_context={
                            'project': new_project,
                            'project_url': project_url,
                            'message': f"Has sido asignado al proyecto '{new_project.name}'.",
                            'title': 'Nuevo Proyecto Asignado'
                        }
                    )
                except Exception:
                    current_app.logger.exception('Failed to notify client %s for new project %s', client.id, new_project.id)

        # Notificación a participantes
        if new_project.members:
            for member in new_project.members:
                try:
                    NotificationService.notify(
                        user_id=member.id,
                        title='Nuevo Proyecto Asignado',
                        message=f"Has sido añadido como participante al proyecto '{new_project.name}'.",
                        notification_type=NotificationService.PROJECT_CREATED,
                        related_entity_type='project',
                        related_entity_id=new_project.id,
                        send_email=True,
                        email_context={
                            'project': new_project,
                            'project_url': project_url,
                            'message': f"Has sido añadido como participante al proyecto '{new_project.name}'.",
                            'title': 'Nuevo Proyecto Asignado'
                        }
                    )
                except Exception:
                    current_app.logger.exception('Failed to notify member %s for new project %s', member.id, new_project.id)
        
        # Asegurar que las notificaciones generadas se reflejen en la BD inmediatamente
        db.session.commit()

        # Webhook: project.created
        try:
            from app.services import webhook_service
            webhook_service.dispatch('project.created', {
                'project_id': new_project.id,
                'project_name': new_project.name,
                'user_name': current_user.name or current_user.email,
                'new_status': new_project.status,
                'department_name': new_project.department.name if new_project.department_id and new_project.department else None,
            })
        except Exception:
            current_app.logger.exception('Failed to dispatch project.created webhook')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear proyecto: {str(e)}', 'danger')

    return redirect(url_for('main.projects'))

@main_bp.route('/calendar')
@login_required
def calendar_view():
    """Calendar view — tasks with dates rendered via FullCalendar."""
    user_role = current_user.role.name if current_user.role else None
    projects = Project.query.order_by(Project.name).all()
    return render_template('calendar.html',
                           projects=projects,
                           current_user_role_name=user_role,
                           now=datetime.now())


@main_bp.route('/api/calendar-events')
@login_required
def calendar_events():
    """Return tasks as FullCalendar-compatible JSON events."""
    user_role = current_user.role.name if current_user.role else None

    # Optional filters from query string
    project_id = request.args.get('project_id', type=int)
    start_range = request.args.get('start')  # ISO date string from FullCalendar
    end_range = request.args.get('end')

    q = Task.query.filter(
        (Task.due_date != None) | (Task.start_date != None)
    )

    # Role-based visibility
    if user_role in ['PMP', 'Admin']:
        pass  # see all tasks
    elif user_role == 'Participante':
        q = q.filter(
            (Task.assigned_to_id == current_user.id) |
            Task.assignees.any(id=current_user.id)
        )
    else:
        # Client: only tasks visible externally on their projects
        client_project_ids = [p.id for p in Project.query.filter(
            Project.clients.any(id=current_user.id)
        ).all()]
        q = q.filter(Task.project_id.in_(client_project_ids), Task.is_external_visible == True)

    if project_id:
        q = q.filter(Task.project_id == project_id)

    if start_range:
        try:
            q = q.filter(Task.due_date >= datetime.fromisoformat(start_range.replace('Z', '')))
        except Exception:
            pass
    if end_range:
        try:
            q = q.filter(Task.due_date <= datetime.fromisoformat(end_range.replace('Z', '')))
        except Exception:
            pass

    STATUS_COLORS = {
        'BACKLOG':     '#94a3b8',
        'TODO':        '#3b82f6',
        'IN_PROGRESS': '#f59e0b',
        'IN_REVIEW':   '#8b5cf6',
        'COMPLETED':   '#10b981',
        'DONE':        '#10b981',
        'ACCEPTED':    '#10b981',
    }
    PRIORITY_BORDER = {
        'LOW':      '#94a3b8',
        'MEDIUM':   '#3b82f6',
        'HIGH':     '#f59e0b',
        'CRITICAL': '#ef4444',
    }

    events = []
    for t in q.all():
        start = None
        end = None
        if t.start_date:
            start = t.start_date.date().isoformat()
        if t.due_date:
            # FullCalendar end is exclusive — add 1 day for all-day events
            from datetime import date as _date
            end_date = t.due_date.date()
            if start and end_date.isoformat() == start:
                from datetime import timedelta as _td
                end_date = end_date + _td(days=1)
            end = end_date.isoformat()

        if not start and not end:
            continue

        color = STATUS_COLORS.get(t.status, '#94a3b8')
        border = PRIORITY_BORDER.get(t.priority, '#64748b')
        project_name = t.project.name if t.project else ''

        events.append({
            'id': t.id,
            'title': t.title,
            'start': start or end,
            'end': end,
            'color': color,
            'borderColor': border,
            'url': url_for('main.task_detail', task_id=t.id),
            'extendedProps': {
                'status': t.status,
                'priority': t.priority or 'MEDIUM',
                'project': project_name,
                'project_id': t.project_id,
                'editable': user_role in ['PMP', 'Admin'] or t.assigned_to_id == current_user.id,
            }
        })

    return jsonify(events)


@main_bp.route('/api/tasks/<int:task_id>/dates', methods=['PATCH'])
@login_required
def update_task_dates(task_id):
    """Update task start_date and due_date from calendar drag & drop."""
    task = Task.query.get_or_404(task_id)
    user_role = current_user.role.name if current_user.role else None

    # Only assignee, PMP or Admin can reschedule
    can_edit = (
        user_role in ['PMP', 'Admin'] or
        task.assigned_to_id == current_user.id or
        current_user in task.assignees
    )
    if not can_edit:
        return jsonify({'error': 'Sin permisos para modificar esta tarea'}), 403

    data = request.get_json(force=True, silent=True) or {}
    try:
        if 'start_date' in data and data['start_date']:
            task.start_date = datetime.fromisoformat(data['start_date'].replace('Z', ''))
        if 'due_date' in data and data['due_date']:
            task.due_date = datetime.fromisoformat(data['due_date'].replace('Z', ''))
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@main_bp.route('/tasks')
@login_required
def tasks():
    """
    Visibilidad de tareas según rol:
    - PMP/Admin: puede ver todas las tareas
    - Participante: solo tareas asignadas a él
    - Cliente: solo tareas de sus proyectos que son visibles externamente
    - Sin Rol: no puede ver nada
    """
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None
    
    # Usuario sin rol: mostrar mensaje y lista vacía
    if not user_role:
        flash('No tienes un rol asignado. Contacta al administrador para obtener acceso.', 'warning')
        return render_template('tasks.html', tasks=[], no_role=True, now=datetime.now(), status_filter='', priority_filter='')
    
    # Fix 11: Asegurar que Admin vea sus tareas asignadas explícitamente si el filtro general fallara
    # (Aunque la lógica 'PMP/Admin' ve todo, esto refuerza la consistencia si el rol tuviera problemas)
    if user_role == 'Admin' and not current_user.is_internal:
        user_role = 'Cliente' # Fallback de seguridad
    
    # Filtros
    status_filter = request.args.get('status', '')
    priority_filter = request.args.get('priority', '')
    
    query = Task.query
    
    # Aplicar filtro según rol
    if user_role in ['PMP', 'Admin']:
        # PMP/Admin ve todas las tareas
        pass
    elif user_role == 'Participante':
        # Participante: solo ver tareas donde está asignado (incluye multi-asignados)
        query = query.filter(
            (Task.assigned_to_id == current_user.id) | (Task.assignees.any(User.id == current_user.id))
        )
    elif user_role == 'Cliente' or not current_user.is_internal:
        # Cliente: ver tareas asignadas explícitamente o visibles externamente en sus proyectos
        from sqlalchemy import or_
        client_project_ids = db.session.query(Project.id).filter(
            Project.clients.any(User.id == current_user.id)
        ).subquery()
        query = query.filter(
            or_(
                Task.assigned_client_id == current_user.id,
                (Task.is_external_visible == True) & (Task.project_id.in_(client_project_ids))
            )
        )
    else:
        # Cualquier otro rol: no ve nada
        query = query.filter(Task.id == -1)  # Query que no devuelve nada
    
    if status_filter:
        # Normalize legacy filter 'DONE' to canonical 'COMPLETED'
        if status_filter == 'DONE':
            status_filter = 'COMPLETED'
        query = query.filter(Task.status == status_filter)
    
    if priority_filter:
        query = query.filter(Task.priority == priority_filter)
    
    tasks = query.order_by(Task.due_date).all()
    
    # Calcular horas por tarea
    for t in tasks:
        t.hours_spent = db.session.query(func.sum(TimeEntry.hours)).filter_by(task_id=t.id).scalar() or 0
    
    return render_template('tasks.html', 
        tasks=tasks, 
        now=datetime.now(),
        status_filter=status_filter,
        priority_filter=priority_filter,
        no_role=False
    )

@main_bp.route('/time')
@main_bp.route('/time-entries')
@login_required
def time_entries():
    # Filters and pagination
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None
    page = int(request.args.get('page', 1))
    per_page = 25

    # Usuario sin rol: mostrar mensaje y lista vacía
    if not user_role:
        flash('No tienes un rol asignado. Contacta al administrador para obtener acceso.', 'warning')
        return render_template('time_entries.html', time_entries=[], total_hours_week=0, page=1, total_pages=0, users=[], total_count=0, page_urls=[], prev_url=None, next_url=None, no_role=True)

    # Parse filters
    start_date_s = request.args.get('start_date')
    end_date_s = request.args.get('end_date')
    filter_user_id = request.args.get('user_id') if (user_role in ['PMP', 'Admin']) else None

    query = TimeEntry.query

    # Apply role-based visibility
    if user_role in ['PMP', 'Admin']:
        pass  # all entries visible, filters below will narrow if provided
    else:
        query = query.filter(TimeEntry.user_id == current_user.id)

    # Apply date filters
    if start_date_s:
        try:
            start_date = datetime.strptime(start_date_s, '%Y-%m-%d').date()
            query = query.filter(TimeEntry.date >= start_date)
        except Exception:
            pass
    if end_date_s:
        try:
            end_date = datetime.strptime(end_date_s, '%Y-%m-%d').date()
            query = query.filter(TimeEntry.date <= end_date)
        except Exception:
            pass

    # Apply user filter (PMP/Admin only)
    if filter_user_id:
        try:
            uid = int(filter_user_id)
            query = query.filter(TimeEntry.user_id == uid)
        except Exception:
            pass

    total_count = query.count()
    total_pages = (total_count + per_page - 1) // per_page

    entries = query.order_by(TimeEntry.date.desc()).offset((page - 1) * per_page).limit(per_page).all()

    # Total hours visible in this filtered view
    total_hours_week = query.with_entities(func.sum(TimeEntry.hours)).filter(TimeEntry.date >= (datetime.now().date() - timedelta(days=7))).scalar() or 0

    # If PMP/Admin, expose list of users for filter (only users that have time entries within the date filters)
    users = []
    if user_role in ['PMP', 'Admin']:
        user_ids_q = db.session.query(TimeEntry.user_id.distinct())
        if start_date_s:
            try:
                s_date = datetime.strptime(start_date_s, '%Y-%m-%d').date()
                user_ids_q = user_ids_q.filter(TimeEntry.date >= s_date)
            except Exception:
                pass
        if end_date_s:
            try:
                e_date = datetime.strptime(end_date_s, '%Y-%m-%d').date()
                user_ids_q = user_ids_q.filter(TimeEntry.date <= e_date)
            except Exception:
                pass
        user_ids = [r[0] for r in user_ids_q.all()]
        if user_ids:
            users = User.query.filter(User.id.in_(user_ids)).order_by(User.first_name).all()
        else:
            users = []

    # Build pagination urls (preserve filters)
    params = request.args.to_dict()
    def make_url_for(p):
        params_copy = params.copy()
        params_copy['page'] = p
        return url_for('main.time_entries', **params_copy)

    page_urls = [{'num': i, 'url': make_url_for(i), 'active': (i == page)} for i in range(1, total_pages + 1)] if total_pages > 0 else []
    prev_url = make_url_for(page - 1) if page > 1 else None
    next_url = make_url_for(page + 1) if page < total_pages else None

    return render_template('time_entries.html', time_entries=entries, total_hours_week=total_hours_week, page=page, total_pages=total_pages, users=users, total_count=total_count, page_urls=page_urls, prev_url=prev_url, next_url=next_url)

@main_bp.route('/team')
@login_required
def team():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
    # Obtener usuarios internos
    internal_users = User.query.filter_by(is_internal=True, is_active=True).all()
    
    # Estadísticas de tareas por estado para cada usuario interno
    for u in internal_users:
        # Contar tareas por estado
        task_stats = db.session.query(
            Task.status, func.count(Task.id)
        ).filter(Task.assigned_to_id == u.id).group_by(Task.status).all()
        
        u.task_by_status = {status: count for status, count in task_stats}
        u.total_tasks = sum(u.task_by_status.values())
        
        # Horas este mes
        u.hours_this_month = db.session.query(func.sum(TimeEntry.hours)).join(Task).filter(
            TimeEntry.user_id == u.id,
            TimeEntry.date >= (datetime.now().date().replace(day=1))
        ).scalar() or 0
    
    return render_template('team.html', internal_users=internal_users)


@main_bp.route('/workload')
@login_required
def workload():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))

    from datetime import date as _date, timedelta as _td

    # Filters
    project_id = request.args.get('project_id', type=int)
    capacity = request.args.get('capacity', default=40, type=int)  # weekly hours capacity

    # Date ranges
    today = _date.today()
    week_start = today - _td(days=today.weekday())   # Monday
    week_end   = week_start + _td(days=6)
    month_start = today.replace(day=1)

    # All active projects for filter dropdown
    projects = Project.query.filter(Project.status != 'ARCHIVED').order_by(Project.name).all()

    # Internal active users
    members = User.query.filter_by(is_internal=True, is_active=True).order_by(User.first_name).all()

    ACTIVE_STATUSES = ('BACKLOG', 'IN_PROGRESS', 'IN_REVIEW')
    DONE_STATUSES   = ('COMPLETED', 'DONE', 'ACCEPTED')

    workload_data = []
    for u in members:
        # Base query for tasks assigned to this user (via many-to-many OR legacy field)
        from sqlalchemy import or_, and_
        from ..models import task_assignees as _ta

        task_q = Task.query.filter(
            or_(
                Task.assigned_to_id == u.id,
                Task.id.in_(
                    db.session.query(_ta.c.task_id).filter(_ta.c.user_id == u.id)
                )
            )
        )
        if project_id:
            task_q = task_q.filter(Task.project_id == project_id)

        all_tasks = task_q.all()

        active_tasks = [t for t in all_tasks if t.status in ACTIVE_STATUSES]
        done_tasks   = [t for t in all_tasks if t.status in DONE_STATUSES]

        # Estimated hours pending (active tasks only)
        est_hours = sum(float(t.estimated_hours or 0) for t in active_tasks)

        # Hours logged this week
        hours_week_q = db.session.query(func.sum(TimeEntry.hours)).filter(
            TimeEntry.user_id == u.id,
            TimeEntry.date >= week_start,
            TimeEntry.date <= week_end,
        )
        if project_id:
            hours_week_q = hours_week_q.join(Task).filter(Task.project_id == project_id)
        hours_week = float(hours_week_q.scalar() or 0)

        # Hours logged this month
        hours_month_q = db.session.query(func.sum(TimeEntry.hours)).filter(
            TimeEntry.user_id == u.id,
            TimeEntry.date >= month_start,
        )
        if project_id:
            hours_month_q = hours_month_q.join(Task).filter(Task.project_id == project_id)
        hours_month = float(hours_month_q.scalar() or 0)

        # Status breakdown (active)
        status_counts = {}
        for s in ACTIVE_STATUSES:
            status_counts[s] = sum(1 for t in active_tasks if t.status == s)

        # Workload % relative to capacity
        workload_pct = min(round(est_hours / capacity * 100), 150) if capacity > 0 else 0

        # Overdue tasks
        overdue = [t for t in active_tasks if t.due_date and t.due_date.date() < today]

        # Sort active tasks: overdue first, then by due_date
        active_tasks_sorted = sorted(
            active_tasks,
            key=lambda t: (t.due_date is None, t.due_date or today)
        )

        workload_data.append({
            'user': u,
            'active_tasks': active_tasks_sorted,
            'active_count': len(active_tasks),
            'done_count': len(done_tasks),
            'est_hours': round(est_hours, 1),
            'hours_week': round(hours_week, 1),
            'hours_month': round(hours_month, 1),
            'status_counts': status_counts,
            'workload_pct': workload_pct,
            'overdue_count': len(overdue),
            'is_overloaded': est_hours > capacity,
        })

    # Sort: overloaded first, then by est_hours desc
    workload_data.sort(key=lambda x: (-x['is_overloaded'], -x['est_hours']))

    return render_template(
        'workload.html',
        workload_data=workload_data,
        projects=projects,
        selected_project_id=project_id,
        capacity=capacity,
        week_start=week_start,
        week_end=week_end,
        today=today,
    )


@main_bp.route('/budget')
@login_required
def budget_tracking():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))

    enabled = SystemSettings.get('budget_tracking_enabled', 'true')
    if isinstance(enabled, str) and enabled.lower() in ('false', '0', 'no'):
        abort(404)

    from datetime import date as _date

    # Settings-based thresholds
    warn_threshold   = int(SystemSettings.get('budget_warn_threshold',   80))
    danger_threshold = int(SystemSettings.get('budget_danger_threshold', 100))
    currency         = SystemSettings.get('budget_currency', 'USD')

    # Default hourly rate
    default_rate_obj = HourlyRate.get_default()
    default_rate     = float(default_rate_obj.rate) if default_rate_obj else None

    # Filters
    status_filter  = request.args.get('status', '')
    manager_filter = request.args.get('manager_id', type=int)

    q = Project.query
    if status_filter:
        q = q.filter(Project.status == status_filter)
    else:
        q = q.filter(Project.status != 'ARCHIVED')
    if manager_filter:
        q = q.filter(Project.manager_id == manager_filter)
    projects = q.order_by(Project.name).all()

    # Managers for filter dropdown
    managers = User.query.filter_by(is_internal=True, is_active=True).order_by(User.first_name).all()

    today = _date.today()

    budget_rows = []
    totals = {'budget_h': 0.0, 'logged_h': 0.0, 'est_cost': 0.0, 'logged_cost': 0.0, 'projects_over': 0}

    for p in projects:
        budget_h = float(p.budget_hours) if p.budget_hours else None

        logged_h = float(
            db.session.query(func.sum(TimeEntry.hours))
            .join(Task)
            .filter(Task.project_id == p.id)
            .scalar() or 0
        )

        est_h = float(
            db.session.query(func.sum(Task.estimated_hours))
            .filter(Task.project_id == p.id)
            .scalar() or 0
        )

        total_tasks     = db.session.query(func.count(Task.id)).filter(Task.project_id == p.id).scalar() or 0
        completed_tasks = db.session.query(func.count(Task.id)).filter(
            Task.project_id == p.id,
            Task.status.in_(['COMPLETED', 'DONE', 'ACCEPTED'])
        ).scalar() or 0
        task_pct = round(completed_tasks / total_tasks * 100) if total_tasks > 0 else 0

        if budget_h and budget_h > 0:
            pct = round(logged_h / budget_h * 100, 1)
        elif est_h > 0:
            pct = round(logged_h / est_h * 100, 1)
        else:
            pct = None

        if pct is None:
            budget_status = 'nodata'
        elif pct >= danger_threshold:
            budget_status = 'danger'
        elif pct >= warn_threshold:
            budget_status = 'warning'
        else:
            budget_status = 'ok'

        est_cost    = round(budget_h    * default_rate, 2) if budget_h    and default_rate else None
        logged_cost = round(logged_h    * default_rate, 2) if default_rate else None
        remaining_h = round((budget_h - logged_h), 1)      if budget_h    else None

        if budget_h:
            totals['budget_h']  += budget_h
        totals['logged_h']      += logged_h
        if est_cost:
            totals['est_cost']  += est_cost
        if logged_cost:
            totals['logged_cost'] += logged_cost
        if budget_status == 'danger':
            totals['projects_over'] += 1

        budget_rows.append({
            'project':       p,
            'budget_h':      budget_h,
            'logged_h':      round(logged_h, 1),
            'est_h':         round(est_h, 1),
            'pct':           pct,
            'remaining_h':   remaining_h,
            'est_cost':      est_cost,
            'logged_cost':   round(logged_cost, 2) if logged_cost is not None else None,
            'budget_status': budget_status,
            'task_pct':      task_pct,
            'total_tasks':   total_tasks,
            'completed_tasks': completed_tasks,
        })

    # Sort: danger first, then warning, then ok, then nodata
    order = {'danger': 0, 'warning': 1, 'ok': 2, 'nodata': 3}
    budget_rows.sort(key=lambda r: (order.get(r['budget_status'], 4), r['project'].name))

    totals['budget_h']   = round(totals['budget_h'], 1)
    totals['logged_h']   = round(totals['logged_h'], 1)
    totals['est_cost']   = round(totals['est_cost'], 2)
    totals['logged_cost']= round(totals['logged_cost'], 2)

    return render_template(
        'budget_tracking.html',
        budget_rows=budget_rows,
        totals=totals,
        warn_threshold=warn_threshold,
        danger_threshold=danger_threshold,
        currency=currency,
        default_rate=default_rate,
        managers=managers,
        status_filter=status_filter,
        manager_filter=manager_filter,
    )


# ── Portfolio Dashboard ────────────────────────────────────────────────────────

@main_bp.route('/portfolio')
@login_required
def portfolio():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
    _v = SystemSettings.get('portfolio_enabled', 'true')
    if isinstance(_v, str) and _v.lower() in ('false', '0', 'no'):
        abort(404)

    from datetime import date as _date

    today = _date.today()

    # Thresholds from settings
    warn_threshold   = int(SystemSettings.get('budget_warn_threshold',   80))
    danger_threshold = int(SystemSettings.get('budget_danger_threshold', 100))
    currency         = SystemSettings.get('budget_currency', 'USD')
    default_rate_obj = HourlyRate.get_default()
    default_rate     = float(default_rate_obj.rate) if default_rate_obj else None
    risks_enabled    = SystemSettings.get('risks_enabled', 'true')
    risks_enabled    = not (isinstance(risks_enabled, str) and risks_enabled.lower() in ('false','0','no'))

    # Filters
    status_filter  = request.args.get('status', 'ACTIVE')
    manager_filter = request.args.get('manager_id', type=int)
    rag_filter     = request.args.get('rag', '')

    q = Project.query
    if status_filter:
        q = q.filter(Project.status == status_filter)
    else:
        q = q.filter(Project.status != 'ARCHIVED')
    if manager_filter:
        q = q.filter(Project.manager_id == manager_filter)
    projects = q.order_by(Project.name).all()

    managers = User.query.filter_by(is_internal=True, is_active=True).order_by(User.first_name).all()

    portfolio_rows = []
    rag_counts = {'green': 0, 'yellow': 0, 'red': 0}

    for p in projects:
        # Task progress
        total_tasks = db.session.query(func.count(Task.id)).filter(Task.project_id == p.id).scalar() or 0
        done_tasks  = db.session.query(func.count(Task.id)).filter(
            Task.project_id == p.id,
            Task.status.in_(['COMPLETED', 'DONE', 'ACCEPTED'])
        ).scalar() or 0
        task_pct = round(done_tasks / total_tasks * 100) if total_tasks > 0 else 0

        # Budget
        logged_h = float(db.session.query(func.sum(TimeEntry.hours)).join(Task).filter(Task.project_id == p.id).scalar() or 0)
        budget_h = float(p.budget_hours) if p.budget_hours else None
        budget_pct = round(logged_h / budget_h * 100, 1) if budget_h and budget_h > 0 else None
        logged_cost = round(logged_h * default_rate, 2) if default_rate else None

        # Risks
        open_risks     = p.risks.filter_by(status='OPEN').count() if risks_enabled else 0
        critical_risks = p.risks.filter(ProjectRisk.status == 'OPEN', ProjectRisk.severity == 'CRITICAL').count() if risks_enabled else 0
        high_risks     = p.risks.filter(ProjectRisk.status == 'OPEN', ProjectRisk.severity == 'HIGH').count() if risks_enabled else 0

        # Team size (unique assignees across tasks)
        from sqlalchemy import distinct
        from ..models import task_assignees as _ta
        team_size = db.session.query(func.count(distinct(_ta.c.user_id))).join(
            Task, Task.id == _ta.c.task_id
        ).filter(Task.project_id == p.id).scalar() or 0

        # Days remaining / overdue
        days_remaining = None
        is_overdue = False
        if p.end_date:
            delta = (p.end_date.date() if hasattr(p.end_date, 'date') else p.end_date) - today
            days_remaining = delta.days
            is_overdue = days_remaining < 0 and p.status not in ('COMPLETED', 'ARCHIVED')

        # RAG calculation
        if (critical_risks > 0
                or (budget_pct is not None and budget_pct >= danger_threshold)
                or is_overdue):
            rag = 'red'
        elif (high_risks > 0
                or (budget_pct is not None and budget_pct >= warn_threshold)
                or (days_remaining is not None and 0 <= days_remaining <= 7)
                or (total_tasks > 0 and task_pct < 30 and p.status == 'ACTIVE' and days_remaining is not None and days_remaining < 30)):
            rag = 'yellow'
        else:
            rag = 'green'

        rag_counts[rag] += 1

        portfolio_rows.append({
            'project':        p,
            'task_pct':       task_pct,
            'total_tasks':    total_tasks,
            'done_tasks':     done_tasks,
            'logged_h':       round(logged_h, 1),
            'budget_h':       budget_h,
            'budget_pct':     budget_pct,
            'logged_cost':    logged_cost,
            'open_risks':     open_risks,
            'critical_risks': critical_risks,
            'days_remaining': days_remaining,
            'is_overdue':     is_overdue,
            'team_size':      team_size,
            'rag':            rag,
        })

    # Apply RAG filter after calculation
    if rag_filter:
        portfolio_rows = [r for r in portfolio_rows if r['rag'] == rag_filter]

    # Sort: red → yellow → green, then by name
    rag_order = {'red': 0, 'yellow': 1, 'green': 2}
    portfolio_rows.sort(key=lambda r: (rag_order[r['rag']], r['project'].name))

    # Global KPIs
    all_projects = Project.query.filter(Project.status != 'ARCHIVED').all()
    kpis = {
        'total':     len(all_projects),
        'active':    sum(1 for p in all_projects if p.status == 'ACTIVE'),
        'completed': sum(1 for p in all_projects if p.status == 'COMPLETED'),
        'planning':  sum(1 for p in all_projects if p.status == 'PLANNING'),
    }

    return render_template(
        'portfolio.html',
        portfolio_rows=portfolio_rows,
        managers=managers,
        status_filter=status_filter,
        manager_filter=manager_filter,
        rag_filter=rag_filter,
        rag_counts=rag_counts,
        kpis=kpis,
        currency=currency,
        default_rate=default_rate,
        risks_enabled=risks_enabled,
        warn_threshold=warn_threshold,
        danger_threshold=danger_threshold,
        today=today,
    )

# ── Risk / Issue Register ─────────────────────────────────────────────────────

@main_bp.route('/risks')
@login_required
def risks_global():
    """Vista global de riesgos e issues de todos los proyectos."""
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
    _v = SystemSettings.get('risks_enabled', 'true')
    if isinstance(_v, str) and _v.lower() in ('false', '0', 'no'):
        abort(404)

    type_filter     = request.args.get('type', '')
    status_filter   = request.args.get('status', '')
    severity_filter = request.args.get('severity', '')
    project_filter  = request.args.get('project_id', type=int)

    q = ProjectRisk.query.join(Project)
    if type_filter:
        q = q.filter(ProjectRisk.type == type_filter)
    if status_filter:
        q = q.filter(ProjectRisk.status == status_filter)
    if severity_filter:
        q = q.filter(ProjectRisk.severity == severity_filter)
    if project_filter:
        q = q.filter(ProjectRisk.project_id == project_filter)

    risks = q.order_by(ProjectRisk.status, ProjectRisk.updated_at.desc()).all()

    projects  = Project.query.filter(Project.status != 'ARCHIVED').order_by(Project.name).all()
    members   = User.query.filter_by(is_internal=True, is_active=True).order_by(User.first_name).all()

    counts = {
        'open':     sum(1 for r in risks if r.status == 'OPEN'),
        'risk':     sum(1 for r in risks if r.type == 'RISK'),
        'issue':    sum(1 for r in risks if r.type == 'ISSUE'),
        'critical': sum(1 for r in risks if r.severity == 'CRITICAL' and r.status == 'OPEN'),
    }

    return render_template('risks.html',
        risks=risks,
        projects=projects,
        members=members,
        counts=counts,
        type_filter=type_filter,
        status_filter=status_filter,
        severity_filter=severity_filter,
        project_filter=project_filter,
        project=None,
    )


@main_bp.route('/project/<int:project_id>/risks')
@login_required
def project_risks(project_id):
    """Vista de riesgos e issues de un proyecto específico."""
    _v = SystemSettings.get('risks_enabled', 'true')
    if isinstance(_v, str) and _v.lower() in ('false', '0', 'no'):
        abort(404)
    p = Project.query.get_or_404(project_id)

    type_filter     = request.args.get('type', '')
    status_filter   = request.args.get('status', '')

    q = ProjectRisk.query.filter_by(project_id=project_id)
    if type_filter:
        q = q.filter(ProjectRisk.type == type_filter)
    if status_filter:
        q = q.filter(ProjectRisk.status == status_filter)

    risks   = q.order_by(ProjectRisk.status, ProjectRisk.updated_at.desc()).all()
    members = User.query.filter_by(is_internal=True, is_active=True).order_by(User.first_name).all()

    counts = {
        'open':     sum(1 for r in risks if r.status == 'OPEN'),
        'risk':     sum(1 for r in risks if r.type == 'RISK'),
        'issue':    sum(1 for r in risks if r.type == 'ISSUE'),
        'critical': sum(1 for r in risks if r.severity == 'CRITICAL' and r.status == 'OPEN'),
    }

    return render_template('risks.html',
        risks=risks,
        projects=[],
        members=members,
        counts=counts,
        type_filter=type_filter,
        status_filter=status_filter,
        severity_filter='',
        project_filter=project_id,
        project=p,
    )


# ── Risk API ───────────────────────────────────────────────────────────────────

@main_bp.route('/api/projects/<int:project_id>/risks', methods=['POST'])
@login_required
def create_risk(project_id):
    Project.query.get_or_404(project_id)
    data = request.get_json(force=True, silent=True) or {}

    if not data.get('title', '').strip():
        return jsonify({'error': 'El título es requerido'}), 400

    risk = ProjectRisk(
        project_id      = project_id,
        type            = data.get('type', 'RISK'),
        title           = data['title'].strip(),
        description     = data.get('description', ''),
        severity        = data.get('severity', 'MEDIUM'),
        probability     = data.get('probability', 'MEDIUM'),
        status          = data.get('status', 'OPEN'),
        mitigation_plan = data.get('mitigation_plan', ''),
        owner_id        = data.get('owner_id') or None,
        created_by_id   = current_user.id,
    )
    db.session.add(risk)
    db.session.commit()

    # Notificar al responsable asignado
    if risk.owner_id and risk.owner_id != current_user.id:
        try:
            from app.services.notifications import NotificationService
            type_label = 'Riesgo' if risk.type == 'RISK' else 'Issue'
            sev_label  = {'LOW':'Baja','MEDIUM':'Media','HIGH':'Alta','CRITICAL':'Crítica'}.get(risk.severity, risk.severity)
            risks_url  = url_for('main.project_risks', project_id=project_id, _external=True)
            NotificationService.notify(
                user_id=risk.owner_id,
                title=f'Se te asignó un {type_label}: {risk.title}',
                message=f'{current_user.name} te asignó el {type_label.lower()} "{risk.title}" '
                        f'(Severidad: {sev_label}) en el proyecto {risk.project.name}.',
                notification_type='general',
                related_entity_type='ProjectRisk',
                related_entity_id=risk.id,
                send_email=True,
                email_context={
                    'title':       f'Se te asignó un {type_label}',
                    'message':     f'{current_user.name} te asignó el {type_label.lower()} "{risk.title}" '
                                   f'(Severidad: {sev_label}) en el proyecto {risk.project.name}.',
                    'task_url':    risks_url,
                    'dashboard_url': risks_url,
                },
            )
        except Exception as e:
            current_app.logger.warning('Error enviando notificación de riesgo: %s', e)

    return jsonify(_risk_to_dict(risk)), 201


@main_bp.route('/api/projects/<int:project_id>/risks/<int:risk_id>', methods=['PATCH'])
@login_required
def update_risk(project_id, risk_id):
    risk = ProjectRisk.query.filter_by(id=risk_id, project_id=project_id).first_or_404()
    data = request.get_json(force=True, silent=True) or {}

    prev_owner_id = risk.owner_id
    for field in ('type', 'title', 'description', 'severity', 'probability',
                  'status', 'mitigation_plan'):
        if field in data:
            setattr(risk, field, data[field])
    if 'owner_id' in data:
        risk.owner_id = data['owner_id'] or None

    db.session.commit()

    # Notificar si el responsable cambió y es distinto al editor
    new_owner_id = risk.owner_id
    if new_owner_id and new_owner_id != prev_owner_id and new_owner_id != current_user.id:
        try:
            from app.services.notifications import NotificationService
            type_label = 'Riesgo' if risk.type == 'RISK' else 'Issue'
            sev_label  = {'LOW':'Baja','MEDIUM':'Media','HIGH':'Alta','CRITICAL':'Crítica'}.get(risk.severity, risk.severity)
            risks_url  = url_for('main.project_risks', project_id=project_id, _external=True)
            NotificationService.notify(
                user_id=new_owner_id,
                title=f'Se te asignó un {type_label}: {risk.title}',
                message=f'{current_user.name} te asignó el {type_label.lower()} "{risk.title}" '
                        f'(Severidad: {sev_label}) en el proyecto {risk.project.name}.',
                notification_type='general',
                related_entity_type='ProjectRisk',
                related_entity_id=risk.id,
                send_email=True,
                email_context={
                    'title':       f'Se te asignó un {type_label}',
                    'message':     f'{current_user.name} te asignó el {type_label.lower()} "{risk.title}" '
                                   f'(Severidad: {sev_label}) en el proyecto {risk.project.name}.',
                    'task_url':    risks_url,
                    'dashboard_url': risks_url,
                },
            )
        except Exception as e:
            current_app.logger.warning('Error enviando notificación de riesgo: %s', e)

    return jsonify(_risk_to_dict(risk))


@main_bp.route('/api/projects/<int:project_id>/risks/<int:risk_id>', methods=['DELETE'])
@login_required
def delete_risk(project_id, risk_id):
    user_role = current_user.role.name if current_user.role else ''
    if not (current_user.is_internal and user_role in ('PMP', 'Admin')):
        return jsonify({'error': 'Sin permisos'}), 403
    risk = ProjectRisk.query.filter_by(id=risk_id, project_id=project_id).first_or_404()
    db.session.delete(risk)
    db.session.commit()
    return jsonify({'deleted': risk_id})


def _risk_to_dict(r):
    return {
        'id':              r.id,
        'project_id':      r.project_id,
        'type':            r.type,
        'title':           r.title,
        'description':     r.description,
        'severity':        r.severity,
        'probability':     r.probability,
        'status':          r.status,
        'mitigation_plan': r.mitigation_plan,
        'owner_id':        r.owner_id,
        'owner_name':      r.owner.name if r.owner else None,
        'created_at':      r.created_at.isoformat() if r.created_at else None,
        'updated_at':      r.updated_at.isoformat() if r.updated_at else None,
    }

# ──────────────────────────────────────────────────────────────────────────────


@main_bp.route('/clients')
@login_required
def clients():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
    # Obtener clientes
    clients = User.query.filter_by(is_internal=False).all()
    
    # Estadísticas de proyectos por estado para cada cliente
    for c in clients:
        # Proyectos asociados al cliente a través de la tabla project_clients
        project_stats = db.session.query(
            Project.status, func.count(Project.id)
        ).join(project_clients).filter(
            project_clients.c.user_id == c.id
        ).group_by(Project.status).all()
        
        c.project_by_status = {status: count for status, count in project_stats}
        c.total_projects = sum(c.project_by_status.values())
        
        # Tareas pendientes de aprobación
        c.pending_approvals = db.session.query(func.count(Task.id)).join(Project).join(
            project_clients
        ).filter(
            project_clients.c.user_id == c.id,
            Task.requires_approval == True,
            Task.approval_status == 'pending'
        ).scalar() or 0
    
    return render_template('clients.html', clients=clients)


@main_bp.route('/clients/create', methods=['POST'])
@login_required
def create_client():
    """Create a new client user"""
    if not _ensure_pmp():
        return redirect(url_for('main.clients'))
    
    from werkzeug.security import generate_password_hash
    
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '')
    first_name = request.form.get('first_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    company = request.form.get('company', '').strip()
    phone = request.form.get('phone', '').strip()
    
    # Validation
    if not email:
        flash('El email es requerido.', 'danger')
        return redirect(url_for('main.clients'))
    
    if User.query.filter_by(email=email).first():
        flash('Ya existe un usuario con ese email.', 'danger')
        return redirect(url_for('main.clients'))
    
    if not password or len(password) < 6:
        flash('La contraseña debe tener al menos 6 caracteres.', 'danger')
        return redirect(url_for('main.clients'))
    
    # Get or create client role
    client_role = Role.query.filter_by(name='Cliente').first()
    if not client_role:
        # Create the role if it doesn't exist
        client_role = Role(name='Cliente')
        db.session.add(client_role)
        db.session.flush()
    
    client = User(
        email=email,
        first_name=first_name or None,
        last_name=last_name or None,
        company=company or None,
        phone=phone or None,
        is_internal=False,
        is_active=True,
        role_id=client_role.id
    )
    client.set_password(password)
    
    db.session.add(client)
    db.session.commit()
    
    flash(f'Cliente {email} creado exitosamente.', 'success')
    return redirect(url_for('main.clients'))


@main_bp.route('/clients/update', methods=['POST'])
@login_required
def update_client():
    """Update an existing client user"""
    if not _ensure_pmp():
        return redirect(url_for('main.clients'))
    
    from werkzeug.security import generate_password_hash
    
    client_id = request.form.get('client_id')
    client = User.query.get_or_404(client_id)
    
    # Only allow editing non-internal users (clients)
    if client.is_internal:
        flash('No puedes editar usuarios internos desde esta página.', 'danger')
        return redirect(url_for('main.clients'))
    
    # Azure users can only have company/phone updated
    is_azure = bool(client.azure_oid)
    
    if is_azure:
        # Only update additional info for Azure users
        client.company = request.form.get('company', '').strip() or None
        client.phone = request.form.get('phone', '').strip() or None
    else:
        # Local users - all fields can be updated
        email = request.form.get('email', '').strip()
        if email and email != client.email:
            if User.query.filter_by(email=email).first():
                flash('Ya existe un usuario con ese email.', 'danger')
                return redirect(url_for('main.clients'))
            client.email = email
        
        client.first_name = request.form.get('first_name', '').strip() or None
        client.last_name = request.form.get('last_name', '').strip() or None
        client.company = request.form.get('company', '').strip() or None
        client.phone = request.form.get('phone', '').strip() or None
        
        # Update password only if provided
        password = request.form.get('password', '')
        if password:
            if len(password) < 6:
                flash('La contraseña debe tener al menos 6 caracteres.', 'danger')
                return redirect(url_for('main.clients'))
            client.set_password(password)
    
    db.session.commit()
    flash(f'Cliente {client.email} actualizado.', 'success')
    return redirect(url_for('main.clients'))


@main_bp.route('/clients/delete', methods=['POST'])
@login_required
def delete_client():
    """Delete a client user (only if they have no projects)"""
    if not _ensure_pmp():
        return redirect(url_for('main.clients'))
    
    client_id = request.form.get('client_id')
    client = User.query.get_or_404(client_id)
    
    # Only allow deleting non-internal users (clients)
    if client.is_internal:
        flash('No puedes eliminar usuarios internos desde esta página.', 'danger')
        return redirect(url_for('main.clients'))
    
    # Check if client has associated projects
    has_projects = db.session.query(project_clients).filter(
        project_clients.c.user_id == client.id
    ).first()
    
    if has_projects:
        flash('No se puede eliminar el cliente porque tiene proyectos asociados.', 'danger')
        return redirect(url_for('main.clients'))
    
    email = client.email
    db.session.delete(client)
    db.session.commit()
    
    flash(f'Cliente {email} eliminado.', 'success')
    return redirect(url_for('main.clients'))


@main_bp.route('/teams')
@login_required
def teams_alias():
    # Alias por compatibilidad: /teams -> /team
    return redirect(url_for('main.team'))

@main_bp.route('/reports')
@login_required
def reports():
    if not _ensure_pmp():
        return redirect(url_for('main.index'))
    # Reportes generales
    total_projects = db.session.query(func.count(Project.id)).scalar()
    total_tasks = db.session.query(func.count(Task.id)).scalar()
    completed_tasks = db.session.query(func.count(Task.id)).filter(Task.status == 'COMPLETED').scalar()
    
    # Budget overview
    projects_budget = db.session.query(
        func.sum(Project.budget_hours),
        func.sum(func.coalesce(
            db.session.query(func.sum(TimeEntry.hours))
            .join(Task)
            .filter(Task.project_id == Project.id)
            .correlate(Project)
            .as_scalar(), 0
        ))
    ).all()
    
    total_budget = projects_budget[0][0] or 0
    total_hours_spent = projects_budget[0][1] or 0
    budget_usage_percent = (total_hours_spent / total_budget * 100) if total_budget > 0 else 0
    
    # Hours by user (last 30 days)
    thirty_days_ago = datetime.now().date() - timedelta(days=30)
    user_hours = db.session.query(
        User.email,
        func.sum(TimeEntry.hours).label('total_hours')
    ).join(TimeEntry).filter(
        TimeEntry.date >= thirty_days_ago
    ).group_by(User.id, User.email).all()

    # Projects list for selector
    projects = Project.query.order_by(Project.name).all()

    # Optional: build project-specific summary rows if a project is selected
    project = None
    task_rows = None
    project_id = request.args.get('project_id')
    if project_id:
        project = Project.query.get(project_id)
        if not project:
            abort(404)

        # Paginated task query for project summary (10 items per page)
        tasks_query = Task.query.filter_by(project_id=project.id).order_by(Task.id)
        project_total_tasks = tasks_query.count()

        # Pagination params
        try:
            page = max(1, int(request.args.get('page', 1)))
        except Exception:
            page = 1
        per_page = 10

        # Load tasks: all if exporting, otherwise page slice
        if request.args.get('export') == 'xlsx':
            tasks = tasks_query.all()
        else:
            tasks = tasks_query.limit(per_page).offset((page - 1) * per_page).all()

        task_rows = []
        today = datetime.now().date()
        for t in tasks:
            assignees = [u.name for u in t.assignee_list]
            # Use first_name when available, otherwise fall back to name or email
            client_name = ''
            if t.assigned_client:
                client_name = t.assigned_client.first_name or getattr(t.assigned_client, 'name', None) or t.assigned_client.email
            hours_logged = db.session.query(func.coalesce(func.sum(TimeEntry.hours), 0)).filter(TimeEntry.task_id == t.id).scalar() or 0

            # Calculate days overdue (Option 1): for non-completed tasks only
            if t.due_date:
                due_val = t.due_date.date() if hasattr(t.due_date, 'date') else t.due_date
                if t.status in ('COMPLETED', 'ACCEPTED'):
                    # Calcular atraso real al momento de completar (usar completed_at)
                    if t.completed_at:
                        comp_val = t.completed_at.date() if hasattr(t.completed_at, 'date') else t.completed_at
                        days_overdue = (comp_val - due_val).days  # Puede ser negativo (anticipación)
                    else:
                        # Si no tiene completed_at, usar hoy como referencia
                        days_overdue = (today - due_val).days
                else:
                    # Tarea no completada: calcular desde hoy
                    days_overdue = max(0, (today - due_val).days)
            else:
                days_overdue = 0

            task_rows.append({
                'id': t.id,
                'title': t.title,
                'status': t.status,
                'priority': t.priority,
                'assignees': assignees,
                'client': client_name,
                'start_date': t.start_date.date() if t.start_date else None,
                'due_date': t.due_date.date() if t.due_date else None,
                'completed_at': t.completed_at.date() if t.completed_at else None,
                'estimated_hours': t.estimated_hours,
                'hours_logged': float(hours_logged),
                'days_overdue': days_overdue
            })

        # Compute pagination metadata for template (only when not exporting)
        pagination = None
        if request.args.get('export') != 'xlsx':
            total_pages = (project_total_tasks + per_page - 1) // per_page if project_total_tasks else 1
            pagination = {'page': page, 'per_page': per_page, 'total': project_total_tasks, 'pages': total_pages}

        # Override KPI cards with project-scoped metrics when a project is selected
        try:
            project_completed_tasks = sum(1 for t in tasks_query if (t.status or '').upper() == 'COMPLETED')
            project_total_budget = project.budget_hours or 0
            project_total_hours_spent = db.session.query(func.coalesce(func.sum(TimeEntry.hours), 0)).join(Task).filter(Task.project_id == project.id).scalar() or 0

            # % usage
            project_budget_usage = (float(project_total_hours_spent) / float(project_total_budget) * 100) if project_total_budget and float(project_total_budget) > 0 else 0

            # Hours by user limited to this project
            project_user_hours = db.session.query(
                User.email,
                func.sum(TimeEntry.hours).label('total_hours')
            ).join(TimeEntry).join(Task).filter(
                Task.project_id == project.id,
                TimeEntry.date >= thirty_days_ago
            ).group_by(User.id, User.email).all()

            # Override variables used in template
            total_projects = 1
            total_tasks = project_total_tasks
            completed_tasks = project_completed_tasks
            total_budget = project_total_budget
            total_hours_spent = project_total_hours_spent
            budget_usage_percent = project_budget_usage
            user_hours = project_user_hours
        except Exception:
            # If anything goes wrong, fall back to global values already computed
            pass

        if request.args.get('export') == 'csv':
            import csv, io
            bio = BytesIO()
            bio.write(b'\xef\xbb\xbf')  # UTF-8 BOM for Excel compatibility
            text_wrapper = io.TextIOWrapper(bio, encoding='utf-8', newline='')
            writer = csv.writer(text_wrapper)
            writer.writerow(['Task ID', 'Título', 'Estado', 'Prioridad', 'Responsables', 'Cliente Externo', 'Fecha Inicio', 'Fecha Vencimiento', 'Completada', 'Atraso (días)', 'Horas Estimadas', 'Horas Registradas'])
            for row in task_rows:
                writer.writerow([
                    row['id'],
                    row['title'],
                    row['status'],
                    row['priority'],
                    ', '.join(row['assignees']) if row['assignees'] else '',
                    row.get('client', ''),
                    row['start_date'].isoformat() if row['start_date'] else '',
                    row['due_date'].isoformat() if row['due_date'] else '',
                    row['completed_at'].isoformat() if row.get('completed_at') else '',
                    row.get('days_overdue', 0),
                    float(row['estimated_hours']) if row['estimated_hours'] is not None else '',
                    row['hours_logged'],
                ])
            text_wrapper.detach()
            bio.seek(0)
            return send_file(bio, mimetype='text/csv; charset=utf-8', as_attachment=True, download_name=f"project_{project.id}_summary.csv")

        if request.args.get('export') == 'xlsx':
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter
            except Exception:
                abort(500, 'openpyxl not installed')

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Project-{project.id}-Summary"

            headers = ['Task ID', 'Title', 'Status', 'Priority', 'Assignees', 'Assigned Client', 'Start Date', 'Due Date', 'Completed At', 'Atraso (días)', 'Estimated Hours', 'Hours Logged']
            ws.append(headers)
            for row in task_rows:
                ws.append([
                    row['id'],
                    row['title'],
                    row['status'],
                    row['priority'],
                    ', '.join(row['assignees']) if row['assignees'] else '',
                    (row['client'] + ' (Cliente Externo)') if row.get('client') else '',
                    row['start_date'].isoformat() if row['start_date'] else '',
                    row['due_date'].isoformat() if row['due_date'] else '',
                    row['completed_at'].isoformat() if row.get('completed_at') else '',
                    row.get('days_overdue', 0),
                    float(row['estimated_hours']) if row['estimated_hours'] is not None else '',
                    row['hours_logged']
                ])

            # Auto width
            for i, column_cells in enumerate(ws.columns, 1):
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws.column_dimensions[get_column_letter(i)].width = min(max(length + 4, 12), 60) # Fix 1: Ajuste más generoso

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            filename = f"project_{project.id}_summary.xlsx"
            return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

        # Otherwise render the report section with task_rows
        return render_template('reports.html',
            total_projects=total_projects,
            total_tasks=total_tasks,
            completed_tasks=completed_tasks,
            total_budget=total_budget,
            total_hours_spent=total_hours_spent,
            budget_usage_percent=budget_usage_percent,
            user_hours=user_hours,
            projects=projects,
            selected_project=project,
            task_rows=task_rows,
            pagination=pagination
        )

    return render_template('reports.html',
        total_projects=total_projects,
        total_tasks=total_tasks,
        completed_tasks=completed_tasks,
        total_budget=total_budget,
        total_hours_spent=total_hours_spent,
        budget_usage_percent=budget_usage_percent,
        user_hours=user_hours,
        projects=projects
    )


# ========== TIME ENTRIES ROUTES ==========

@main_bp.route('/time-entries/new', methods=['GET', 'POST'])
@login_required
def create_time_entry():
    if request.method == 'POST':
        try:
            task_id = request.form.get('task_id')
            task = Task.query.get_or_404(task_id)
            
            # Validar que no hay predecesoras incompletas
            incomplete_preds = task.incomplete_predecessors()
            if incomplete_preds:
                pred_names = ', '.join([p.title for p in incomplete_preds[:3]])
                if len(incomplete_preds) > 3:
                    pred_names += f' y {len(incomplete_preds) - 3} más'
                flash(f'No se puede registrar tiempo en esta tarea. Primero deben completarse las tareas predecesoras: {pred_names}', 'warning')
                return redirect(url_for('main.create_time_entry', task_id=task_id))
            
            date_str = request.form.get('date')
            hours = float(request.form.get('hours'))
            description = request.form.get('description')
            is_billable = request.form.get('is_billable') == 'on'
            
            time_entry = TimeEntry(
                task_id=task_id,
                user_id=current_user.id,
                date=datetime.fromisoformat(date_str).date() if date_str else datetime.now().date(),
                hours=hours,
                description=description,
                is_billable=is_billable
            )
            
            db.session.add(time_entry)
            db.session.flush()
            
            # Auto-update task status if not already in progress/done
            if task.status == 'BACKLOG' or task.status == 'TODO':
                task.status = 'IN_PROGRESS'
            
            db.session.commit()
            # Notify project manager or PMP users about new time entry
            try:
                from app.services.notifications import NotificationService
                project = task.project
                notified = False
                if project and project.manager_id and project.manager_id != current_user.id:
                    NotificationService.notify(
                        user_id=project.manager_id,
                        title='Nuevo registro de tiempo',
                        message=f"{current_user.name} registró {time_entry.hours}h en la tarea '{task.title}'",
                        notification_type=NotificationService.GENERAL,
                        related_entity_type='task',
                        related_entity_id=task.id,
                        send_email=True
                    )
                    notified = True
                if not notified:
                    from ..models import User, Role
                    pmps = User.query.join(Role).filter(Role.name == 'PMP', User.is_active == True).all()
                    for u in pmps:
                        if u.id == current_user.id:
                            continue
                        try:
                            NotificationService.notify(
                                user_id=u.id,
                                title='Nuevo registro de tiempo',
                                message=f"{current_user.name} registró {time_entry.hours}h en la tarea '{task.title}'",
                                notification_type=NotificationService.GENERAL,
                                related_entity_type='task',
                                related_entity_id=task.id,
                                send_email=True
                            )
                        except Exception:
                            current_app.logger.exception('Failed to notify PMP %s about time entry %s', u.id, time_entry.id)
            except Exception:
                current_app.logger.exception('Error while sending time entry notifications')

            # Audit creation of time entry
            try:
                audit = AuditLog(
                    entity_type='time_entry',
                    entity_id=time_entry.id,
                    action='CREATE',
                    user_id=current_user.id,
                    changes={
                        'task_id': time_entry.task_id,
                        'hours': float(time_entry.hours) if time_entry.hours is not None else None,
                        'description': time_entry.description
                    }
                )
                db.session.add(audit)
                db.session.commit()
            except Exception:
                current_app.logger.exception('Failed to write AuditLog for time entry create (web)')

            flash('Tiempo registrado exitosamente', 'success')
            return redirect(url_for('main.task_detail', task_id=task_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    # Show tasks based on role: PMP/Admin sees all, others see only their assigned tasks
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None
    # Respect role: PMP/Admin see all tasks, others should see tasks they are assigned to (either legacy assigned_to or in assignees many-to-many)
    if user_role in ['PMP', 'Admin']:
        tasks = Task.query.order_by(Task.title).all()
    else:
        from ..models import User
        tasks = Task.query.filter(
            (Task.assigned_to_id == user.id) | (Task.assignees.any(User.id == user.id))
        ).order_by(Task.title).all()

    # Marcar tareas bloqueadas por predecesoras
    for t in tasks:
        t.has_incomplete_predecessors = len(t.incomplete_predecessors()) > 0

    # Preserve selected task id if link passed it (e.g., from task detail) so the select pre-selects it
    selected_task_id = request.args.get('task_id', type=int)

    # If a selected_task_id was provided but not in the tasks list (could happen), try to include it
    if selected_task_id and not any(t.id == selected_task_id for t in tasks):
        try:
            sel_t = Task.query.get(int(selected_task_id))
            # only add if the current user should be able to log time for it (PMP/Admin or assigned)
            can_attach = False
            if user_role in ['PMP', 'Admin']:
                can_attach = True
            else:
                if sel_t and (sel_t.assigned_to_id == user.id or (getattr(sel_t, 'assignees', None) and any(u.id == user.id for u in sel_t.assignees))):
                    can_attach = True
            if can_attach and sel_t:
                sel_t.has_incomplete_predecessors = len(sel_t.incomplete_predecessors()) > 0
                tasks = tasks + [sel_t]
        except Exception:
            pass

    return render_template('time_entry_edit.html', entry=None, tasks=tasks, selected_task_id=selected_task_id, now=datetime.now())

    selected_task_id = request.args.get('task_id', type=int)
    return render_template('time_entry_edit.html', tasks=tasks, now=datetime.now(), selected_task_id=selected_task_id)


@main_bp.route('/time-entry/<int:entry_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_time_entry(entry_id):
    entry = TimeEntry.query.get_or_404(entry_id)
    
    # Only PMP or Admin can edit time entries (participants/owners cannot)
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    role_name = user.role.name if (user and user.role) else None
    if role_name not in ('PMP', 'Admin'):
        flash('No tienes permiso para editar registros de tiempo. Solo PMP o Admin pueden hacerlo.', 'danger')
        return redirect(url_for('main.time_entries'))
    
    if request.method == 'POST':
        try:
            old_values = {'date': str(entry.date), 'hours': float(entry.hours), 'is_billable': entry.is_billable}
            
            # Only PMP or Admin can modify time entries and billable flag
            entry.date = datetime.fromisoformat(request.form.get('date')).date()
            entry.hours = float(request.form.get('hours'))
            entry.description = request.form.get('description')
            entry.is_billable = request.form.get('is_billable') == 'on'
            
            new_values = {'date': str(entry.date), 'hours': float(entry.hours), 'is_billable': entry.is_billable}
            changes = {k: {'old': old_values[k], 'new': new_values[k]} for k in old_values if old_values[k] != new_values[k]}
            
            if changes:
                audit = AuditLog(
                    entity_type='TimeEntry',
                    entity_id=entry.id,
                    action='UPDATE',
                    user_id=current_user.id,
                    changes=changes
                )
                db.session.add(audit)
            
            db.session.commit()
            flash('Registro actualizado.', 'success')
            return redirect(url_for('main.time_entries'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    tasks = Task.query.all()
    return render_template('time_entry_edit.html', entry=entry, tasks=tasks)


@main_bp.route('/time-entry/<int:entry_id>/delete', methods=['POST'])
@login_required
def delete_time_entry(entry_id):
    entry = TimeEntry.query.get_or_404(entry_id)
    # Only PMP or Admin may delete time entries
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    role_name = user.role.name if (user and user.role) else None
    if role_name not in ('PMP', 'Admin'):
        flash('No tienes permiso para eliminar registros de tiempo.', 'danger')
        return redirect(url_for('main.time_entries'))
    
    try:
        # Auditoría antes de eliminar
        audit = AuditLog(
            entity_type='TimeEntry',
            entity_id=entry.id,
            action='DELETE',
            user_id=current_user.id,
            changes={'task_id': entry.task_id, 'hours': float(entry.hours), 'date': str(entry.date)}
        )
        db.session.add(audit)
        
        db.session.delete(entry)
        db.session.commit()
        flash('Registro eliminado.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('main.time_entries'))


# ========== TASK ROUTES (ENHANCED) ==========

@main_bp.route('/task/<int:task_id>/status', methods=['POST'])
@internal_required
def update_task_status(task_id):
    from flask import session
    task = Task.query.get_or_404(task_id)
    project = task.project

    # Load fresh user from session
    from ..models import User
    uid = session.get('_user_id')
    user = User.query.get(int(uid)) if uid else None

    role_name = user.role.name if (user and user.role) else None
    is_assigned = bool(
        user and (
            task.assigned_to_id == user.id or
            (getattr(task, 'assignees', None) and any(u.id == user.id for u in task.assignees))
        )
    )

    # PMP/Admin/Supervisor can change task status. Participantes only their assigned tasks.
    if not user or not user.is_internal:
        return jsonify({'error': 'Permission denied'}), 403
    if role_name in ('PMP', 'Admin'):
        pass
    elif role_name == 'Supervisor' and user in project.members:
        pass
    elif role_name == 'Participante' and is_assigned:
        pass
    else:
        return jsonify({'error': 'Permission denied'}), 403
    
    try:
        new_status = request.form.get('status')
        old_status = task.status

        # Validate if task can advance to new status
        if new_status:
            can_advance, error_msg, blockers = task.can_advance_status(new_status)
            if not can_advance:
                return jsonify({
                    'error': error_msg,
                    **(blockers or {})
                }), 400
            # Also check incomplete predecessors when marking COMPLETED (strict sequential enforcement)
            if new_status in ('COMPLETED', 'DONE'):
                incomplete_preds = task.incomplete_predecessors()
                if incomplete_preds:
                    pred_titles = ', '.join([p.title for p in incomplete_preds[:3]])
                    return jsonify({
                        'error': f'No se puede completar la tarea: tiene predecesoras incompletas ({pred_titles})',
                        'incomplete_predecessors': [{'id': p.id, 'title': p.title, 'status': p.status} for p in incomplete_preds]
                    }), 400

        task.set_status(new_status)

        # AuditLog del cambio de estado
        if new_status and new_status != old_status:
            from app.models import AuditLog as _AuditLog
            audit = _AuditLog(
                entity_type='Task',
                entity_id=task.id,
                action='STATUS_CHANGE',
                user_id=current_user.id,
                changes={'status': {'old': old_status, 'new': task.status}}
            )
            db.session.add(audit)

        db.session.commit()

        # Notificaciones de cambio de estado
        if new_status and new_status != old_status:
            try:
                from app.services.notifications import NotificationService
                NotificationService.notify_task_status_changed(task, old_status, user)
            except Exception:
                current_app.logger.exception('Error notifying status change for task %s', task.id)

        # Dispatch webhooks (non-blocking)
        try:
            from app.services import webhook_service
            if new_status and new_status != old_status:
                webhook_data = {
                    'task_id': task.id,
                    'task_title': task.title,
                    'project_id': task.project_id,
                    'project_name': project.name if project else None,
                    'user_name': user.name if user else None,
                    'old_status': old_status,
                    'new_status': task.status,
                }
                if task.status == 'COMPLETED':
                    webhook_service.dispatch('task.completed', webhook_data)
                else:
                    webhook_service.dispatch('task.status_changed', webhook_data)
        except Exception:
            current_app.logger.exception('Error dispatching webhook for task %s', task.id)

        return redirect(url_for('main.task_detail', task_id=task.id))
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400


@main_bp.route('/task/<int:task_id>')
@login_required
def task_detail(task_id):
    task = Task.query.get(task_id)
    if not task:
        # Task not found — determine if it ever existed by checking audit logs
        from app.models import AuditLog
        any_log = AuditLog.query.filter_by(entity_type='Task', entity_id=task_id).order_by(AuditLog.created_at.desc()).first()
        if not any_log:
            return render_template('item_status.html', entity_type='tarea', entity_id=task_id, status='never', canonical_entity='Task'), 404
        # If we have a DELETE record, show deleted message, otherwise show 'unavailable'
        deletion = AuditLog.query.filter_by(entity_type='Task', entity_id=task_id, action='DELETE').order_by(AuditLog.created_at.desc()).first()
        if deletion:
            deleted_by = deletion.user.name if deletion.user else (f'Usuario {deletion.user_id}' if deletion.user_id else None)
            deleted_at = deletion.created_at
            return render_template('item_status.html', entity_type='tarea', entity_id=task_id, status='deleted', deleted_by=deleted_by, deleted_at=deleted_at, canonical_entity='Task'), 404
        return render_template('item_status.html', entity_type='tarea', entity_id=task_id, status='unavailable', canonical_entity='Task'), 404
    project = task.project
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None

    # Control de acceso
    can_view = False
    can_edit = False
    
    user_id = user.id if user else None
    if user_role in ['PMP', 'Admin']:
        can_view = True
        can_edit = True
    elif user_role == 'Supervisor':
        # Supervisor puede ver y editar tareas en proyectos donde es miembro
        is_project_member = task.project and (user in task.project.members)
        can_view = bool(is_project_member)
        can_edit = bool(is_project_member)
    elif user_role == 'Participante':
        # Participante (internal) puede ver cualquier tarea; editar solo si esta asignado
        is_assigned = task.assigned_to_id == user_id or (
            getattr(task, 'assignees', None) and any(u.id == user_id for u in task.assignees)
        )
        # Participante can view any task (internal user), but edit only if assigned
        can_view = True
        can_edit = bool(is_assigned)
    elif user_role == 'Cliente' or not (user.is_internal if user else True):
        # Cliente puede ver tareas de sus proyectos solo si la tarea es visible externamente o está asignada a él (solo lectura)
        project_client_ids = [u.id for u in project.clients] if project else []
        if user_id in project_client_ids:
            if task.is_external_visible or task.assigned_client_id == user_id:
                can_view = True
                can_edit = False

    if not can_view:
        flash('No tienes permiso para ver esta tarea.', 'danger')
        return redirect(url_for('main.project_detail', project_id=task.project_id))

    # can_edit_task_form: puede acceder al formulario de edición completo (PMP/Admin/Supervisor)
    # Distinto de can_edit (que para Participante es True si está asignado, pero solo puede cambiar estado)
    can_edit_task_form = user_role in ('PMP', 'Admin', 'Supervisor') and can_edit

    # Solo usuarios asignados pueden comentar o subir archivos.
    assigned_user_ids = set([u.id for u in task.assignees]) if getattr(task, 'assignees', None) else set()
    can_comment_upload = bool(
        user_role in ('PMP', 'Admin') or
        (user_role == 'Supervisor' and task.project and user in task.project.members) or
        task.assigned_to_id == user_id or
        user_id in assigned_user_ids or
        task.assigned_client_id == user_id
    )
    can_delete_task = user_role in ('PMP', 'Admin')

    time_entries = TimeEntry.query.filter_by(task_id=task_id).all()
    total_hours = db.session.query(func.sum(TimeEntry.hours)).filter_by(task_id=task_id).scalar() or 0

    return render_template('task_detail.html', task=task, time_entries=time_entries, total_hours=total_hours,
                           now=datetime.now(), can_edit=can_edit,
                           can_edit_task_form=can_edit_task_form,
                           can_delete_task=can_delete_task,
                           can_comment_upload=can_comment_upload)


@main_bp.route('/task/<int:task_id>/comments', methods=['GET', 'POST'])
@login_required
def task_comments(task_id):
    from app.models import TaskComment, Task, Project
    task = Task.query.get(task_id)
    if not task:
        return jsonify({'error': 'Tarea no encontrada'}), 404

    # Determine if current user can view the task (reuse same logic as task_detail)
    can_view = False
    user_role = current_user.role.name if (current_user and current_user.role) else None
    project = task.project
    if user_role in ['PMP', 'Admin']:
        can_view = True
    elif user_role == 'Supervisor':
        can_view = bool(project and current_user in project.members)
    elif user_role == 'Participante':
        is_assigned = task.assigned_to_id == current_user.id or (
            getattr(task, 'assignees', None) and any(u.id == current_user.id for u in task.assignees)
        )
        is_project_member = project and (current_user in project.members)
        if is_project_member or is_assigned:
            can_view = True
    elif user_role == 'Cliente' or not current_user.is_internal:
        if current_user in project.clients:
            if task.is_external_visible or task.assigned_client_id == current_user.id:
                can_view = True

    if not can_view:
        return jsonify({'error': 'No autorizado'}), 403

    assigned_user_ids = set([u.id for u in task.assignees]) if getattr(task, 'assignees', None) else set()
    can_comment_upload = bool(
        user_role in ('PMP', 'Admin') or
        (user_role == 'Supervisor' and project and current_user in project.members) or
        task.assigned_to_id == current_user.id or
        current_user.id in assigned_user_ids or
        task.assigned_client_id == current_user.id
    )

    if request.method == 'GET':
        # Paginate top-level comments (parent_id is NULL). Return latest comments first
        try:
            page = int(request.args.get('page', 1))
        except Exception:
            page = 1
        try:
            per_page = int(request.args.get('per_page', 10))
        except Exception:
            per_page = 10
        if per_page <= 0:
            per_page = 10
        if per_page > 50:
            per_page = 50

        base_q = TaskComment.query.filter_by(task_id=task_id, parent_id=None)
        total = base_q.count()
        top_comments = base_q.order_by(TaskComment.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

        out = []

        def build_children_recursive(parent_comment):
            children_rows = TaskComment.query.filter_by(parent_id=parent_comment.id).order_by(TaskComment.created_at.asc()).all()
            children_list = []
            for ch in children_rows:
                ch_dict = {
                    'id': ch.id,
                    'body': ch.body,
                    'created_at': ch.created_at.isoformat() if ch.created_at else None,
                    'parent_id': ch.parent_id,
                    'user': {'id': ch.user.id, 'name': ch.user.name},
                    'children': build_children_recursive(ch)
                }
                children_list.append(ch_dict)
            return children_list

        for c in top_comments:
            out.append({
                'id': c.id,
                'body': c.body,
                'created_at': c.created_at.isoformat() if c.created_at else None,
                'parent_id': c.parent_id,
                'user': {'id': c.user.id, 'name': c.user.name},
                'children': build_children_recursive(c)
            })

        has_more = (page * per_page) < total
        return jsonify({'comments': out, 'page': page, 'per_page': per_page, 'total': total, 'has_more': has_more})

    # POST - create comment
    if not can_comment_upload:
        return jsonify({'error': 'Solo usuarios asignados pueden comentar en esta tarea'}), 403

    data = request.get_json(force=True, silent=True) or request.form
    body = (data.get('body') or '').strip() if data else ''
    if not body:
        return jsonify({'error': 'El comentario está vacío'}), 400

    try:
        parent_id = data.get('parent_id') if isinstance(data, dict) else None
        if parent_id:
            try:
                parent_id = int(parent_id)
            except Exception:
                parent_id = None

        # If parent_id provided, verify it exists and belongs to same task
        parent = None
        if parent_id:
            parent = TaskComment.query.get(parent_id)
            if not parent or parent.task_id != task_id:
                return jsonify({'error': 'Parent comment inválido'}), 400

        comment = TaskComment(task_id=task_id, user_id=current_user.id, body=body, parent_id=parent_id)
        db.session.add(comment)

        # AuditLog del comentario
        from app.models import AuditLog as _AuditLog
        _snippet = body[:200] + ('...' if len(body) > 200 else '')
        _audit_comment = _AuditLog(
            entity_type='Task',
            entity_id=task_id,
            action='COMMENT',
            user_id=current_user.id,
            changes={'body_snippet': _snippet, 'is_reply': bool(parent_id)}
        )
        db.session.add(_audit_comment)
        db.session.commit()

        # Send notifications: if this is a reply, notify the parent author (unless it's the same user).
        # Otherwise, notify assignees + PMP principal + PMP adicionales + Supervisores of the task.
        try:
            import re as _re
            from app.services.notifications import NotificationService
            from app.models import SystemSettings as _SysSettings

            # Resolve email send preference from system settings
            _comment_setting = _SysSettings.get('notify_task_comment', True)
            _send_comment_email = _comment_setting.lower() in ('true', '1', 'yes') if isinstance(_comment_setting, str) else bool(_comment_setting)
            # Mentions always notify (separate setting could be added later)
            _mention_setting = _SysSettings.get('notify_task_mention', True)
            _send_mention_email = _mention_setting.lower() in ('true', '1', 'yes') if isinstance(_mention_setting, str) else bool(_mention_setting)

            recipients = set()
            # If reply to a comment, notify that comment's author
            if parent and parent.user_id and parent.user_id != current_user.id:
                recipients.add(parent.user_id)
            else:
                # Notify primary assignee
                if getattr(task, 'assigned_to_id', None):
                    recipients.add(task.assigned_to_id)
                # Notify additional assignees
                if getattr(task, 'assignees', None):
                    for u in task.assignees:
                        if getattr(u, 'id', None) and u.id != current_user.id:
                            recipients.add(u.id)
                # Notify PMP principal
                if task.project and task.project.manager_id:
                    recipients.add(task.project.manager_id)
                # Notify PMP adicionales y Supervisores del proyecto
                if task.project and getattr(task.project, 'members', None):
                    for _m in task.project.members:
                        if getattr(_m, 'role', None) and _m.role.name in ('PMP', 'Supervisor'):
                            recipients.add(_m.id)

            # Parse @[Name] mentions and notify those users directly
            mention_names = _re.findall(r'@\[([^\]]+)\]', comment.body)
            mention_recipients = set()
            if mention_names:
                all_active = User.query.filter_by(is_active=True).all()
                name_to_id = {u.name.lower(): u.id for u in all_active}
                for mname in mention_names:
                    uid_m = name_to_id.get(mname.lower())
                    if uid_m and uid_m != current_user.id:
                        mention_recipients.add(uid_m)

            # Remove the commenter from recipients if present
            recipients.discard(current_user.id)
            mention_recipients.discard(current_user.id)

            snippet = (comment.body[:120] + '...') if len(comment.body) > 120 else comment.body

            # Build task URL using the actual request host (no hardcoded domain)
            try:
                _task_url = url_for('main.task_detail', task_id=task.id, _external=True)
            except Exception:
                _task_url = f"/task/{task.id}"

            # Notify regular recipients (comment on task)
            title = 'Nuevo comentario en tarea'
            message = f"{current_user.name or current_user.email} comentó en la tarea '{task.title}': {snippet}"
            for uid in recipients - mention_recipients:
                try:
                    NotificationService.notify(
                        user_id=uid,
                        title=title,
                        message=message,
                        notification_type=NotificationService.TASK_COMMENT,
                        related_entity_type='task',
                        related_entity_id=task.id,
                        send_email=_send_comment_email,
                        email_context={'task': task, 'comment': comment, 'commenter': current_user,
                                       'message': message, 'title': title, 'task_url': _task_url}
                    )
                except Exception:
                    current_app.logger.exception(f'Failed to notify user {uid} about comment {comment.id}')

            # Notify mentioned users with a specific message
            mention_title = f'{current_user.name or current_user.email} te mencionó en un comentario'
            mention_msg = f"Te mencionaron en la tarea '{task.title}': {snippet}"
            for uid in mention_recipients:
                try:
                    NotificationService.notify(
                        user_id=uid,
                        title=mention_title,
                        message=mention_msg,
                        notification_type=NotificationService.MENTION,
                        related_entity_type='task',
                        related_entity_id=task.id,
                        send_email=_send_mention_email,
                        email_context={'task': task, 'comment': comment, 'commenter': current_user,
                                       'message': mention_msg, 'title': mention_title, 'task_url': _task_url}
                    )
                except Exception:
                    current_app.logger.exception(f'Failed to notify mentioned user {uid}')
        except Exception:
            current_app.logger.exception('Failed to dispatch comment notifications')

        # Webhook: comment.created
        try:
            from app.services import webhook_service
            snippet = (comment.body[:200] + '...') if len(comment.body) > 200 else comment.body
            webhook_service.dispatch('comment.created', {
                'task_id': task.id,
                'task_title': task.title,
                'project_name': task.project.name if task.project else None,
                'user_name': current_user.name or current_user.email,
                'comment_snippet': snippet,
                'is_reply': bool(parent_id),
            })
        except Exception:
            current_app.logger.exception('Failed to dispatch comment.created webhook')

        return jsonify({
            'id': comment.id,
            'body': comment.body,
            'created_at': comment.created_at.isoformat() if comment.created_at else None,
            'parent_id': comment.parent_id,
            'user': {'id': current_user.id, 'name': current_user.name}
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@main_bp.route('/api/mention-users')
@login_required
def mention_users():
    """Return active users matching a name query, for @mention autocomplete."""
    q = request.args.get('q', '').strip().lower()
    users = User.query.filter_by(is_active=True).all()
    results = []
    for u in users:
        full_name = u.name  # computed property: first_name + last_name or email
        if not q or q in full_name.lower():
            results.append({'id': u.id, 'name': full_name})
    results.sort(key=lambda x: x['name'])
    return jsonify(results[:10])


@main_bp.route('/task/<int:task_id>/upload', methods=['POST'])
@login_required
def task_upload(task_id):
    """Upload attachment(s) to a task (form-based, redirects back to task detail)."""
    task = Task.query.get_or_404(task_id)
    project = task.project

    # Check permissions: internal users or clients belonging to this project
    is_project_client = project and (
        (project.client_id == current_user.id) or
        (current_user.id in [u.id for u in project.clients])
    )
    can_upload = (
        current_user.is_internal or
        is_project_client
    )

    if not can_upload:
        flash('No tienes permiso para subir archivos a esta tarea.', 'danger')
        return redirect(url_for('main.task_detail', task_id=task_id))

    if 'file' not in request.files:
        flash('No se encontró archivo.', 'warning')
        return redirect(url_for('main.task_detail', task_id=task_id))

    files = request.files.getlist('file')
    saved_count = 0
    skipped_filenames = []

    ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx',
                          'png', 'jpg', 'jpeg', 'gif', 'svg', 'webp',
                          'txt', 'csv', 'zip', 'rar', '7z',
                          'mp4', 'mov', 'avi', 'mp3', 'wav'}

    for file in files:
        if not file or not file.filename:
            continue
        from werkzeug.utils import secure_filename
        filename = secure_filename(file.filename)
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        if ext not in ALLOWED_EXTENSIONS:
            skipped_filenames.append(file.filename)
            continue

        import mimetypes
        mime_type = mimetypes.guess_type(filename)[0] or 'application/octet-stream'
        file_data = file.read()
        file_size = len(file_data)

        task_folder = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), f'task_{task_id}')
        os.makedirs(task_folder, exist_ok=True)

        import uuid
        stored_filename = f"{uuid.uuid4().hex}_{filename}"
        filepath = os.path.join(task_folder, stored_filename)
        with open(filepath, 'wb') as f:
            f.write(file_data)

        att = TaskAttachment(
            task_id=task_id,
            filename=filename,
            stored_filename=stored_filename,
            file_size=file_size,
            mime_type=mime_type,
            uploaded_by_id=current_user.id
        )
        db.session.add(att)
        saved_count += 1

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash('Error al guardar los archivos.', 'danger')
        return redirect(url_for('main.task_detail', task_id=task_id))

    if skipped_filenames:
        flash(f'Algunos archivos no fueron subidos porque su extensión no está permitida: {", ".join(skipped_filenames)}', 'warning')
    if saved_count:
        flash(f'{saved_count} archivo(s) subido(s) correctamente.', 'success')

    return redirect(url_for('main.task_detail', task_id=task_id))


@main_bp.route('/attachment/<int:attachment_id>/download')
@login_required
def download_attachment(attachment_id):
    """Download a task attachment"""
    attachment = TaskAttachment.query.get_or_404(attachment_id)
    task = attachment.task
    
    # Check permissions - user must be able to view the task
    from ..auth.decorators import _get_user_from_session
    _current_user = _get_user_from_session()
    can_view = False
    if _current_user and _current_user.is_internal:
        can_view = True
    elif _current_user:
        # Clients can only download attachments from tasks they can view
        if task.is_external_visible or task.assigned_client_id == _current_user.id:
            can_view = True

    if not can_view:
        return ('No tienes permiso para acceder a este archivo.', 403)
    
    # Build the file path
    task_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], f'task_{task.id}')
    filepath = os.path.join(task_folder, attachment.stored_filename)
    
    if not os.path.exists(filepath):
        flash('Archivo no encontrado.', 'danger')
        return redirect(url_for('main.task_detail', task_id=task.id))
    
    # choose inline or attachment depending on query string (preview functionality)
    as_attachment = not request.args.get('inline')
    return send_file(
        filepath,
        mimetype=attachment.mime_type or 'application/octet-stream',
        as_attachment=as_attachment,
        download_name=attachment.filename
    )


@main_bp.route('/task/<int:task_id>/move', methods=['POST'])
@login_required
def move_task(task_id):
    task = Task.query.get_or_404(task_id)
    project = task.project

    # Only internal users can move tasks between statuses
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    if not user or not user.is_internal:
        return jsonify({'error': 'No tienes permiso para mover tareas.'}), 403

    role_name = user.role.name if (user and user.role) else None
    is_assigned = bool(
        task.assigned_to_id == user.id or
        (getattr(task, 'assignees', None) and any(u.id == user.id for u in task.assignees))
    )

    if role_name in ('PMP', 'Admin'):
        pass
    elif role_name == 'Supervisor' and user in project.members:
        pass
    elif role_name == 'Participante' and is_assigned:
        pass
    else:
        return jsonify({'error': 'No tienes permiso para mover esta tarea.'}), 403

    data = request.get_json() or request.form
    new_status = data.get('status')
    old_status = task.status

    VALID_STATUSES = ['BACKLOG', 'IN_PROGRESS', 'IN_REVIEW', 'COMPLETED']
    if not new_status or new_status not in VALID_STATUSES:
        return jsonify({'error': 'Estado inválido.'}), 400

    # Validate if task can advance to new status (blocked by predecessors or children)
    can_advance, error_msg, blockers = task.can_advance_status(new_status)
    if not can_advance:
        return jsonify({
            'error': error_msg,
            **(blockers or {})
        }), 400

    try:
        # Normalize and set status
        task.set_status(new_status)
        normalized_new = task.status
        
        # Registrar auditoría de cambio de estado (use session-bound user)
        audit = AuditLog(
            entity_type='Task',
            entity_id=task.id,
            action='UPDATE',
            user_id=user.id,
            changes={'status': {'old': old_status, 'new': normalized_new}}
        )
        db.session.add(audit)
        
        # Si la tarea se completa, notificar a los clientes
        if normalized_new == 'COMPLETED' and old_status != 'COMPLETED':
            notify_clients_task_completed(task, completed_by_user=user)
        # Notificar cambio de estado
        elif old_status != normalized_new:
            NotificationService.notify_task_status_changed(
                task=task,
                old_status=old_status,
                changed_by_user=user,
            )
        
        db.session.commit()

        # Dispatch webhooks (non-blocking)
        try:
            from app.services import webhook_service
            if old_status != normalized_new:
                webhook_data = {
                    'task_id': task.id,
                    'task_title': task.title,
                    'project_id': task.project_id,
                    'project_name': project.name if project else None,
                    'user_name': user.name if user else None,
                    'old_status': old_status,
                    'new_status': normalized_new,
                }
                if normalized_new == 'COMPLETED':
                    webhook_service.dispatch('task.completed', webhook_data)
                else:
                    webhook_service.dispatch('task.status_changed', webhook_data)
        except Exception:
            current_app.logger.exception('Error dispatching webhook for task %s', task.id)

        return jsonify({'status': 'ok', 'task_id': task.id, 'new_status': task.status})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@main_bp.route('/user/<int:user_id>/photo')
@login_required
def user_photo(user_id):
    user = User.query.get_or_404(user_id)

    # If we already have a stored photo serve it
    if user.photo:
        resp = send_file(BytesIO(user.photo), mimetype=(user.photo_mime or 'image/jpeg'))
        resp.headers['Cache-Control'] = 'public, max-age=3600'
        return resp

    # Fallback: try to fetch from Microsoft Graph using application credentials (if configured)
    if user.azure_oid:
        try:
            from ..auth.utils import get_msal_app
            msal_app = get_msal_app()
            if msal_app:
                token = msal_app.acquire_token_for_client(scopes=['https://graph.microsoft.com/.default'])
                access = token.get('access_token')
                if access:
                    g = requests.get(f'https://graph.microsoft.com/v1.0/users/{user.azure_oid}/photo/$value', headers={'Authorization': f'Bearer {access}'}, timeout=5)
                    if g.status_code == 200 and g.content:
                        user.photo = g.content
                        user.photo_mime = g.headers.get('Content-Type', 'image/jpeg') or 'image/jpeg'
                        user.photo_updated_at = datetime.now()
                        db.session.add(user)
                        db.session.commit()
                        resp = send_file(BytesIO(user.photo), mimetype=user.photo_mime)
                        resp.headers['Cache-Control'] = 'public, max-age=3600'
                        return resp
        except Exception:
            # best-effort fallback — do not block or raise
            current_app.logger.debug('Graph fallback for user photo failed or not configured')

    # Nothing available — return 404 so templates can fall back to initials
    abort(404)


@main_bp.route('/user/<int:user_id>')
@login_required
def user_profile(user_id):
    user = User.query.get_or_404(user_id)
    # Only internal users can view other users' profiles (PMP/Admins primarily). Allow self view.
    if current_user.id != user.id and not current_user.is_internal:
        flash('No tienes permiso para ver este perfil.', 'danger')
        return redirect(url_for('main.profile'))

    # Compute simple stats
    tasks_assigned = Task.query.filter_by(assigned_to_id=user.id).count()
    tasks_completed = Task.query.filter_by(assigned_to_id=user.id, status='COMPLETED').count()
    projects_managed = Project.query.filter_by(manager_id=user.id).count() if user.is_internal else 0
    total_hours = db.session.query(func.coalesce(func.sum(TimeEntry.hours), 0)).filter(TimeEntry.user_id == user.id).scalar() or 0

    recent_tasks = Task.query.filter_by(assigned_to_id=user.id).order_by(Task.due_date.desc().nullslast()).limit(5).all()

    # Attach last activity timestamp for each recent task (from AuditLog, TimeEntry, or approval timestamp)
    for t in recent_tasks:
        last_audit = db.session.query(func.max(AuditLog.created_at)).filter(AuditLog.entity_type == 'Task', AuditLog.entity_id == t.id).scalar()
        last_entry = db.session.query(func.max(TimeEntry.created_at)).filter(TimeEntry.task_id == t.id).scalar()
        candidates = [dt for dt in (last_audit, last_entry, t.approved_at) if dt is not None]
        t.last_activity = max(candidates) if candidates else None

    stats = {
        'tasks_assigned': tasks_assigned,
        'tasks_completed': tasks_completed,
        'projects_managed': projects_managed,
        'total_hours': float(total_hours)
    }

    return render_template('user_profile.html', user=user, stats=stats, recent_tasks=recent_tasks)


@main_bp.route('/task/<int:task_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_task(task_id):
    task = Task.query.get_or_404(task_id)
    project = task.project

    # Permitir a usuarios internos o al cliente del proyecto (si está asignado a la tarea)
    is_project_client = (project.client_id == current_user.id) or (current_user in project.clients)
    can_edit = current_user.is_internal or (is_project_client and task.assigned_client_id == current_user.id)
    if not can_edit:
        flash('No tienes permiso para editar esta tarea.', 'danger')
        return redirect(url_for('main.task_detail', task_id=task.id))
    
    # Fix 7: Rol Participante no debe editar tareas (solo registrar tiempo)
    if current_user.role and current_user.role.name == 'Participante':
        flash('Los participantes no tienen permiso para editar tareas. Solo pueden registrar tiempos.', 'danger')
        return redirect(url_for('main.task_detail', task_id=task.id))

    if request.method == 'POST':
        try:
            # Guardar valores anteriores para detectar cambios y auditoría
            old_assigned_to_id = task.assigned_to_id
            old_assigned_client_id = task.assigned_client_id
            old_parent_id = task.parent_task_id
            old_assignees = set([u.id for u in task.assignees]) if getattr(task, 'assignees', None) else set()
            old_values = {
                'title': task.title,
                'description': task.description,
                'status': task.status,
                'priority': task.priority,
                'assigned_to_id': task.assigned_to_id,
                'is_internal_only': task.is_internal_only,
                'parent_task_id': task.parent_task_id,
                'start_date': str(task.start_date),
                'due_date': str(task.due_date),
                'estimated_hours': str(task.estimated_hours),
                'predecessor_ids': sorted([p.id for p in task.predecessors]),
            }

            # Role checks: only PMP/Admin can modify most fields. Participants and Clients may only change status and upload files.
            is_pmp_admin = (current_user.is_authenticated and current_user.is_internal and current_user.role and current_user.role.name in ('PMP', 'Admin', 'Supervisor'))
            is_participant = (current_user.is_authenticated and current_user.is_internal and current_user.role and current_user.role.name == 'Participante')
            is_client = (not current_user.is_internal) and (project.client_id == current_user.id)
            is_limited_editor = is_participant or is_client

            if is_limited_editor:
                # Only allow status in form. If any other form fields are present, reject and redirect.
                allowed = {'status'}
                extra = set([k for k in request.form.keys() if k not in allowed and not k.startswith('_')])
                if extra:
                    flash('No tienes permiso para modificar campos además de estado y archivos.', 'danger')
                    return redirect(url_for('main.edit_task', task_id=task.id))

            # Process fields (PMP/Admin or allowed status for limited editors)
            task.title = request.form.get('title') or task.title
            task.description = request.form.get('description')

            # Handle predecessors (many-to-many) EARLY: only if provided in form, validate and assign BEFORE status validation
            # Fix 3: Permitir desmarcar predecesoras (procesar siempre si es POST)
            predecessor_ids = [int(x) for x in request.form.getlist('predecessor_ids') if x and x.strip()]
            try:
                # validate before assignment
                task.validate_predecessor_ids(predecessor_ids)
                # Si la lista está vacía, se limpiarán las predecesoras (correcto para desmarcar)
                preds = Task.query.filter(Task.id.in_(predecessor_ids)).all() if predecessor_ids else []
                task.predecessors = preds
            except ValueError as ve:
                raise ve

            status_from_form = request.form.get('status')
            if status_from_form and status_from_form != task.status:
                can_advance, error_msg, blockers = task.can_advance_status(status_from_form)
                if not can_advance:
                    raise ValueError(error_msg)
                # Also check incomplete predecessors when marking COMPLETED (strict sequential enforcement)
                if status_from_form in ('COMPLETED', 'DONE'):
                    incomplete_preds = task.incomplete_predecessors()
                    if incomplete_preds:
                        pred_titles = ', '.join([p.title for p in incomplete_preds[:3]])
                        raise ValueError(f'No se puede completar la tarea: tiene predecesoras incompletas ({pred_titles})')
                task.set_status(status_from_form)
            # If no status provided, keep existing task.status unchanged

            task.priority = request.form.get('priority') or task.priority
            # Only update is_internal_only if provided in form (checkboxes absent when not submitted)
            if 'is_internal_only' in request.form:
                new_internal_flag = request.form.get('is_internal_only') == 'on'
                task.is_internal_only = new_internal_flag
                # Propagate the internal-only flag to all hierarchical descendants (children recursively)
                # so that when a parent is marked private, its subtasks are not visible to clients.
                stack = list(getattr(task, 'children', []) or [])
                visited = set()
                while stack:
                    node = stack.pop()
                    if node.id in visited:
                        continue
                    visited.add(node.id)
                    node.is_internal_only = new_internal_flag
                    for c in getattr(node, 'children', []) or []:
                        if c.id not in visited:
                            stack.append(c)
            start_date_str = request.form.get('start_date')
            due_date_str = request.form.get('due_date')

            # Only PMP/Admin may modify dates
            if (start_date_str or due_date_str) and not is_pmp_admin:
                flash('No tienes permiso para modificar fechas de la tarea.', 'danger')
                return redirect(url_for('main.edit_task', task_id=task.id))

            # Fix 8: Validación de fechas en edición
            if start_date_str and due_date_str and start_date_str > due_date_str:
                flash('La fecha de vencimiento no puede ser anterior a la de inicio.', 'danger')
                return redirect(url_for('main.edit_task', task_id=task.id))

            if start_date_str:
                task.start_date = datetime.fromisoformat(start_date_str)
            else:
                task.start_date = None

            if due_date_str:
                task.due_date = datetime.fromisoformat(due_date_str)
            else:
                task.due_date = None

            # Update parent task (validate no cycles)
            # Update parent task only if provided in form
            if 'parent_task_id' in request.form:
                parent_task_id = request.form.get('parent_task_id')
                if parent_task_id and parent_task_id.strip():
                    new_parent_id = int(parent_task_id)
                    if new_parent_id == task.id:
                        raise ValueError('La tarea no puede ser su propia tarea padre')
                    parent_task = Task.query.get(new_parent_id)
                    if not parent_task or parent_task.project_id != project.id:
                        raise ValueError('Tarea padre inválida')
                    # Ensure not assigning a descendant as parent (would create cycle)
                    descendant_ids = [d.id for d in task.descendants()]
                    if new_parent_id in descendant_ids:
                        raise ValueError('No se puede asignar una subtarea como tarea padre')
                    task.parent_task_id = new_parent_id
                else:
                    task.parent_task_id = None

            # Update estimated_hours only if provided in form
            if 'estimated_hours' in request.form:
                estimated_hours = request.form.get('estimated_hours')
                if estimated_hours and estimated_hours.strip():
                    task.estimated_hours = float(estimated_hours)
                else:
                    task.estimated_hours = None

            # Update assignees from three separate role pickers: PMP / Supervisor / Participante
            # Also accept legacy 'assignees' field for backward compat (API calls, etc.)
            SPLIT_FIELDS = ('pmp_assignee_ids', 'supervisor_assignee_ids', 'participant_assignee_ids')
            uses_split = any(f in request.form for f in SPLIT_FIELDS)
            uses_legacy = 'assignees' in request.form and not uses_split

            if uses_split or uses_legacy:
                if uses_split:
                    raw_ids = []
                    for field in SPLIT_FIELDS:
                        raw_ids += request.form.getlist(field)
                else:
                    raw_ids = request.form.getlist('assignees')
                current_app.logger.debug('edit_task: combined assignee ids = %s', raw_ids)
                assignee_ids = list({int(x) for x in raw_ids if x and x.strip()})
                if assignee_ids:
                    users_assigned = User.query.filter(User.id.in_(assignee_ids)).all()
                    task.assignees = users_assigned
                    task.assigned_to_id = users_assigned[0].id if users_assigned else None
                    # Auto-add assignees to project members if not already members
                    existing_member_ids = {m.id for m in project.members}
                    for u in users_assigned:
                        if u.id not in existing_member_ids:
                            project.members.append(u)
                else:
                    task.assignees = []
                    task.assigned_to_id = None

            # Update assigned clients (multi-select); keep assigned_client_id as primary/legacy
            assigned_client_provided = False
            new_assigned_client_id = None
            if 'assigned_client_id' in request.form:
                assigned_client_provided = True
                client_ids = [int(x) for x in request.form.getlist('assigned_client_id') if x and x.strip()]
                if client_ids:
                    client_users = User.query.filter(User.id.in_(client_ids)).all()
                    task.assigned_clients = client_users
                    new_assigned_client_id = client_ids[0]
                    task.assigned_client_id = new_assigned_client_id
                else:
                    task.assigned_clients = []
                    task.assigned_client_id = None
            # Predecessor handling moved earlier to occur BEFORE status validation to ensure
            # that status changes take the new predecessors into account. See above.

            # Handle file attachments
            files = request.files.getlist('attachments')
            has_files_to_upload = any(f and getattr(f, 'filename', None) for f in files)
            assigned_user_ids = set([u.id for u in task.assignees]) if getattr(task, 'assignees', None) else set()
            assigned_client_ids = set([u.id for u in task.assigned_clients]) if getattr(task, 'assigned_clients', None) else set()
            can_upload_attachments = bool(
                is_pmp_admin or
                task.assigned_to_id == current_user.id or
                current_user.id in assigned_user_ids or
                task.assigned_client_id == current_user.id or
                current_user.id in assigned_client_ids
            )
            if has_files_to_upload and not can_upload_attachments:
                flash('Solo usuarios asignados pueden subir archivos a esta tarea.', 'danger')
                return redirect(url_for('main.task_detail', task_id=task.id))

            invalid_files = []
            for file in files:
                if file and file.filename:
                    if not allowed_file(file.filename):
                        invalid_files.append(file.filename)
                        continue
                    filename = secure_filename(file.filename)
                    # get_unique_filename returns (stored_filename, task_folder)
                    stored_filename, task_folder = get_unique_filename(task.id, filename)
                    
                    # task_folder is already created inside get_unique_filename
                    file_path = os.path.join(task_folder, stored_filename)
                    file.save(file_path)
                    
                    attachment = TaskAttachment(
                        task_id=task.id,
                        filename=filename,
                        stored_filename=stored_filename,
                        file_size=os.path.getsize(file_path),
                        mime_type=file.content_type,
                        uploaded_by_id=current_user.id
                    )
                    db.session.add(attachment)

            if invalid_files:
                flash('Algunos archivos no fueron subidos porque su extensión no está permitida: ' + ', '.join(invalid_files), 'warning')

            # Registrar auditoría de cambios en tarea
            new_values = {
                'title': task.title,
                'description': task.description,
                'status': task.status,
                'priority': task.priority,
                'assigned_to_id': task.assigned_to_id,
                'is_internal_only': task.is_internal_only,
                'parent_task_id': task.parent_task_id,
                'start_date': str(task.start_date),
                'due_date': str(task.due_date),
                'estimated_hours': str(task.estimated_hours),
                'predecessor_ids': sorted([p.id for p in task.predecessors]),
            }
            changes = {k: {'old': old_values[k], 'new': new_values[k]} for k in old_values if old_values[k] != new_values[k]}
            
            if changes:
                audit = AuditLog(
                    entity_type='Task',
                    entity_id=task.id,
                    action='UPDATE',
                    user_id=current_user.id,
                    changes=changes
                )
                db.session.add(audit)

            db.session.commit()

            # Dispatch webhooks for edits
            try:
                from app.services import webhook_service
                _old_st = old_values.get('status')
                _wh_base = {
                    'task_id': task.id,
                    'task_title': task.title,
                    'project_id': task.project_id,
                    'project_name': task.project.name if task.project else None,
                    'user_name': current_user.name if current_user else None,
                }
                # Status-specific events
                if _old_st and _old_st != task.status:
                    _wh_status = {**_wh_base, 'old_status': _old_st, 'new_status': task.status}
                    if task.status == 'COMPLETED':
                        webhook_service.dispatch('task.completed', _wh_status)
                    else:
                        webhook_service.dispatch('task.status_changed', _wh_status)
                # General update event for any field change
                # Build new_values matching the same keys as old_values
                _new_values = {
                    'title': task.title,
                    'description': task.description,
                    'status': task.status,
                    'priority': task.priority,
                    'assigned_to_id': task.assigned_to_id,
                    'is_internal_only': task.is_internal_only,
                    'parent_task_id': task.parent_task_id,
                    'start_date': str(task.start_date),
                    'due_date': str(task.due_date),
                    'estimated_hours': str(task.estimated_hours),
                    'predecessor_ids': sorted([p.id for p in task.predecessors]),
                }
                _changed = [k for k in old_values if str(old_values[k]) != str(_new_values.get(k))]
                current_app.logger.info('Webhook task.updated: changed_fields=%s task=%s', _changed, task.id)
                if _changed:
                    webhook_service.dispatch('task.updated', {**_wh_base, 'changed_fields': _changed})
            except Exception:
                current_app.logger.exception('Error dispatching webhook for task %s', task.id)

            send_email_setting = SystemSettings.get('notify_task_assigned', 'true')
            send_email = send_email_setting == 'true' or send_email_setting == True
            email_sent = False

            # Notify status change: PMP principal, PMP adicionales, Supervisores, Clientes asignados
            _old_st = old_values.get('status')
            if _old_st and _old_st != task.status:
                try:
                    NotificationService.notify_task_status_changed(
                        task=task,
                        old_status=_old_st,
                        changed_by_user=current_user,
                    )
                except Exception:
                    current_app.logger.exception('Error notifying status change for task %s', task.id)

            # Notify newly added assignees (for multi-assign) - including self-assignment
            new_assignees = set([u.id for u in task.assignees]) if getattr(task, 'assignees', None) else set()
            added = new_assignees - old_assignees
            if added:
                project_obj = Project.query.get(task.project_id) if task.project_id else None
                for uid in added:
                    try:
                        NotificationService.notify(
                            user_id=uid,
                            title='Nueva tarea asignada',
                            message=f"Se te ha asignado la tarea '{task.title}'{(' en el proyecto '+project_obj.name) if project_obj else ''}",
                            notification_type=NotificationService.TASK_ASSIGNED,
                            related_entity_type='task',
                            related_entity_id=task.id,
                            send_email=send_email,
                            email_context={'task': task, 'project': project_obj, 'assigned_by': current_user, 'task_url': NotificationService._build_task_url(task)}
                        )
                        email_sent = True
                    except Exception:
                        current_app.logger.exception('Failed to notify new assignee %s for task %s', uid, task.id)

            # Notificar si se asignó a un nuevo cliente (solo si fue provisto en el formulario)
            if assigned_client_provided and new_assigned_client_id is not None and new_assigned_client_id != old_assigned_client_id and new_assigned_client_id != current_user.id:
                try:
                    NotificationService.notify_task_assigned(
                        task=task,
                        assigned_by_user=current_user,
                        send_email=send_email,
                        notify_client=True
                    )
                    email_sent = True
                except Exception:
                    current_app.logger.exception('Failed to notify assigned client %s for task %s', new_assigned_client_id, task.id)

            if email_sent and send_email:
                flash('Se ha enviado una notificación por correo al usuario asignado.', 'info')            
            
            # Asegurar que las notificaciones generadas se reflejen en la BD inmediatamente
            db.session.commit()
            
            flash('Tarea actualizada.', 'success')
            return redirect(url_for('main.task_detail', task_id=task.id))
        except Exception as e:
            db.session.rollback()
            current_app.logger.exception('Error updating task %s: %s', task.id if task else None, e)
            flash(f'Error al actualizar: {str(e)}', 'danger')

    # Usuarios asignables: miembros del proyecto + manager + cualquier usuario que ya tenga
    # tareas asignadas en este proyecto (datos previos a auto-membresía).
    # Además, sincronizamos: si un usuario tiene tareas asignadas pero no es miembro, lo agregamos.
    task_assignee_ids = db.session.query(Task.assigned_to_id).filter(
        Task.project_id == project.id,
        Task.assigned_to_id.isnot(None)
    ).distinct().all()
    task_assignee_ids = {row[0] for row in task_assignee_ids}

    # Multi-assignees también
    from sqlalchemy import text as _text
    multi_rows = db.session.execute(
        _text('SELECT user_id FROM task_assignees ta JOIN tasks t ON ta.task_id = t.id WHERE t.project_id = :pid'),
        {'pid': project.id}
    ).fetchall()
    task_assignee_ids |= {row[0] for row in multi_rows}

    # Backfill: agregar como miembros del proyecto a quienes tienen tareas asignadas
    existing_member_ids = {m.id for m in project.members}
    new_members_added = False
    if task_assignee_ids - existing_member_ids:
        extra_users = User.query.filter(
            User.id.in_(task_assignee_ids - existing_member_ids),
            User.is_internal == True,
            User.is_active == True
        ).all()
        for u in extra_users:
            project.members.append(u)
            new_members_added = True
        if new_members_added:
            db.session.commit()

    # Candidate predecessors: tasks within the same project (exclude self)
    candidate_predecessors = Task.query.filter(Task.project_id == project.id, Task.id != task.id).order_by(Task.title).all()

    # Role flags for template: limit edit capabilities for Participants and Clients
    is_pmp_admin = (current_user.is_authenticated and current_user.is_internal and current_user.role and current_user.role.name in ('PMP', 'Admin', 'Supervisor'))
    is_participant = (current_user.is_authenticated and current_user.is_internal and current_user.role and current_user.role.name == 'Participante')
    is_client = (not current_user.is_internal) and (project.client_id == current_user.id)
    is_limited_editor = is_participant or is_client

    # Selector de asignados: miembros del proyecto (internos activos), agrupados por rol.
    # También incluir al manager del proyecto si no está ya en members.
    TASK_ROLE_GROUPS = ['PMP', 'Supervisor', 'Participante']
    users_by_role = {r: [] for r in TASK_ROLE_GROUPS}
    users_other = []
    seen_ids = set()
    project_team = list(project.members)
    if project.manager and project.manager.is_internal and project.manager.is_active:
        if project.manager.id not in {m.id for m in project_team}:
            project_team.append(project.manager)
    project_team.sort(key=lambda u: (u.first_name or u.email))
    for u in project_team:
        if u.id in seen_ids or not u.is_internal or not u.is_active:
            continue
        seen_ids.add(u.id)
        rn = u.role.name if u.role else None
        if rn in users_by_role:
            users_by_role[rn].append(u)
        else:
            users_other.append(u)
    users = project_team  # lista plana para compatibilidad

    # Clientes disponibles para asignar a la tarea: del proyecto, si no hay, todos los clientes del sistema
    _client_role = Role.query.filter_by(name='Cliente').first()
    if project.clients:
        available_task_clients = project.clients
    elif _client_role:
        available_task_clients = User.query.filter_by(role_id=_client_role.id, is_active=True).order_by(User.first_name).all()
    else:
        available_task_clients = []

    return render_template('task_edit.html', task=task, project=project,
                           users=users, users_by_role=users_by_role, users_other=users_other,
                           available_task_clients=available_task_clients,
                           candidate_predecessors=candidate_predecessors,
                           is_pmp_admin=is_pmp_admin, is_limited_editor=is_limited_editor,
                           allowed_extensions=list(current_app.config.get('ALLOWED_EXTENSIONS', [])))


@main_bp.route('/task/<int:task_id>/client_accept', methods=['POST'])
@login_required
def client_accept_task(task_id):
    """Allow a project client to accept a task and mark it as completed (with approval metadata)."""
    task = Task.query.get_or_404(task_id)
    project = task.project
    
    # Only allow project client to accept
    is_project_client = (project.client_id and project.client_id == current_user.id) or (current_user in project.clients)
    if current_user.is_internal or not is_project_client:
        flash('No tienes permiso para realizar esta acción.', 'danger')
        return redirect(url_for('main.task_detail', task_id=task.id))

    # Additionally, client can only accept tasks that are explicitly assigned to them
    if task.assigned_client_id != current_user.id:
        flash('No tienes permiso para realizar esta acción.', 'danger')
        return redirect(url_for('main.task_detail', task_id=task.id))

    # Prevent completing if predecessors or descendants incomplete
    blockers = task.get_completion_blockers()
    if blockers['incomplete_predecessors']:
        flash('No se puede completar la tarea mientras existan predecesoras incompletas', 'danger')
        return redirect(url_for('main.task_detail', task_id=task.id))
    if blockers['incomplete_children']:
        flash('No se puede completar la tarea mientras existan subtareas incompletas', 'danger')
        return redirect(url_for('main.task_detail', task_id=task.id))

    try:
        old_status = task.status
        task.status = 'COMPLETED'
        task.approval_status = 'APPROVED'
        task.approved_by_id = current_user.id
        task.completed_at = datetime.now()
        task.approved_at = datetime.now()

        audit = AuditLog(
            entity_type='Task',
            entity_id=task.id,
            action='UPDATE',
            user_id=current_user.id,
            changes={'status': {'old': old_status, 'new': task.status}, 'approval_status': {'old': None, 'new': 'APPROVED'}}
        )
        db.session.add(audit)
        db.session.commit()

        try:
            from app.services import webhook_service
            webhook_service.dispatch('task.completed', {
                'task_id': task.id,
                'task_title': task.title,
                'project_id': task.project_id,
                'project_name': task.project.name if task.project else None,
                'user_name': current_user.name if current_user else None,
                'old_status': old_status,
                'new_status': 'COMPLETED',
            })
        except Exception:
            current_app.logger.exception('Error dispatching webhook for task %s', task.id)

        # Notificar al asignado y al PMP que la tarea fue aprobada por el cliente
        try:
            from app.services.notifications import NotificationService
            from app.models import SystemSettings
            _send = SystemSettings.get('notify_task_approved', True)
            _send_email = _send.lower() in ('true', '1', 'yes') if isinstance(_send, str) else bool(_send)
            NotificationService.notify_task_approved(task=task, approved_by_user=current_user, send_email=_send_email)
            # También notificar al PMP principal y adicionales del proyecto
            _recipients = set()
            if task.project and task.project.manager_id:
                _recipients.add(task.project.manager_id)
            if task.project and getattr(task.project, 'members', None):
                for _m in task.project.members:
                    if getattr(_m, 'role', None) and _m.role.name in ('PMP', 'Supervisor'):
                        _recipients.add(_m.id)
            _recipients.discard(current_user.id)
            _recipients.discard(task.assigned_to_id)  # ya notificado por notify_task_approved
            _task_url = NotificationService._build_task_url(task)
            for _uid in _recipients:
                try:
                    NotificationService.notify(
                        user_id=_uid,
                        title='Tarea aprobada por cliente',
                        message=f"La tarea '{task.title}' fue aprobada por {current_user.name or current_user.email}",
                        notification_type=NotificationService.TASK_APPROVED,
                        related_entity_type='task',
                        related_entity_id=task.id,
                        send_email=_send_email,
                        email_context={'task': task, 'approved_by': current_user, 'task_url': _task_url,
                                       'message': f"La tarea '{task.title}' fue aprobada por {current_user.name or current_user.email}",
                                       'title': 'Tarea aprobada por cliente'},
                    )
                except Exception:
                    current_app.logger.exception('Failed to notify uid %s of task approval', _uid)
        except Exception:
            current_app.logger.exception('Error sending approval notifications for task %s', task.id)

        flash('Tarea aceptada y marcada como completada.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('main.pending_approvals'))


@main_bp.route('/task/<int:task_id>/reject', methods=['POST'])
@login_required
def reject_task(task_id):
    """Cliente rechaza una tarea completada"""
    task = Task.query.get_or_404(task_id)
    project = task.project
    
    # Verificar que el usuario es cliente del proyecto
    if current_user not in project.clients:
        flash('No tienes permiso para rechazar esta tarea.', 'danger')
        return redirect(url_for('main.index'))
        return redirect(url_for('main.dashboard'))
    
    notes = request.form.get('notes', '')
    
    if not notes:
        flash('Debes indicar el motivo del rechazo.', 'warning')
        return redirect(url_for('main.pending_approvals'))
    
    try:
        task.approval_status = 'REJECTED'
        task.approved_by_id = current_user.id
        task.approved_at = datetime.now()
        task.approval_notes = notes
        
        # Volver a IN_REVIEW para que el equipo revise
        task.status = 'IN_REVIEW'
        
        # Auditoría de rechazo
        audit = AuditLog(
            entity_type='Task',
            entity_id=task.id,
            action='UPDATE',
            user_id=current_user.id,
            changes={'approval_status': {'old': 'PENDING', 'new': 'REJECTED'}, 'status': {'old': 'COMPLETED', 'new': 'IN_REVIEW'}, 'rejection_reason': notes}
        )
        db.session.add(audit)
        
        # Marcar notificaciones del cliente como leídas
        SystemNotification.query.filter_by(
            user_id=current_user.id,
            related_entity_type='Task',
            related_entity_id=task.id
        ).update({'is_read': True})
        
        db.session.commit()

        try:
            from app.services import webhook_service
            webhook_service.dispatch('task.status_changed', {
                'task_id': task.id,
                'task_title': task.title,
                'project_id': task.project_id,
                'project_name': project.name if project else None,
                'user_name': current_user.name if current_user else None,
                'old_status': 'COMPLETED',
                'new_status': 'IN_REVIEW',
            })
        except Exception:
            current_app.logger.exception('Error dispatching webhook for task %s', task.id)

        # Notificar al responsable de la tarea que fue rechazada (con email)
        NotificationService.notify_task_rejected(
            task=task,
            rejected_by_user=current_user,
            rejection_reason=notes,
            send_email=SystemSettings.get('notify_task_rejected', True)
        )
        
        # Asegurar que la notificación de rechazo se refleje en la BD inmediatamente
        db.session.commit()
        
        flash(f'Tarea "{task.title}" rechazada. El equipo será notificado.', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al rechazar: {str(e)}', 'danger')
    
    return redirect(url_for('main.pending_approvals'))


# ========== AUDIT LOG ROUTES ==========

def pmp_or_admin_required(f):
    """Decorator to require PMP or Admin role"""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        if not current_user.role or current_user.role.name not in ['PMP', 'Admin']:
            flash('Solo usuarios PMP o Admin pueden acceder a esta sección.', 'danger')
            return redirect(url_for('main.index'))
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated


@main_bp.route('/audit')
@login_required
@pmp_or_admin_required
def audit_log():
    """Vista del registro de auditoría - Solo PMP y Admin"""
    # Limpiar registros antiguos (más de 6 meses)
    cleanup_old_audit_logs()
    
    # Filtros
    # Canonical filters
    entity_type = request.args.get('entity_type', '')
    entity_id = request.args.get('entity_id', type=int)

    # Backward-compatible aliases from older links
    if not entity_type:
        entity_type = request.args.get('entity', '')
    if entity_id is None:
        entity_id = request.args.get('id', type=int)

    action = request.args.get('action', '')
    user_id = request.args.get('user_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    page = request.args.get('page', 1, type=int)
    per_page = 15
    
    # Construir query
    query = AuditLog.query
    
    if entity_type:
        query = query.filter(AuditLog.entity_type == entity_type)
    if entity_id:
        query = query.filter(AuditLog.entity_id == entity_id)
    if action:
        query = query.filter(AuditLog.action == action)
    if user_id:
        query = query.filter(AuditLog.user_id == user_id)
    if date_from:
        try:
            date_from_dt = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(AuditLog.created_at >= date_from_dt)
        except ValueError:
            pass
    if date_to:
        try:
            date_to_dt = datetime.strptime(date_to, '%Y-%m-%d').date() + timedelta(days=1)
            query = query.filter(AuditLog.created_at < date_to_dt)
        except ValueError:
            pass
    
    # Ordenar y paginar
    query = query.order_by(AuditLog.created_at.desc())
    total = query.count()
    logs = query.offset((page - 1) * per_page).limit(per_page).all()
    total_pages = (total + per_page - 1) // per_page
    
    # Obtener registro más antiguo
    oldest = AuditLog.query.order_by(AuditLog.created_at.asc()).first()
    oldest_record = oldest.created_at if oldest else None
    
    # Obtener usuarios para filtro
    users = User.query.filter(User.is_internal == True).order_by(User.first_name).all()
    
    filters = {
        'entity_type': entity_type,
        'entity_id': entity_id,
        'action': action,
        'user_id': user_id,
        'date_from': date_from,
        'date_to': date_to
    }
    
    return render_template('audit_log.html',
        logs=logs,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
        filters=filters,
        users=users,
        oldest_record=oldest_record
    )


def cleanup_old_audit_logs():
    """Elimina registros de auditoría con más de 6 meses de antigüedad"""
    try:
        cutoff_date = datetime.now() - timedelta(days=180)  # 6 meses
        deleted_count = AuditLog.query.filter(AuditLog.created_at < cutoff_date).delete()
        if deleted_count > 0:
            db.session.commit()
            current_app.logger.info(f'Limpieza de auditoría: {deleted_count} registros eliminados')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error limpiando auditoría: {e}')


# ========== ADMIN SETTINGS ROUTES ==========

def admin_required(f):
    """Decorator to require Admin role only for settings"""
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        if not current_user.role or current_user.role.name != 'Admin':
            flash('Solo los administradores pueden acceder a esta sección.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated


@main_bp.route('/admin/settings', methods=['GET'])
@login_required
@pmp_or_admin_required
def admin_settings_page():
    """Admin settings page"""
    from app.models import SystemSettings, Role
    from app.services import license_service
    
    # Get all users
    users = User.query.order_by(User.created_at.desc()).all()
    
    # Get all roles
    roles = Role.query.all()
    
    # Get all settings as dict (raw DB values)
    all_settings = SystemSettings.query.all()
    settings = {s.key: s.value for s in all_settings}
    # ensure logo_url and favicon_url entries for templates; ignore if file missing
    logo = SystemSettings.get('logo_path') or ''
    if logo:
        # resolve disk path similar to context processor
        rel = logo.lstrip('/')
        parts = rel.split('/')
        if parts and parts[0] == 'uploads':
            parts = parts[1:]
        fs = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), *parts)
        if not os.path.exists(fs):
            logo = ''
    settings.setdefault('logo_url', logo)
    favicon = SystemSettings.get('favicon_path') or ''
    settings.setdefault('favicon_url', favicon)
    
    # Ensure sensible defaults: notifications ON by default
    defaults = {
        'portfolio_enabled':        'true',
        'budget_tracking_enabled': 'true',
        'risks_enabled':           'true',
        'budget_warn_threshold':   '80',
        'budget_danger_threshold': '100',
        'budget_currency':         'USD',
        'notify_task_assigned': 'true',
        'notify_task_completed': 'true',
        'notify_task_status_changed': 'true',
        'notify_task_approved': 'true',
        'notify_task_rejected': 'true',
        'notify_task_comment': 'true',
        'notify_task_mention': 'true',
        'notify_due_date_reminder': 'true',
        'show_notification_center': 'true',
        'enable_azure_auth': 'true',
        'enable_local_auth': 'true'
    }
    # Default for global alert
    defaults.setdefault('global_alert_enabled', 'false')
    defaults.setdefault('global_alert_message', '')
    for k, v in defaults.items():
        settings.setdefault(k, v)

    # Normalize boolean-like options so templates can rely on Python booleans
    notify_keys = [
        'notify_task_assigned', 'notify_task_completed', 'notify_task_status_changed',
        'notify_task_approved', 'notify_task_rejected', 'notify_task_comment',
        'notify_task_mention', 'notify_due_date_reminder',
        'show_notification_center', 'enable_push_notifications',
        'enable_azure_auth', 'enable_local_auth',
        'portfolio_enabled',
        'budget_tracking_enabled',
        'risks_enabled',
        'departments_enabled',
    ]
    # Treat global alert enabled as boolean for templates
    notify_keys.append('global_alert_enabled')
    for k in notify_keys:
        raw = settings.get(k, defaults.get(k, 'true'))
        if isinstance(raw, str):
            settings[k] = raw.lower() not in ('false', '0', 'no')
        else:
            settings[k] = bool(raw)
    
    # Get system stats
    stats = {
        'total_users': User.query.count(),
        'total_projects': Project.query.count(),
        'total_tasks': Task.query.count(),
        'active_users': User.query.filter_by(is_active=True).count()
    }
    
    # Get license info
    license_info = license_service.check_license_status()
    hardware_id = license_service.get_hardware_id()
    
    from ..models import Department
    departments = Department.query.order_by(Department.name).all()
    return render_template('admin_settings.html',
                         users=users,
                         roles=roles,
                         settings=settings,
                         stats=stats,
                         license_info=license_info,
                         hardware_id=hardware_id,
                         departments=departments)


@main_bp.route('/admin/settings', methods=['POST'])
@login_required
@pmp_or_admin_required
def admin_settings():
    """Handle admin settings form submission"""
    from app.models import SystemSettings
    
    section = request.form.get('section', 'general')
    
    # always handle logo/favicon uploads if present, regardless of section
    if 'logo' in request.files:
        file = request.files['logo']
        if file and file.filename:
            from werkzeug.utils import secure_filename
            filename = secure_filename(file.filename)
            logo_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), 'branding')
            os.makedirs(logo_path, exist_ok=True)
            filepath = os.path.join(logo_path, f'logo_{filename}')
            file.save(filepath)
            current_app.logger.info('Saved branding logo to %s', filepath)
            SystemSettings.set('logo_path', f'/uploads/branding/logo_{filename}', 'branding', 'Ruta del logo', user_id=current_user.id)
    if 'favicon' in request.files:
        file = request.files['favicon']
        if file and file.filename:
            from werkzeug.utils import secure_filename
            filename = secure_filename(file.filename)
            logo_path = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), 'branding')
            os.makedirs(logo_path, exist_ok=True)
            filepath = os.path.join(logo_path, f'favicon_{filename}')
            file.save(filepath)
            current_app.logger.info('Saved branding favicon to %s', filepath)
            SystemSettings.set('favicon_path', f'/uploads/branding/favicon_{filename}', 'branding', 'Ruta del favicon', user_id=current_user.id)
    
    # Get all form fields (except section and csrf)
    fields_to_save = {k: v for k, v in request.form.items() 
                     if k not in ('section', 'csrf_token')}
    
    # Handle checkboxes (they only submit when checked)
    # Fix: Only process checkboxes belonging to the current section to avoid disabling others
    section_checkboxes = {
        'notifications': [
            'smtp_use_tls', 'notify_task_assigned', 'notify_task_completed',
            'notify_task_status_changed', 'notify_task_approved', 'notify_task_rejected',
            'notify_task_comment', 'notify_task_mention', 'notify_due_date_reminder',
            'show_notification_center', 'enable_push_notifications'
        ],
        'authentication': [
            'enable_azure_auth', 'enable_local_auth',
            'allow_public_registration', 'password_require_complexity'
        ],
        'general': [
            'allow_projects_without_manager', 'require_task_estimation',
            'block_parent_until_children_complete'
        ],
        'projects': [
            'portfolio_enabled',
            'budget_tracking_enabled',
            'risks_enabled',
            'departments_enabled',
        ],
        'global_alert': [
            'global_alert_enabled'
        ]
    }
    
    # Get checkboxes for current section (support 'security' as alias for 'authentication')
    target_section = 'authentication' if section == 'security' else section
    checkboxes_to_process = section_checkboxes.get(target_section, [])

    # Flatten all known checkbox fields so we can detect boolean settings when saving
    checkbox_fields = set()
    for lst in section_checkboxes.values():
        checkbox_fields.update(lst)
    # Include global alert flag as a checkbox field
    checkbox_fields.add('global_alert_enabled')

    for cb in checkboxes_to_process:
        if cb not in fields_to_save:
            fields_to_save[cb] = 'false'
        else:
            fields_to_save[cb] = 'true'
    
    # Handle password field - don't update if empty
    if 'smtp_password' in fields_to_save and not fields_to_save['smtp_password']:
        del fields_to_save['smtp_password']
    
    # Save each setting; for checkbox fields, save as boolean value_type
    for key, value in fields_to_save.items():
        if value is not None and value != '':
            value_type = 'boolean' if key in checkbox_fields else 'string'
            SystemSettings.set(
                key=key,
                value=str(value),
                category=section,
                description=None,
                value_type=value_type,
                user_id=current_user.id
            )
    
    db.session.commit()
    flash('Configuración guardada correctamente.', 'success')
    return redirect(url_for('main.admin_settings_page') + f'#{section}')


@main_bp.route('/admin/settings/user/create', methods=['POST'])
@login_required
@admin_required
def admin_settings_create_user():
    """Create a new user from settings page"""
    from werkzeug.security import generate_password_hash
    from app.models import Role
    
    email = request.form.get('email')
    password = request.form.get('password')
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    role_id = request.form.get('role_id')
    is_internal = request.form.get('is_internal') == '1'
    
    # Check if email already exists
    if User.query.filter_by(email=email).first():
        flash('Ya existe un usuario con ese email.', 'danger')
        return redirect(url_for('main.admin_settings_page') + '#users')
    
    # Create user
    user = User(
        email=email,
        password_hash=generate_password_hash(password),
        first_name=first_name,
        last_name=last_name,
        role_id=int(role_id) if role_id else None,
        is_internal=is_internal,
        is_active=True
    )
    db.session.add(user)
    db.session.commit()
    
    flash(f'Usuario {email} creado correctamente.', 'success')
    return redirect(url_for('main.admin_settings_page') + '#users')


@main_bp.route('/admin/settings/user/update', methods=['POST'])
@login_required
@admin_required  
def admin_settings_update_user():
    """Update an existing user from settings page"""
    from werkzeug.security import generate_password_hash
    
    user_id = request.form.get('user_id')
    user = User.query.get_or_404(user_id)
    
    # Azure Entra ID users can only have their role updated
    is_azure_user = bool(user.azure_oid)
    
    if is_azure_user:
        # Only allow role update for Azure users
        user.role_id = int(request.form.get('role_id')) if request.form.get('role_id') else user.role_id
    else:
        # Local users - all fields can be updated
        user.email = request.form.get('email', user.email)
        user.first_name = request.form.get('first_name', user.first_name)
        user.last_name = request.form.get('last_name', user.last_name)
        user.role_id = int(request.form.get('role_id')) if request.form.get('role_id') else user.role_id
        user.is_internal = request.form.get('is_internal') == '1'
        
        # Update password only if provided
        password = request.form.get('password')
        if password:
            user.password_hash = generate_password_hash(password)
    
    db.session.commit()
    flash(f'Usuario {user.email} actualizado.', 'success')
    return redirect(url_for('main.admin_settings_page') + '#users')


@main_bp.route('/api/users/<int:user_id>/toggle-status', methods=['POST'])
@login_required
@admin_required
def toggle_user_status(user_id):
    """Toggle user active status"""
    user = User.query.get_or_404(user_id)
    
    if user.id == current_user.id:
        return jsonify({'success': False, 'error': 'No puedes desactivar tu propia cuenta'}), 400
    
    user.is_active = not user.is_active
    db.session.commit()
    
    return jsonify({'success': True, 'is_active': user.is_active})


@main_bp.route('/admin/test-email', methods=['POST'])
@login_required
@admin_required
def admin_test_email():
    """Test SMTP email configuration"""
    from app.models import SystemSettings
    import smtplib
    from email.mime.text import MIMEText

    try:
        # Get email settings
        host = SystemSettings.get('smtp_host', '')
        port = int(SystemSettings.get('smtp_port', '587'))
        username = SystemSettings.get('smtp_username', '')
        password = SystemSettings.get('smtp_password', '')
        use_tls = SystemSettings.get('smtp_use_tls', 'true') == 'true'
        from_email = SystemSettings.get('email_from', username)
        from_name = SystemSettings.get('email_from_name', 'BridgeWork')

        if not host or not username:
            return jsonify({'success': False, 'error': 'Configura el servidor SMTP primero'}), 400

        # Create test message
        msg = MIMEText(f'Este es un correo de prueba desde {from_name}.\n\nSi recibes este mensaje, la configuración SMTP es correcta.')
        msg['Subject'] = f'[{from_name}] Prueba de conexión SMTP'
        msg['From'] = f'{from_name} <{from_email}>'
        msg['To'] = current_user.email

        # Connect and send
        if use_tls and port == 587:
            server = smtplib.SMTP(host, port)
            server.starttls()
        elif port == 465:
            server = smtplib.SMTP_SSL(host, port)
        else:
            server = smtplib.SMTP(host, port)

        if username and password:
            server.login(username, password)

        server.sendmail(from_email, [current_user.email], msg.as_string())
        server.quit()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/admin/run-due-reminders', methods=['POST'])
@login_required
@admin_required
def admin_run_due_reminders():
    """Trigger generation of due date reminders manually (admin-only)."""
    from app.tasks.alerts import generate_alerts

    try:
        res = generate_alerts()
        created = res.get('created', [])
        groups = res.get('groups', {})
        # Ensure keys are JSON-serializable (strings)
        simple_groups = {str(k): len(v) for k, v in groups.items()}
        return jsonify({'success': True, 'created': len(created), 'groups': simple_groups})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/admin/enable-all-notifications', methods=['POST'])
@login_required
@admin_required
def admin_enable_all_notifications():
    """Enable all email notification toggles (admin-only convenience endpoint)."""
    from app.models import SystemSettings, AuditLog

    keys = [
        'notify_task_assigned', 'notify_task_completed', 'notify_task_approved',
        'notify_task_rejected', 'notify_task_comment', 'notify_due_date_reminder',
        'show_notification_center', 'enable_push_notifications'
    ]
    try:
        for k in keys:
            SystemSettings.set(k, 'true', category='notifications', value_type='boolean', user_id=current_user.id)
        # Audit log
        audit = AuditLog(
            entity_type='SystemSettings',
            entity_id=0,
            action='UPDATE',
            user_id=current_user.id,
            changes={'enabled_all_notifications': True}
        )
        db.session.add(audit)
        db.session.commit()
        return jsonify({'success': True, 'enabled': keys})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/admin/send-test-notification', methods=['POST'])
@login_required
@admin_required
def admin_send_test_notification():
    """Create a test notification for the current admin user and attempt send (admin-only)."""
    from app.services.notifications import NotificationService

    try:
        # Create an in-app notification
        note = NotificationService.create(
            user_id=current_user.id,
            title='Prueba de Notificación',
            message='Este es un mensaje de prueba generado desde la configuración de administrador',
            notification_type=NotificationService.GENERAL
        )
        # Attempt to send email for this notification
        sent = NotificationService.send_email(user_id=current_user.id, subject=note.title, notification_type=note.notification_type, context={'message': note.message, 'title': note.title})
        return jsonify({'success': True, 'notification_id': note.id, 'email_sent': bool(sent)})
    except Exception as e:
        current_app.logger.exception('Error sending test notification: %s', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/search')
@login_required
def global_search():
    """Global search endpoint for projects, tasks, users and comments."""
    from ..models import TaskComment
    from sqlalchemy import or_, func

    query = request.args.get('q', '').strip()
    type_filter = request.args.get('type', '')   # projects|tasks|users|comments|clients
    LIMIT = 8

    if not query or len(query) < 2:
        return jsonify({'projects': [], 'tasks': [], 'users': [], 'comments': [], 'clients': [], 'query': query})

    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None
    if not user_role:
        return jsonify({'error': 'No tienes permisos para buscar'}), 403
    is_pmp_admin = user_role in ('PMP', 'Admin')

    # ── Projects ──────────────────────────────────────────────────────────────
    projects = []
    if not type_filter or type_filter == 'projects':
        pq = Project.query.filter(
            or_(Project.name.ilike(f'%{query}%'), Project.description.ilike(f'%{query}%'))
        )
        if not is_pmp_admin:
            if user_role == 'Participante':
                pq = pq.filter(Project.members.contains(current_user))
            elif not current_user.is_internal:
                pq = pq.filter(Project.clients.contains(current_user))
            else:
                pq = pq.filter(Project.id == -1)
        projects = pq.limit(LIMIT).all()

    # ── Tasks ─────────────────────────────────────────────────────────────────
    tasks = []
    if not type_filter or type_filter == 'tasks':
        tq = Task.query.filter(
            or_(Task.title.ilike(f'%{query}%'), Task.description.ilike(f'%{query}%'))
        )
        if not is_pmp_admin:
            if user_role == 'Participante':
                tq = tq.filter(
                    or_(Task.assigned_to_id == current_user.id,
                        Task.assignees.any(User.id == current_user.id))
                )
            elif not current_user.is_internal:
                # Client sees tasks in their linked projects (external visible) OR directly assigned
                client_project_ids = [p.id for p in Project.query.filter(Project.clients.contains(current_user)).all()]
                tq = tq.filter(
                    or_(Task.assigned_client_id == current_user.id,
                        Task.project_id.in_(client_project_ids) if client_project_ids else False)
                )
            else:
                tq = tq.filter(Task.id == -1)
        tasks = tq.limit(LIMIT).all()

    # ── Comments ──────────────────────────────────────────────────────────────
    comments = []
    if not type_filter or type_filter == 'comments':
        cq = TaskComment.query.filter(TaskComment.body.ilike(f'%{query}%'))
        if not is_pmp_admin:
            if user_role == 'Participante':
                cq = cq.join(Task).filter(
                    or_(Task.assigned_to_id == current_user.id,
                        Task.assignees.any(User.id == current_user.id))
                )
            elif not current_user.is_internal:
                cq = cq.join(Task).filter(Task.assigned_client_id == current_user.id)
            else:
                cq = cq.filter(TaskComment.id == -1)
        comments = cq.limit(LIMIT).all()

    # ── Users (internal) ─────────────────────────────────────────────────────
    users = []
    if current_user.is_internal and (not type_filter or type_filter == 'users'):
        full_name = func.concat(User.first_name, ' ', User.last_name)
        users = User.query.filter(
            User.is_internal == True,
            or_(User.first_name.ilike(f'%{query}%'),
                User.last_name.ilike(f'%{query}%'),
                full_name.ilike(f'%{query}%'),
                User.email.ilike(f'%{query}%'))
        ).limit(LIMIT).all()

    # ── Clients ───────────────────────────────────────────────────────────────
    clients_res = []
    if is_pmp_admin and (not type_filter or type_filter == 'clients'):
        full_name_c = func.concat(User.first_name, ' ', User.last_name)
        clients_res = User.query.filter(
            User.is_internal == False,
            or_(User.first_name.ilike(f'%{query}%'),
                User.last_name.ilike(f'%{query}%'),
                full_name_c.ilike(f'%{query}%'),
                User.email.ilike(f'%{query}%'),
                User.company.ilike(f'%{query}%'))
        ).limit(LIMIT).all()

    return jsonify({
        'query': query,
        'projects': [{
            'name':        p.name,
            'status':      p.status,
            'description': (p.description or '')[:80],
            'manager':     p.manager.name if p.manager else '',
            'url':         url_for('main.project_detail', project_id=p.id),
        } for p in projects],
        'tasks': [{
            'title':        t.title,
            'status':       t.status,
            'priority':     t.priority,
            'project_name': t.project.name if t.project else '',
            'description':  (t.description or '')[:80],
            'due_date':     t.due_date.strftime('%d/%m/%Y') if t.due_date else None,
            'is_overdue':   bool(t.due_date and (t.due_date.date() if hasattr(t.due_date, 'date') else t.due_date) < datetime.now().date() and t.status not in ('COMPLETED', 'DONE', 'ACCEPTED')),
            'assignee':     t.assignees[0].name if t.assignees else (t.assigned_to.name if t.assigned_to else ''),
            'url':          url_for('main.task_detail', task_id=t.id),
        } for t in tasks],
        'comments': [{
            'content':      (c.body or '')[:100],
            'task_title':   c.task.title if c.task else '',
            'author':       c.user.name if c.user else '',
            'url':          url_for('main.task_detail', task_id=c.task_id) + '#comments',
        } for c in comments],
        'users': [{
            'name':        f"{u.first_name or ''} {u.last_name or ''}".strip() or u.email,
            'email':       u.email,
            'role':        u.role.name if u.role else '',
            'is_internal': u.is_internal,
            'url':         url_for('main.team'),
        } for u in users],
        'clients': [{
            'name':    f"{u.first_name or ''} {u.last_name or ''}".strip() or u.email,
            'email':   u.email,
            'company': u.company or '',
            'url':     url_for('main.clients'),
        } for u in clients_res],
    })


@main_bp.route('/notifications/recent')
@login_required
def recent_notifications():
    """Get recent notifications for the current user (JSON API)"""
    from .. import db
    try:
        unread_count = SystemNotification.query.filter_by(
            user_id=current_user.id,
            is_read=False
        ).count()
        
        notifications = SystemNotification.query.filter_by(
            user_id=current_user.id
        ).order_by(SystemNotification.created_at.desc()).limit(10).all()
        
        notif_list = [{
            'id': n.id,
            'title': n.title,
            'message': n.message,
            'type': n.notification_type,
            'created_at': n.created_at.isoformat() if n.created_at else None,
            'is_read': n.is_read,
            'related_entity_type': n.related_entity_type,
            'related_entity_id': n.related_entity_id
        } for n in notifications]
            
        return jsonify({
            'unread_count': unread_count,
            'notifications': notif_list
        })
    except Exception as e:
        current_app.logger.exception('Error fetching recent notifications: %s', e)
        # If the DB session is in a broken state, rollback to clear it
        try:
            db.session.rollback()
        except Exception:
            current_app.logger.exception('Error rolling back DB session')
        return jsonify({'error': 'Error fetching notifications'}), 500


# Missing notifications route
@main_bp.route('/notifications')
@login_required
def notifications():
    """User notifications center - displays in-app system notifications"""
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    
    # Get unread notifications for current user
    unread = SystemNotification.query.filter_by(
        user_id=current_user.id,
        is_read=False
    ).order_by(SystemNotification.created_at.desc()).all()
    
    # Get all notifications (paginated)
    page = request.args.get('page', 1, type=int)
    per_page = 20
    all_notifications = SystemNotification.query.filter_by(
        user_id=current_user.id
    ).order_by(SystemNotification.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Count stats
    unread_count = len(unread)
    
    return render_template('notifications.html', 
                          notifications=all_notifications,
                          unread_count=unread_count,
                          now=datetime.now())


@main_bp.route('/notifications/mark-all-read', methods=['POST'])
@login_required
def mark_all_notifications_read():
    SystemNotification.query.filter_by(user_id=current_user.id, is_read=False).update({'is_read': True})
    db.session.commit()
    return jsonify({'success': True})


@main_bp.route('/notification/<int:notification_id>/read', methods=['POST'])
@login_required
def mark_notification_read(notification_id):
    notification = SystemNotification.query.get_or_404(notification_id)
    if notification.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403
    notification.is_read = True
    db.session.commit()
    return jsonify({'success': True})


@main_bp.route('/pending-approvals')
@login_required
def pending_approvals():
    """Show pending task approvals for project clients"""
    from ..auth.decorators import _get_user_from_session
    user = _get_user_from_session()
    user_role = user.role.name if (user and user.role) else None
    
    # Only clients can view pending approvals (or show empty for others)
    if current_user.is_internal:
        pending_tasks = []
    else:
        # Get tasks pending client approval that are assigned to this client
        pending_tasks = Task.query.filter(
            Task.assigned_client_id == current_user.id,
            Task.requires_approval == True,
            Task.approval_status == 'PENDING',
            Task.status == 'COMPLETED'
        ).order_by(Task.due_date.desc().nullslast()).all()
    
    return render_template('pending_approvals.html',
                          tasks=pending_tasks,
                          now=datetime.now())